import os
import re
import csv
import argparse
import sys
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from shared.locale_parser import extract_locale_from_filename

def normalize_locale(filename):
    return extract_locale_from_filename(filename)

def scan_apps(aso_demo_dir):
    # Any directory inside aso_demo_dir that contains an "Input" folder is an app
    apps = []
    for item in os.listdir(aso_demo_dir):
        item_path = os.path.join(aso_demo_dir, item)
        if os.path.isdir(item_path) and not item.startswith(".") and item != "Keyword_Tracker" and item != "Docs_and_Templates" and item != "Master_Keywords":
            input_dir = os.path.join(item_path, "Input")
            if os.path.isdir(input_dir):
                apps.append(item)
    return apps

def get_unique_keywords_from_csv(app_dir):
    input_dir = os.path.join(app_dir, "Input")
    if not os.path.isdir(input_dir):
        return {}
        
    keywords_by_locale = {} # locale -> set of lowercase keywords
    
    for month_folder in os.listdir(input_dir):
        month_path = os.path.join(input_dir, month_folder)
        if os.path.isdir(month_path) and re.match(r'^\d{6}$', month_folder):
            for file_name in os.listdir(month_path):
                if file_name.endswith(".csv"):
                    csv_path = os.path.join(month_path, file_name)
                    locale = normalize_locale(file_name)
                    if not locale:
                        continue
                        
                    if locale not in keywords_by_locale:
                        keywords_by_locale[locale] = set()
                        
                    try:
                        with open(csv_path, mode='r', encoding='utf-8-sig') as f:
                            reader = csv.DictReader(f)
                            fieldnames = reader.fieldnames
                            if not fieldnames:
                                continue
                                
                            # Find keyword column name (case insensitive/trimmed)
                            kw_col = None
                            for f_name in fieldnames:
                                if f_name.strip().lower() == "keyword":
                                    kw_col = f_name
                                    break
                                    
                            if not kw_col:
                                continue
                                
                            for row in reader:
                                kw = row.get(kw_col)
                                if kw and kw.strip():
                                    keywords_by_locale[locale].add(kw.strip())
                    except Exception as e:
                        print(f"  [Warning] Error reading CSV {file_name}: {e}")
                        
    return keywords_by_locale

def get_excluded_keywords_from_excel(app_dir):
    output_dir = os.path.join(app_dir, "Output")
    if not os.path.isdir(output_dir):
        return {}
        
    excluded_by_locale = {} # locale -> set of lowercase keywords
    
    for month_folder in os.listdir(output_dir):
        month_path = os.path.join(output_dir, month_folder)
        if os.path.isdir(month_path) and re.match(r'^\d{6}$', month_folder):
            for file_name in os.listdir(month_path):
                if file_name.endswith(".xlsx") and not file_name.startswith("~$") and "AllMarkets" not in file_name:
                    excel_path = os.path.join(month_path, file_name)
                    locale = normalize_locale(file_name)
                    if not locale:
                        continue
                        
                    if locale not in excluded_by_locale:
                        excluded_by_locale[locale] = set()
                        
                    wb = None
                    try:
                        wb = load_workbook(excel_path, read_only=True, data_only=True)
                        audit_sheet = next((name for name in wb.sheetnames if name.endswith("Dropped_Audit")), None)
                        if audit_sheet:
                            ws = wb[audit_sheet]
                            
                            # Read headers to find columns
                            headers = []
                            for row in ws.iter_rows(max_row=1, values_only=True):
                                headers = row
                                break
                                
                            if not headers:
                                continue
                                
                            kw_idx = -1
                            for idx, h in enumerate(headers):
                                if h:
                                    clean_h = str(h).strip().lower()
                                    if clean_h == "keyword":
                                        kw_idx = idx
                                        
                            if kw_idx == -1:
                                continue
                                
                            # Read data rows
                            row_iter = ws.iter_rows(min_row=2, values_only=True)
                            for row in row_iter:
                                if len(row) > kw_idx:
                                    kw = row[kw_idx]
                                    if kw:
                                        excluded_by_locale[locale].add(str(kw).strip().lower())
                    except Exception as e:
                        print(f"  [Warning] Error parsing Excel {file_name}: {e}")
                    finally:
                        if wb is not None:
                            wb.close()
                        
    return excluded_by_locale

def export_master_list(app_name, aso_demo_dir):
    print(f"Exporting Master Keywords for {app_name}...")
    app_dir = os.path.join(aso_demo_dir, app_name)
    
    # 1. Gather all keywords from CSV
    keywords_by_locale = get_unique_keywords_from_csv(app_dir)
    
    # 2. Gather excluded keywords from Excel output
    excluded_by_locale = get_excluded_keywords_from_excel(app_dir)
    
    if not keywords_by_locale:
        print(f"  [Error] No CSV keywords found for app {app_name}.")
        return
        
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Styling definitions
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
    regular_font = Font(name=font_family, size=11)
    
    thin_border_side = Side(border_style="thin", color="DADCE0")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    # Process each locale
    locales_processed = 0
    for locale in sorted(keywords_by_locale.keys()):
        raw_keywords = keywords_by_locale[locale]
        excluded = excluded_by_locale.get(locale, set())
        
        # Filter: Keep keywords whose lowercase version is NOT in excluded set
        filtered_keywords = []
        for kw in raw_keywords:
            if kw.lower() not in excluded:
                filtered_keywords.append(kw)
                
        # Sort keywords alphabetically
        filtered_keywords.sort()
        
        if not filtered_keywords:
            print(f"  [Info] Locale {locale} has 0 keywords after filtering. Skipping sheet.")
            continue
            
        # Create sheet for locale
        ws = wb.create_sheet(title=locale)
        ws.views.sheetView[0].showGridLines = True
        
        # Write headers
        ws.cell(row=1, column=1, value="#").font = title_font
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=1).alignment = align_center
        ws.cell(row=1, column=1).border = thin_border
        
        ws.cell(row=1, column=2, value="Keyword").font = title_font
        ws.cell(row=1, column=2).fill = header_fill
        ws.cell(row=1, column=2).alignment = align_left
        ws.cell(row=1, column=2).border = thin_border
        
        # Write data rows
        for r_idx, kw in enumerate(filtered_keywords, start=2):
            c_index = ws.cell(row=r_idx, column=1, value=r_idx - 1)
            c_index.font = regular_font
            c_index.alignment = align_center
            c_index.border = thin_border
            
            c_kw = ws.cell(row=r_idx, column=2, value=kw)
            c_kw.font = regular_font
            c_kw.alignment = align_left
            c_kw.border = thin_border
            
        # Set dimensions
        ws.column_dimensions['A'].width = 8
        
        # Find maximum length for keyword column to autofit
        max_len = max(len(k) for k in filtered_keywords)
        ws.column_dimensions['B'].width = max(max_len + 4, 25)
        
        print(f"  Locale {locale}: {len(filtered_keywords)} unique keywords (Filtered out {len(excluded)} irrelevant/noise terms).")
        locales_processed += 1
        
    if locales_processed == 0:
        print(f"  [Error] No sheets written for app {app_name}. Workbook empty.")
        return
        
    # Save file to output directory
    output_dir = os.path.join(aso_demo_dir, "Master_Keywords")
    os.makedirs(output_dir, exist_ok=True)
    
    output_file = os.path.join(output_dir, f"{app_name}_master.xlsx")
    wb.save(output_file)
    print(f"  [Success] Saved Master Keywords workbook to: {os.path.relpath(output_file, aso_demo_dir)}")

def main():
    parser = argparse.ArgumentParser(description="Export consolidated unique keywords from ASO pipeline CSV inputs, filtering out irrelevant and noise-only keywords classified in Excel outputs.")
    
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--app", type=str, help="Export master keywords for a specific app (e.g. AR_Filter)")
    group.add_argument("--all", action="store_true", help="Export master keywords for all detected apps")
    
    args = parser.parse_args()
    
    script_dir = os.path.dirname(os.path.abspath(__file__))
    aso_demo_dir = script_dir # Since script is running directly inside ASO-DEMO/
    
    detected_apps = scan_apps(aso_demo_dir)
    
    if args.app:
        if args.app not in detected_apps:
            print(f"Error: App '{args.app}' not detected in workspace. Available apps: {detected_apps}")
            sys.exit(1)
        export_master_list(args.app, aso_demo_dir)
    elif args.all:
        print(f"Detected {len(detected_apps)} apps: {detected_apps}")
        for app in detected_apps:
            export_master_list(app, aso_demo_dir)
            print("-" * 50)
            
if __name__ == "__main__":
    main()
