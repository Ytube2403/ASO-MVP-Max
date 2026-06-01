import pandas as pd
import numpy as np
import re
import os
import unicodedata
import json
import urllib.request
import urllib.parse
import html as html_lib
import ssl
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import argparse
import sys

_SHARED_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _SHARED_ROOT not in sys.path:
    sys.path.insert(0, _SHARED_ROOT)
from shared import text_dedup as _shared_text_dedup
from shared import profile_service as _shared_profile_service
from shared import translation_service as _shared_translation_service

# Parse arguments
parser = argparse.ArgumentParser(description="ASO Keyword Planner for Game Emulator")
parser.add_argument("--csv", type=str, default=None, help="Path to input CSV")
parser.add_argument("--market", type=str, default="US_EN", help="Market code (e.g. US_EN)")
parser.add_argument("--output", type=str, default="", help="Path to output Excel file")
parser.add_argument("--interactive", action="store_true", help="Run interactive Web UI selector")
args, unknown = parser.parse_known_args()

INPUT_PATH = args.csv
if args.output:
    OUTPUT_PATH = args.output
else:
    # Update OUTPUT_PATH dynamically
    csv_dir = os.path.dirname(os.path.abspath(INPUT_PATH))
    OUTPUT_PATH = os.path.join(csv_dir, "Game_Emulator", f"GameEmulator_{args.market.replace('_', '-')}_Output.xlsx")

# Game Emulator configuration
config = {
    "app_id": "com.game.emulator.gb4.retro.gameboy.collection",
    "app_name": "Game Emulator: GB4 Retro Games",
    "category": "Game Emulator",
    "market": args.market,
    "platform_mode": "google_play",
    "semantic_mode": "game_emulator",
    "dedup_policy": {
        "auto_merge_token_bag": True,
        "review_overlap_threshold": 0.80,
        "accent_fold_auto_merge_locales": [],
        "enable_review_log": True,
    },

    "intent_core_terms": [
        "game emulator", "retro game emulator", "retro games emulator", 
        "gba emulator", "gameboy emulator", "arcade emulator", "handheld emulator",
        "gb4 emulator", "gb4 emulador", "gb4", "gba4 emulator", "gba4 emulador", "gba4"
    ],
    
    "feature_terms": [
        # Nintendo
        'gameboy', 'gba', 'gbc', 'gb', 'game boy', 'game boy advance', 'gb advance',
        'nes', 'snes', 'n64', 'super nintendo',
        '3ds', 'ds', 'nds', 'nintendo ds',
        'gamecube', 'wii', 'wii u', 'nintendo',
        # Sony
        'ps2', 'ps3', 'ps4', 'ps5', 'playstation', 'playstation 2', 'playstation 3',
        'playstation 4', 'playstation 5',
        'psp', 'psx', 'ps vita',
        # Microsoft/Sega/Others
        'xbox', 'sega', 'dreamcast', 'arcade', 'fliperama',
        'handheld', 'console', 'portable',
        # Retro/Classic indicators
        'retro', 'classic', 'classicos', 'clássico', 'klasik',
        '8bit', '8-bit', '16bit', '16-bit', '32bit',
        'old', 'vintage', 'nostalgic', 'nostalgia',
        'jadul', 'lawas', 'advance', 'collection'
    ],
    
    "style_terms": [
        # Nintendo IP
        'pokemon', 'pokemon red', 'pokemon blue', 'pokemon fire red',
        'pokemon emerald', 'pokemon yellow', 'pokemon gold', 'pokemon silver',
        'mario', 'super mario', 'mario kart', 'zelda', 'legend of zelda',
        'metroid', 'kirby', 'donkey kong', 'star fox', 'fire emblem',
        'animal crossing', 'smash bros', 'super smash bros',
        # Sega/Sonic
        'sonic', 'sonic the hedgehog', 'tails', 'knuckles',
        'golden axe', 'streets of rage', 'shinobi',
        # Capcom/Konami/Square
        'street fighter', 'mega man', 'rockman',
        'castlevania', 'metal gear', 'silent hill',
        'final fantasy', 'dragon quest', 'chrono trigger',
        'resident evil', 'devil may cry', 'monster hunter',
        # Namco/Others
        'pacman', 'pac-man', 'galaga', 'dig dug',
        'tekken', 'soulcalibur', 'ridge racer',
        'naruto', 'dragon ball', 'dbz', 'bleach', 'one piece',
        'tetris', 'puzzle',
        # Sony IP
        'god of war', 'crash', 'crash bandicoot', 'spyro',
        'gran turismo', 'twisted metal', 'parappa',
        # Fighting/Shooters
        'mortal kombat', 'mk', 'killer instinct',
        'doom', 'quake', 'wolfenstein', 'duke nukem',
        'contra', 'metal slug', 'gunstar heroes',
        # RPG/JRPG
        'earthbound', 'mother', 'undertale',
        'persona', 'shin megami tensei', 'smt',
        'suikoden', 'wild arms', 'vagrant story',
        # Platform/Adventure
        'banjo kazooie', 'conker', 'rareware',
        'banjo-tooie', 'perfect dark',
        # Misc classic
        'bomberman', 'ice climber', 'excitebike',
        'duck hunt', 'punch out', 'kid icarus',
        'wario', 'waluigi', 'yoshi', 'luigi'
    ],
    
    "competitor_brands": [
        # Major multi-system
        'ppsspp', 'ppsspp gold', 'dolphin emulator',
        'retroarch', 'retroarch emulator',
        'delta emulator', 'delta nintendo emulator',
        'citra', 'citra emulator', 'lime3ds', 'lime3ds emulator',
        'aethersx2', 'aethersx2 emulator',
        'lemuroid', 'lemuroid emulator',
        # GBA specific
        'my boy', 'my boy emulator', 'my boy free', 'my boy gba',
        'john gba', 'john gba lite', 'john gba emulator',
        'gameboid', 'gameboid emulator',
        'gba4ios', 'gba4ios emulator',
        'vgbanext', 'vgbanext emulator',
        'gamma emulator', 'gamma game emulator',
        # NDS/3DS
        'drastic', 'drastic emulator', 'drastic ds',
        'melonDS', 'melon ds', 'desmume',
        # NES/SNES
        'snes9x', 'snes9x ex', 'zsnes',
        'nestopia', 'fceux', 'quicknes',
        'super retro 16', 'super retro 16 plus',
        # PS1/PS2
        'epsxe', 'epsxe emulator',
        'fpse', 'fpse emulator',
        'pcsx2', 'pcsx2 emulator', 'damonps2', 'damon ps2',
        'play emulator',
        # N64
        'mupen64plus', 'mupen64', 'mupen',
        'project64', 'project 64', 'n64oid',
        # Arcade/MAME
        'mame', 'mame4droid', 'mame4ios',
        'fba', 'fbneo', 'final burn alpha', 'final burn neo',
        'kawaks',
        # All-in-one
        'classicboy', 'classicboy lite',
        'retro game boy',
        'emu games', 'emu paradise',
        'romsmania', 'loveroms', 'romhustler',
        # Cloud/Remote
        'netboom', 'netboom cloud gaming',
        'airconsole', 'air console',
        'starparks', 'chikii', 'chikii cloud',
        'psplay', 'psplay remote play',
        'xbplay', 'xbplay remote play',
        # Misc
        'superpsx', 'super psx',
        'bitboy', 'bitboy emulator',
        'onecast', 'onecast xbox',
        'happy chick', 'happy chick emulator',
        'pizza emulator', 'pizza boy', 'pizza boy gba',
        'easy emu', 'mock emulator', 'lucky emulator',
        'gk emulator', 'gas emulator', 'folium emulator', 'jeans emulator',
        'emulsio', 'emulator guia', 'emulator md2',
        'emulator anak permainan', 'emulator juegos pro',
        'retro game master', 'retro game hub',
        # Browser/Non-emulator
        'dolphin browser', '870 fitness'
    ],
    
    "noise_terms": [
        'game', 'games', 'gaming', 'gamer', 'gamers', 'gameplay',
        'video game', 'videogame', 'video games', 'computer game',
        'android game', 'mobile game', 'phone game', 'tablet game',
        'play', 'playing', 'player', 'fun', 'entertainment',
        'download', 'free', 'gratis', 'premium', 'pro', 'lite',
        'best', 'top', 'new', 'old', 'latest', 'update', 'version',
        'android', 'ios', 'iphone', 'ipad', 'phone', 'mobile', 'tablet',
        'app', 'application', 'software', 'tool', 'utility', 'program',
        'device', 'system', 'platform', 'technology', 'digital',
        'emulator', 'emulador', 'emulation', 'emu', 'emulators', 'emuladores',
        'simulador', 'simulator', 'simulate', 'virtual', 'virtual machine',
        'rom', 'roms', 'iso', 'bios', 'cheat', 'cheats', 'hack', 'mod',
        'save', 'load', 'state', 'slot', 'backup', 'restore',
        'controller', 'control', 'controle', 'kontrol', 'kontroler',
        'gamepad', 'joypad', 'joystick', 'pad', 'button', 'buttons',
        'd-pad', 'dpad', 'analog', 'stick', 'trigger', 'bumper',
        'bluetooth controller', 'wireless controller', 'usb controller',
        'remote', 'remoto', 'remote play', 'second screen',
        'cloud', 'cloud gaming', 'streaming', 'stream',
        'geforce now', 'xbox cloud', 'playstation now', 'stadia',
        'nvidia', 'shadow', 'boosteroid', 'blacknut',
        'screen', 'display', 'monitor', 'resolution', 'fps', 'hz',
        'battery', 'storage', 'memory', 'ram', 'cpu', 'gpu',
        'speed', 'fast', 'slow', 'lag', 'latency', 'ping',
        'online', 'offline', 'multiplayer', 'coop', 'pvp', 'pve',
        'wifi', 'internet', 'network', 'connection', 'server',
        'account', 'login', 'register', 'profile', 'avatar',
        'chat', 'message', 'friend', 'social', 'community', 'forum',
        'rate', 'review', 'feedback', 'support', 'help', 'faq',
        'guide', 'tutorial', 'walkthrough', 'tips', 'tricks',
        'news', 'blog', 'patch', 'dlc', 'expansion',
        'skin', 'theme', 'wallpaper', 'icon', 'font', 'sound',
        'music', 'song', 'audio', 'soundtrack', 'ost', 'bgm',
        'record', 'recording', 'screenshot', 'capture', 'clip',
        'share', 'export', 'import', 'sync', 'transfer',
        'the', 'a', 'an', 'and', 'or', 'but', 'for', 'with', 'without',
        'in', 'on', 'at', 'to', 'from', 'by', 'of', 'about', 'into',
        'through', 'during', 'before', 'after', 'above', 'below',
        'between', 'among', 'within', 'against', 'under', 'over',
        'good', 'great', 'awesome', 'amazing', 'excellent', 'perfect',
        'bad', 'terrible', 'awful', 'horrible', 'worst',
        'big', 'small', 'huge', 'tiny', 'large', 'mini',
        'easy', 'hard', 'difficult', 'simple', 'complex',
        'first', 'last', 'next', 'previous', 'final',
        '1990', '1995', '2000', '2005', '2010', '2015', '2020',
        '90s', '80s', '00s', 'year', 'years', 'decade',
        'one', 'two', 'three', 'four', 'five', 'six', 'seven', 'eight', 'nine', 'ten'
    ],
    
    "typo_blacklist": [
        'gretro', 'restro', 'gaem', 'gam emulador', 'imulator', 'gbã', 'gbã emulator', 'emulsio',
        '0s5', 'pspusado', 'ps4ps5', 'ps ps ps', 'ps5ps4', 'pxp', 'pps', 'pc5', 'eio', 'psdp', 'pspp', 'ppsp', 'ssip', 'ds3',
        'pintasan', '870 fitness', 'maldives', 'dolphin browser', 'restaurați poza', 'memperlambat',
        'ukuran panjang', 'seperti apa', 'seperti apa itu', 'tujuan', 'posisi', 'tercepat',
        'calculator', 'calendar', 'weather', 'clock', 'alarm', 'reminder',
        'note', 'notes', 'file manager', 'gallery', 'camera', 'video player',
        'music player', 'audio player', 'podcast', 'radio', 'news', 'magazine',
        'shopping', 'delivery', 'food', 'restaurant', 'hotel', 'travel',
        'booking', 'ticket', 'flight', 'train', 'bus', 'map', 'navigation',
        'gps', 'tracker', 'fitness', 'health', 'medical', 'diet', 'yoga',
        'meditation', 'sleep', 'water', 'step', 'calorie', 'workout',
        'education', 'learning', 'course', 'lesson', 'quiz', 'exam', 'study',
        'language', 'dictionary', 'translator', 'keyboard', 'input method',
        'launcher', 'home screen', 'lock screen', 'live wallpaper', 'widget',
        'keyboard theme', 'icon pack', 'font', 'ringtone', 'notification sound'
    ],
    
    "risky_platform_terms": [
        "ios", "iphone", "apple", "os 17", "os 18", "os17", "os18", "ipad"
    ],

    "risky_ip_terms": ["pokemon", "mario", "zelda", "nintendo", "playstation"],
    "ambiguous_brand_terms": ["dolphin", "delta", "play"],
    "platform_affiliation_terms": [],
    "truncation_policy": {
        "enabled": True,
        "min_prefix_length": 2,
        "allowed_partial_terms": []
    },
    "risk_policy": {
        "competitor_brand_action": "drop",
        "ambiguous_brand_action": "consider",
        "risky_ip_action": "consider",
        "platform_context_action": "consider",
        "platform_only_action": "drop",
        "platform_affiliation_action": "drop",
        "style_only_action": "reserve",
        "core_intent_override": True
    },
    "user_overrides": {
        "force_top30_terms": [],
        "force_consider_terms": [],
        "force_drop_terms": []
    },
    
    "balanced_weights": {
        "VolumeN": 0.20,
        "DifficultyN": 0.15,
        "KEIN": 0.15,
        "RelevancyScore": 0.30,
        "CurrentRankN": 0.10,
        "ExpansionValue": 0.10
    }
}

from app_config import FILTER_POLICY
config.update(FILTER_POLICY)

# --- Shared Google Play profile service ---
# Build or load App Profile using seed query 'Game Emulator'
app_profile = _shared_profile_service.get_app_profile(config, "Game Emulator", os.path.dirname(__file__))

# --- Local HTTP Server for Selection & ASO Dashboard ---
def start_interactive_server(df, config, app_profile):
    import http.server
    import socketserver
    import webbrowser
    import threading
    import time
    
    keywords_data = []
    for idx, row in df.iterrows():
        keywords_data.append({
            "Keyword": row['Keyword'],
            "Volume": int(row['Volume']),
            "Difficulty": int(row['Difficulty']),
            "KEI": float(row['KEI']),
            "RelevancyScore": float(row['RelevancyScore']),
            "BalancedScore": float(row['BalancedScore']),
            "Bucket": row.get('Bucket', 'Unclassified'),
            "Rank": str(row.get('Rank', '')),
            "CompetitorProven": row.get('CompetitorProven', 'No'),
            "ProvenDetails": row.get('ProvenDetails', '')
        })
        
    data_payload = {
        "app_name": config["app_name"],
        "app_id": config["app_id"],
        "category": config["category"],
        "competitors": app_profile.get("competitors", []),
        "keywords": keywords_data
    }
    
    server_data = {"confirmed_payload": None}
    
    class SelectionHandler(http.server.BaseHTTPRequestHandler):
        def log_message(self, format, *args):
            return
            
        def do_GET(self):
            if self.path == '/':
                self.send_response(200)
                self.send_header('Content-Type', 'text/html; charset=utf-8')
                self.end_headers()
                html_path = os.path.join(os.path.dirname(__file__), 'interactive_optimizer.html')
                with open(html_path, 'r', encoding='utf-8') as f:
                    self.wfile.write(f.read().encode('utf-8'))
            elif self.path == '/data':
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps(data_payload, ensure_ascii=False).encode('utf-8'))
            else:
                self.send_response(404)
                self.end_headers()
                
        def do_POST(self):
            if self.path == '/confirm':
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                payload = json.loads(post_data.decode('utf-8'))
                server_data["confirmed_payload"] = payload
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(b'{"status": "ok"}')
                
                def shutdown_soon():
                    time.sleep(0.5)
                    server.shutdown()
                threading.Thread(target=shutdown_soon).start()
                
    port = 8000
    for p in range(8000, 8050):
        try:
            server = socketserver.TCPServer(("", p), SelectionHandler)
            port = p
            break
        except OSError:
            continue
            
    print(f"\n[INTERACTIVE SELECTOR] Starting local server at http://localhost:{port}...")
    print("Opening browser automatically... Please select keywords and write ASO descriptions in the dashboard, then click Confirm.")
    
    webbrowser.open(f"http://localhost:{port}/")
    server.serve_forever()
    server.server_close()
    print("[INTERACTIVE SELECTOR] Configuration received. Proceeding with Excel generation...")
    return server_data["confirmed_payload"]



# Step 1: Load and Clean
print("[Step 1] Loading raw candidates...")
df_raw = pd.read_csv(INPUT_PATH, encoding="utf-8-sig")

df = pd.DataFrame()
raw_cols = ['Keyword', 'Volume', 'Difficulty', 'KEI', 'Rank', 'CurrentRank', 'RankStatus', 'MaximumReach']
for col in raw_cols:
    if col in df_raw.columns:
        df[col] = df_raw[col]
    elif col == 'CurrentRank' and 'Rank' in df_raw.columns:
        df['CurrentRank'] = df_raw['Rank']
    elif col == 'Rank' and 'CurrentRank' in df_raw.columns:
        df['Rank'] = df_raw['CurrentRank']
    elif col == 'RankStatus' and 'Rank Status' in df_raw.columns:
        df['RankStatus'] = df_raw['Rank Status']
    elif col == 'MaximumReach' and 'Maximum Reach' in df_raw.columns:
        df['MaximumReach'] = df_raw['Maximum Reach']

if 'Volume' not in df.columns:
    df['Volume'] = 0
if 'Difficulty' not in df.columns:
    df['Difficulty'] = 0
if 'KEI' not in df.columns:
    df['KEI'] = 0
if 'Rank' not in df.columns:
    df['Rank'] = 'Unranked'
if 'MaximumReach' not in df.columns:
    df['MaximumReach'] = 0

df['Volume'] = pd.to_numeric(df['Volume'], errors='coerce').fillna(0).astype(int)
df['Difficulty'] = pd.to_numeric(df['Difficulty'], errors='coerce').fillna(0).astype(int)
df['KEI'] = pd.to_numeric(df['KEI'], errors='coerce').fillna(0).astype(float)
df['MaximumReach'] = pd.to_numeric(df['MaximumReach'].astype(str).str.replace(',', '', regex=False), errors='coerce').fillna(0).astype(float)

# Load Max. Volume
max_vol_col = None
for col in df_raw.columns:
    if col.lower().replace('.', '').strip() == 'max volume':
        max_vol_col = col
        break

if max_vol_col is not None:
    df['Max. Volume'] = df_raw[max_vol_col]
else:
    df['Max. Volume'] = df['Volume']

df['Max. Volume'] = pd.to_numeric(df['Max. Volume'], errors='coerce').fillna(df['Volume']).astype(int)
df['Traffic Stability'] = (df['Volume'] / df['Max. Volume']).fillna(1.0).clip(0.0, 1.0)

def get_stability_class(ratio):
    if ratio >= 0.85:
        return 'Stable'
    elif ratio >= 0.50:
        return 'Moderate'
    else:
        return 'Volatile'

df['Stability Class'] = df['Traffic Stability'].apply(get_stability_class)
df['Rank_numeric'] = pd.to_numeric(df['Rank'], errors='coerce').fillna(999)

# Singularization / Normalization
def normalize_text(text):
    return _shared_text_dedup.normalize_text(text)

df['keyword_normalized'] = df['Keyword'].apply(normalize_text)

# Step 2: Language Detection
print("[Step 2] Language classification...")

try:
    from langdetect import detect_langs, DetectorFactory
    DetectorFactory.seed = 0
    HAS_LANGDETECT = True
except ImportError:
    HAS_LANGDETECT = False

def load_english_vocab():
    vocab = set()
    # Use relative path from project root for portability
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(script_dir)
    path = os.path.join(project_root, "Docs_and_Templates", "english_words_10k.txt")
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                for line in f:
                    vocab.add(line.strip().lower())
        except Exception:
            pass
    return vocab

english_vocab = load_english_vocab()

def get_root_word(w):
    w = w.lower()
    if len(w) > 4:
        if w.endswith('ing'):
            return w[:-3]
        if w.endswith('ed'):
            return w[:-2]
        if w.endswith('es'):
            return w[:-2]
        if w.endswith('er'):
            return w[:-2]
        if w.endswith('est'):
            return w[:-3]
    if len(w) > 3:
        if w.endswith('s'):
            return w[:-1]
    return w

# --- Dynamic Google Play Country-Language Mapping Loader ---
_COUNTRY_LANG_MAP = None

def _load_country_language_map():
    global _COUNTRY_LANG_MAP
    if _COUNTRY_LANG_MAP is not None:
        return _COUNTRY_LANG_MAP
    mapping = {}
    try:
        import openpyxl
        script_dir = os.path.dirname(os.path.abspath(__file__))
        map_path = os.path.join(script_dir, "google_play_country_language_map.xlsx")
        if not os.path.exists(map_path):
            map_path = os.path.join(os.path.dirname(script_dir), "google_play_country_language_map.xlsx")
        
        if os.path.exists(map_path):
            wb = openpyxl.load_workbook(map_path, read_only=True)
            if 'Country-Language Map' in wb.sheetnames:
                ws = wb['Country-Language Map']
                for row in ws.iter_rows(min_row=5, values_only=True):
                    if len(row) >= 7:
                        country = str(row[3] or '').strip().upper()
                        primary_locale = str(row[4] or '').strip()
                        secondary_locale = str(row[6] or '').strip()
                        if not country:
                            continue
                        
                        p_langs = [primary_locale.split('-')[0].lower()] if primary_locale else []
                        s_langs = []
                        if secondary_locale and secondary_locale.lower() != 'none':
                            for s in secondary_locale.split(';'):
                                s = s.strip()
                                if s:
                                    s_langs.append(s.split('-')[0].lower())
                        mapping[country] = {'primary': p_langs, 'secondary': s_langs}
            wb.close()
    except Exception as e:
        print(f"Warning loading country-language map: {e}")
    _COUNTRY_LANG_MAP = mapping
    return _COUNTRY_LANG_MAP

def _get_language_policy(config, primary_lang):
    """Get or auto-derive market language policy, incorporating the spreadsheet mapping."""
    policy = config.get('market_language_policy', {})
    
    # 1. If explicit policy exists in the configuration, use it
    if policy.get('primary_languages') or policy.get('secondary_languages'):
        policy_primary = [l.split('-')[0].lower() for l in policy.get('primary_languages', [])]
        secondary_langs = [l.split('-')[0].lower() for l in policy.get('secondary_languages', [])]
        if not policy_primary:
            policy_primary = [primary_lang]
        return policy_primary, secondary_langs
    
    # 2. Derive policy dynamically from the Country-Language map
    market = config.get('market', 'US_EN')
    if "_" in market:
        parts = market.split("_")
        country_code = parts[0].upper()
        target_lang = parts[1].lower()
    else:
        country_code = market.upper()
        target_lang = primary_lang
        
    cmap = _load_country_language_map()
    if cmap and country_code in cmap:
        sheet_primary = cmap[country_code]['primary']
        sheet_secondary = cmap[country_code]['secondary']
        
        # If the target language is part of the spreadsheet's primary languages
        if target_lang in sheet_primary:
            policy_primary = [target_lang]
            secondary_langs = [l for l in sheet_secondary if l != target_lang]
        else:
            policy_primary = [target_lang]
            # Include the spreadsheet's primary and other secondary languages as secondary for this run
            secondary_langs = [l for l in (sheet_primary + sheet_secondary) if l != target_lang]
            
        # Ensure secondary_langs is deduplicated and doesn't contain target_lang
        secondary_langs = list(dict.fromkeys(secondary_langs))
    else:
        # Fallback to standard derivation if spreadsheet is not loaded or country not found
        policy_primary = [target_lang]
        secondary_langs = ['en'] if target_lang != 'en' else []
        
    return policy_primary, secondary_langs

def _build_eng_words_only(config):
    """Build English-only whitelist from config terms that were defined in English.
    Only uses the BASE config keys, not localized extensions."""
    eng_words = {
        'fyp', 'app', 'apps', 'free', 'download', 'android', 'new', 'best', 'top',
        'doggy', 'dogy', 'shrek', 'diy', 'pro', 'lite', 'tiktok', 'snapchat', 'instagram',
        'youtube', 'facebook', 'whatsapp', 'messenger', 'pinterest', 'google', 'play',
        '3d', 'arstudio', 'augmented', 'virtual', 'scanning', 'scanner', 'doge'
    }
    for key in ['intent_core_words', 'intent_core_terms', 'feature_terms', 'style_terms', 'visual_terms', 'noise_terms']:
        if key in config:
            for term in config[key]:
                for w in str(term).lower().split():
                    if all(c.isascii() for c in w):
                        eng_words.add(w)
    return eng_words

_eng_words_cache = _build_eng_words_only(config)

_LANGDETECT_CONFUSION = {
    'no': ['en'],
    'da': ['en'],
    'it': ['en', 'es'],
    'ro': ['en'],
    'sl': ['en'],
    'so': ['en'],
    'tl': ['es'],
    'pt': ['es'],
    'id': ['en'],
    'tr': ['es'],
    'af': ['en'],
    'cy': ['en'],
    'sw': ['en'],
}

def detect_keyword_language(kw, market_lang, config):
    def lang_match(l1, l2):
        l1, l2 = str(l1).lower(), str(l2).lower()
        return l1 == l2 or (l1 == 'fil' and l2 == 'tl') or (l1 == 'tl' and l2 == 'fil')

    kw_lower = str(kw).lower().strip()
    primary_lang = market_lang.split("_")[1].lower() if "_" in market_lang else "en"
    
    if not kw_lower:
        return primary_lang, 'PRIMARY'
    
    policy_primary, secondary_langs = _get_language_policy(config, primary_lang)
    
    words = [re.sub(r'[^a-z0-9]', '', w) for w in kw_lower.split()]
    words = [w for w in words if w]
    
    if not words:
        return primary_lang, 'PRIMARY'
    
    all_english = True
    for w in words:
        root = get_root_word(w)
        if w not in _eng_words_cache and w not in english_vocab and root not in _eng_words_cache and root not in english_vocab:
            all_english = False
            break
    
    if all_english:
        if any(lang_match('en', p) for p in policy_primary):
            return 'en', 'PRIMARY'
        elif any(lang_match('en', s) for s in secondary_langs):
            return 'en', 'SECONDARY'
        else:
            return 'en', 'FOREIGN'
    
    if HAS_LANGDETECT:
        try:
            langs = detect_langs(kw_lower)
            best_lang = langs[0].lang
            prob = langs[0].prob
            
            # Apply confusion matrix corrections for short text
            word_count = len(words)
            confusion_corrected = False
            if word_count <= 3 and best_lang in _LANGDETECT_CONFUSION:
                likely_langs = _LANGDETECT_CONFUSION[best_lang]
                # If primary_lang is in the confusion list, prefer it
                if any(lang_match(primary_lang, l) for l in likely_langs):
                    best_lang = primary_lang
                    confusion_corrected = True
                # If any secondary lang is in the confusion list
                elif any(any(lang_match(s, l) for l in likely_langs) for s in secondary_langs):
                    for s in secondary_langs:
                        if any(lang_match(s, l) for l in likely_langs):
                            best_lang = s
                            confusion_corrected = True
                            break
                # If 'en' is in confusion list and primary is English
                elif any(lang_match('en', l) for l in likely_langs) and any(lang_match('en', p) for p in policy_primary):
                    best_lang = 'en'
                    confusion_corrected = True
                # For very short keywords (1-2 words), default to primary
                elif word_count <= 2:
                    best_lang = primary_lang
                    confusion_corrected = True
            
            # Classify the detected language
            if lang_match(best_lang, primary_lang) or any(lang_match(best_lang, p) for p in policy_primary):
                return best_lang, 'PRIMARY'
            
            # For confusion-corrected results, trust the correction directly
            if confusion_corrected:
                if any(lang_match(best_lang, s) for s in secondary_langs):
                    return best_lang, 'SECONDARY'
                return best_lang, 'FOREIGN'
            
            # For non-corrected results, require higher confidence for short keywords
            min_prob = 0.85 if word_count <= 2 else 0.7 if word_count <= 3 else 0.6
            if prob >= min_prob:
                if any(lang_match(best_lang, s) for s in secondary_langs):
                    return best_lang, 'SECONDARY'
                # Only mark as FOREIGN with sufficient confidence
                return best_lang, 'FOREIGN'
            
        except Exception:
            pass
    
    return primary_lang, 'PRIMARY'

# Override the legacy local detector with the shared, market-aware implementation.
try:
    import sys
    _PROJECT_ROOT_FOR_SHARED = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    if _PROJECT_ROOT_FOR_SHARED not in sys.path:
        sys.path.insert(0, _PROJECT_ROOT_FOR_SHARED)
    from shared.language_detector import detect_keyword_language as _shared_detect_keyword_language

    def detect_keyword_language(kw, market_lang, config):
        return _shared_detect_keyword_language(kw, market_lang, config, english_vocab=english_vocab)
except Exception as e:
    print(f"Warning loading shared language detector: {e}. Falling back to legacy detector.")

# Populate language columns in df
detected_langs = []
lang_groups = []

for idx, row in df.iterrows():
    # If columns already exist in raw data, use them
    raw_lang = df_raw.loc[idx, 'DetectedLanguage'] if 'DetectedLanguage' in df_raw.columns else None
    raw_group = df_raw.loc[idx, 'LanguageGroup'] if 'LanguageGroup' in df_raw.columns else None
    
    if pd.notna(raw_lang) and pd.notna(raw_group):
        detected_langs.append(raw_lang)
        lang_groups.append(raw_group)
    else:
        lang, group = detect_keyword_language(row['Keyword'], config.get('market', 'US_EN'), config)
        detected_langs.append(lang)
        lang_groups.append(group)

df['DetectedLanguage'] = detected_langs
df['LanguageGroup'] = lang_groups

# Translate non-English keywords to English
print("[Step 2.5] Translating non-English keywords to English...")
translation_frame = _shared_translation_service.translate_dataframe(
    df, cache_path=os.path.join(_SHARED_ROOT, ".cache", "translations.sqlite3")
)
df[['EN', 'TranslationStatus', 'TranslationError']] = translation_frame

from shared import keyword_filter as _shared_keyword_filter

# Step 3: Hard filters
print("[Step 3] Hard filters...")
for warning in _shared_keyword_filter.validate_filter_config(config):
    print(f"Warning: {warning}")
_filter_runtime = _shared_keyword_filter.build_filter_runtime(config)
hard_filter_results = df.apply(lambda row: _shared_keyword_filter.evaluate_hard_filters(row, _filter_runtime), axis=1)
for column in _shared_keyword_filter.HARD_FILTER_COLUMNS:
    df[column] = hard_filter_results.apply(lambda result: result.get(column, ""))

# Naturalness Filter
print("[Step 4] Naturalness checking...")
def check_naturalness(kw, config):
    kw_lower = str(kw).lower()
    words = kw_lower.split()
    if len(words) == 0:
        return 'UNNATURAL', 'Empty keyword'
    word_counts = {}
    for w in words:
        word_counts[w] = word_counts.get(w, 0) + 1
    max_repeat = max(word_counts.values()) if word_counts else 0
    ratio = max_repeat / len(words) if words else 0
    if ratio > 0.5 and len(words) > 2:
        return 'STUFFING', 'Too many repeated words'
    has_core = any(term.lower() in kw_lower for term in config['intent_core_terms'])
    if len(words) > 6:
        if has_core:
            pass
        else:
            return 'TOO_LONG', f'Keyword has too many words ({len(words)})'
    grammar_patterns = [
        r"\b(game game|emulator emulator|play play)\b",
        r"\b(what is|how to|why do|when is|where is)\b"
    ]
    for pat in grammar_patterns:
        if re.search(pat, kw_lower):
            return 'UNNATURAL', 'Fails structural validation'
    for char in kw_lower:
        if ord(char) > 127 and char not in 'áéíóúüñ¿¡íóú':
            return 'LANGUAGE_BLEED', 'Foreign script character detected'
    return 'OK', 'Natural enough for keyword research'

naturalness = df.apply(lambda r: _shared_keyword_filter.check_naturalness(r, config), axis=1)
df['NaturalnessFlag'] = [n[0] for n in naturalness]
df['NaturalnessReason'] = [n[1] for n in naturalness]

# Scoring Logic (Game Emulator spec formulas)
print("[Step 5] Relevancy Scoring...")

# Pre-calculate competitor-proven keywords and score boost
def normalize_match_text(text):
    if not isinstance(text, str):
        return ""
    text = text.lower()
    text = re.sub(r'[^a-z0-9\s]', ' ', text)
    return re.sub(r'\s+', ' ', text).strip()

def check_keyword_in_text(kw_norm, text_norm):
    if not kw_norm or not text_norm:
        return False
    pattern = r'\b' + re.escape(kw_norm) + r'\b'
    return bool(re.search(pattern, text_norm))

competitor_proven_list = []
proven_details_list = []
competitor_boost_list = []

for idx, row in df.iterrows():
    kw_norm = normalize_match_text(row['Keyword'])
    
    comp_title_matches = []
    comp_short_matches = []
    comp_desc_matches = []
    
    for comp in app_profile.get("competitors", []):
        comp_title_norm = normalize_match_text(comp.get("title", ""))
        comp_short_norm = normalize_match_text(comp.get("short_description", ""))
        comp_desc_norm = normalize_match_text(comp.get("desc200", ""))
        comp_name = comp.get("title", comp.get("package_id", ""))
        
        if check_keyword_in_text(kw_norm, comp_title_norm):
            comp_title_matches.append(comp_name)
        if check_keyword_in_text(kw_norm, comp_short_norm):
            comp_short_matches.append(comp_name)
        if check_keyword_in_text(kw_norm, comp_desc_norm):
            comp_desc_matches.append(comp_name)
            
    boost = 0.0
    details = []
    if comp_title_matches:
        boost += 0.12
        details.append(f"Title: {', '.join(comp_title_matches)}")
    if comp_short_matches:
        boost += 0.08
        details.append(f"ShortDesc: {', '.join(comp_short_matches)}")
    if comp_desc_matches:
        boost += 0.05
        details.append(f"Desc200: {', '.join(comp_desc_matches)}")
        
    boost = min(boost, 0.15)
    
    if boost > 0:
        competitor_proven_list.append("Yes")
        proven_details_list.append("; ".join(details))
        competitor_boost_list.append(boost)
    else:
        competitor_proven_list.append("No")
        proven_details_list.append("")
        competitor_boost_list.append(0.0)

df['CompetitorProven'] = competitor_proven_list
df['ProvenDetails'] = proven_details_list
df['CompetitorBoost'] = competitor_boost_list

def calculate_relevancy(row, config):
    kw = str(row.get('EN', row['Keyword'])).lower()
    score = 0.3  # Base score
    
    # +0.40: Core emulator intent
    if 'emulator' in kw or 'emulador' in kw:
        score += 0.40
        
    # +0.15: Specific Console/Feature
    feature_match = [
        'gameboy', 'gba', 'gbc', 'nes', 'snes', 'n64',
        'ps2', 'psp', '3ds', 'sega', 'arcade', 'gamecube', 'dreamcast', 'fliperama',
        'gb4', 'gba4'
    ]
    if any(c in kw for c in feature_match):
        score += 0.15
        
    # +0.10: Retro/Classic style
    retro_match = [
        'retro', 'retrô', 'classic', 'klasik', 'clássico',
        '8bit', '8-bit', '16bit', '16-bit',
        'old', 'nostalgic', 'nostalgia', 'nostálgia', 'vintage', 'jadul', 'lawas'
    ]
    if any(r in kw for r in retro_match):
        score += 0.10
        
    # +0.05: Game title IP
    ip_match = [
        'pokemon', 'pokémon', 'mario', 'zelda', 'naruto', 'sonic',
        'tetris', 'pacman', 'metroid', 'street fighter', 'crash',
        'god of war', 'gta', 'tekken', 'wwe', 'super smash bros'
    ]
    if any(g in kw for g in ip_match):
        score += 0.05
        
    # Penalties
    if row['is_competitor']:
        score -= 0.25
    if row['is_noise']:
        score -= 0.20
    if row['LanguageGroup'] == 'FOREIGN':
        score -= 0.35
    if row['NaturalnessFlag'] != 'OK':
        score -= 0.35
    # Competitor Boost
    score += row.get('CompetitorBoost', 0.0)
        
    return max(0.0, min(1.0, score))

if 'RelevancyScore' in df_raw.columns:
    raw_relevancy = df_raw['RelevancyScore'].fillna(0.3).astype(float) + df['CompetitorBoost']
    shared_relevancy = df.apply(lambda r: _shared_keyword_filter.calculate_relevancy(r, config), axis=1)
    df['RelevancyScore'] = np.maximum(raw_relevancy, shared_relevancy)
    df['RelevancyScore'] = df['RelevancyScore'].clip(0.0, 1.0)
else:
    df['RelevancyScore'] = df.apply(lambda r: _shared_keyword_filter.calculate_relevancy(r, config), axis=1)

# Normalization & Balanced Score
print("[Step 6] Balanced Score Normalization...")
max_reach = df['MaximumReach'].max()
max_kei = df['KEI'].max()

df['VolumeN'] = df.apply(
    lambda r: _shared_keyword_filter.calculate_volume_score(
        r['Volume'], r['Max. Volume'], r['MaximumReach'], max_reach, config
    ),
    axis=1,
)
df['DifficultyN'] = 1.0 - (df['Difficulty'].clip(0, 100) / 100.0)
df['KEIN'] = np.log1p(df['KEI']) / np.log1p(max_kei) if max_kei > 0 else 0

def get_rank_n(rank_val):
    try:
        if pd.isna(rank_val) or rank_val in ['Unranked', 'unranked', 'empty', '']:
            return 0.0
        r = float(rank_val)
        if r <= 10: return 1.0
        elif r <= 30: return 0.8
        elif r <= 50: return 0.6
        elif r <= 100: return 0.3
        else: return 0.1
    except:
        return 0.0
df['CurrentRankN'] = df['Rank'].apply(get_rank_n)

def calculate_expansion(row):
    kw = str(row['Keyword']).lower()
    words = kw.split()
    n = len(words)
    if n == 1:
        score = 0.9
    elif n == 2:
        score = 0.7
    elif n == 3:
        score = 0.5
    else:
        score = 0.3
    if 'emulator' in kw or 'emulador' in kw:
        score += 0.1
    if row['is_competitor']:
        score = 0.1
    # Check if keyword has style/game IP
    has_style = any(re.search(r'\b' + re.escape(s.lower()) + r'\b', kw) for s in config['style_terms'])
    if has_style:
        score = min(score, 0.3)
    return max(0.0, min(1.0, score))

df['ExpansionValue'] = df.apply(lambda r: _shared_keyword_filter.calculate_expansion(r, config), axis=1)

bw = config['balanced_weights']
df['BalancedScore'] = (
    bw['VolumeN'] * df['VolumeN'] +
    bw['DifficultyN'] * df['DifficultyN'] +
    bw['KEIN'] * df['KEIN'] +
    bw['RelevancyScore'] * df['RelevancyScore'] +
    bw['CurrentRankN'] * df['CurrentRankN'] +
    bw['ExpansionValue'] * df['ExpansionValue']
)

# Add a small bonus based on LanguageGroup: +0.02 for PRIMARY, +0.01 for SECONDARY
def get_language_bonus(row):
    lg = str(row.get('LanguageGroup', 'PRIMARY')).upper()
    if lg == 'PRIMARY':
        return 0.02
    elif lg == 'SECONDARY':
        return 0.01
    return 0.0

df['BalancedScore'] = (df['BalancedScore'] + df.apply(get_language_bonus, axis=1)).round(4)
df['RelevancyScore'] = df['RelevancyScore'].round(4)

# Bucket classification (Game Emulator Mode)
print("[Step 7] Bucket classification...")
def classify_keyword(row, config):
    kw = str(row['Keyword']).lower()
    
    # Hard drops
    if row['is_competitor']:
        return 'Dropped', 'competitor_brand', 'Dropped: Competitor brand'
    if row['is_typo']:
        return 'Dropped', 'typo_truncated_broken', 'Dropped: Typo, truncated, or broken'
    if row['is_noise']:
        return 'Dropped', 'noise_only', 'Dropped: Noise-only generic term'
    if row['NaturalnessFlag'] != 'OK':
        return 'Dropped', 'unnatural', f"Dropped: Unnatural phrase ({row['NaturalnessReason']})"
        
    # Language Policy
    if row['LanguageGroup'] == 'FOREIGN':
        return 'Language Mismatch Audit', 'foreign_language_mismatch', 'Foreign language mismatch'
    if row['LanguageGroup'] in ['MIXED', 'UNKNOWN']:
        return 'Manual Review', 'manual_review', 'Mixed or unknown language'
    if row['LanguageGroup'] == 'SECONDARY':
        return 'Consider Keywords', 'secondary_language_handling', 'Spanish keyword for US_EN (Secondary language)'
        
    # Platform Risk
    has_platform_risk = any(term in kw for term in config['risky_platform_terms'])
    if has_platform_risk:
        return 'Consider Keywords', 'platform_style_risk', 'Platform-style risk; keep for review only'
        
    # Check console/feature and game titles
    has_core = any(term in kw for term in config['intent_core_terms'])
    has_feature = any(re.search(r'\b' + re.escape(f.lower()) + r'\b', kw) for f in config['feature_terms'])
    has_style = any(re.search(r'\b' + re.escape(s.lower()) + r'\b', kw) for s in config['style_terms'])
    
    if has_core:
        return 'Core Intent Final', 'core_intent_final', 'Strong core game emulator search intent'
        
    if has_style:
        generic_game_terms = ["retro games", "classic games", "gba games", "arcade games", "jogos retrô", "jogos gba", "game emulator"]
        if any(term in kw for term in generic_game_terms):
            return 'Broad Expansion', 'broad_expansion', 'Generic game/emulator variant'
        else:
            return 'Game Keywords', 'game_keywords', 'Game Title/Franchise candidate (Research Only)'
            
    if has_feature:
        return 'System Keywords', 'system_keywords', 'System/Console candidate'
        
    if row['RelevancyScore'] < 0.45:
        return 'Dropped', 'dropped', 'Dropped: Weak app intent after scoring'
        
    return 'Broad Expansion', 'broad_expansion', 'Moderately relevant emulator expansion'

classifications = df.apply(lambda r: _shared_keyword_filter.classify_keyword(r, config), axis=1)
df['Bucket'] = [c[0] for c in classifications]
df['DecisionRule'] = [c[1] for c in classifications]
df['Reason'] = [c[2] for c in classifications]

def override_row(row):
    bucket, rule, reason = _shared_keyword_filter.apply_user_overrides(row, config)
    return pd.Series([bucket, rule, reason])

df[['Bucket', 'DecisionRule', 'Reason']] = df.apply(override_row, axis=1)

# Shortlist building & duplicate checking
print("[Step 8] Near-Duplicate Cleanup & Shortlist building...")
def build_shortlist(df_all, config):
    eligible_buckets = ['Core Intent Final', 'System Keywords', 'Broad Expansion', 'Game Keywords', 'Consider Keywords']
    df_candidates = df_all[df_all['Bucket'].isin(eligible_buckets)]
    df_sorted, dedup_log = _shared_text_dedup.prepare_dataframe(df_candidates, '01_Main_Keyword_Shortlist', config)
    df_sorted = df_sorted.sort_values(by=['BalancedScore', 'Rank_numeric', 'KEI', 'Difficulty'], ascending=[False, True, False, True]).copy()
    selected_core, selected_broad, selected_consider = [], [], []
    selected_normalized, selected_tokens = set(), set()

    def volume_eligible(row, section):
        low_tier_count = sum(_shared_keyword_filter.is_low_volume_tier(item, config) for item in selected_consider)
        return _shared_keyword_filter.is_shortlist_volume_eligible(row, section, low_tier_count, config)
    
    def check_duplicate(kw, original_bucket):
        norm = normalize_text(kw)
        tokens = " ".join(sorted(norm.split()))
        if not norm:
            return True, "Empty normalized keyword", ""
        if norm in selected_normalized:
            all_selected = selected_core + selected_broad + selected_consider
            kept_kw = ""
            for item in all_selected:
                if normalize_text(item['Keyword']) == norm:
                    kept_kw = item['Keyword']
                    break
            return True, f"Exact normalized duplicate of '{kept_kw}'", kept_kw
        if tokens in selected_tokens:
            all_selected = selected_core + selected_broad + selected_consider
            kept_kw = ""
            for item in all_selected:
                t = " ".join(sorted(normalize_text(item['Keyword']).split()))
                if t == tokens:
                    kept_kw = item['Keyword']
                    break
            return True, f"Same normalized token set as '{kept_kw}'", kept_kw
        return False, "", ""
        
    def add_to_shortlist(item, section):
        norm = normalize_text(item['Keyword'])
        tokens = " ".join(sorted(norm.split()))
        selected_normalized.add(norm)
        selected_tokens.add(tokens)
        entry = item.to_dict()
        entry['Section'] = section
        entry['QuotaStatus'] = 'EXACT'
        entry['FillSource'] = ''
        entry['FillReason'] = ''
        return entry

    # Core (Quota: 25)
    core_candidates = df_sorted[df_sorted['Bucket'] == 'Core Intent Final']
    for _, row in core_candidates.iterrows():
        if len(selected_core) >= 25:
            break
        if not volume_eligible(row, 'Core Intent Final'):
            continue
        is_dup, reason, kept_kw = check_duplicate(row['Keyword'], 'Core Intent Final')
        if is_dup:
            dedup_log.append({
                'Table': '01_Main_Keyword_Shortlist', 'RemovedKeyword': row['Keyword'],
                'OriginalSection': 'Core Intent Final', 'KeptKeyword': kept_kw,
                'DedupReason': reason, 'BalancedScore': row['BalancedScore'],
                'Note': 'Keyword remains in All Candidates pool'
            })
        else:
            selected_core.append(add_to_shortlist(row, 'Core Intent Final'))
            
    # Core Fallback
    if len(selected_core) < 25:
        fallback_candidates = df_sorted[df_sorted['Bucket'].isin(['System Keywords', 'Broad Expansion'])]
        for _, row in fallback_candidates.iterrows():
            if len(selected_core) >= 25:
                break
            if not volume_eligible(row, 'Core Intent Final'):
                continue
            norm = normalize_text(row['Keyword'])
            if norm in selected_normalized:
                continue
            if row['RelevancyScore'] < 0.45:
                continue
            is_dup, reason, kept_kw = check_duplicate(row['Keyword'], 'Core Intent Final (Fallback)')
            if is_dup:
                dedup_log.append({
                    'Table': '01_Main_Keyword_Shortlist', 'RemovedKeyword': row['Keyword'],
                    'OriginalSection': 'Core Intent Final (Fallback)', 'KeptKeyword': kept_kw,
                    'DedupReason': reason, 'BalancedScore': row['BalancedScore'],
                    'Note': 'Keyword remains in All Candidates pool'
                })
            else:
                entry = add_to_shortlist(row, 'Core Intent Final')
                entry['QuotaStatus'] = 'FILLED'
                entry['FillSource'] = row['Bucket']
                entry['FillReason'] = 'Core Intent Quota Fallback'
                selected_core.append(entry)

    # Broad (Quota: 5)
    broad_candidates = df_sorted[df_sorted['Bucket'] == 'Broad Expansion']
    for _, row in broad_candidates.iterrows():
        if len(selected_broad) >= 5:
            break
        if not volume_eligible(row, 'Broad Expansion'):
            continue
        is_dup, reason, kept_kw = check_duplicate(row['Keyword'], 'Broad Expansion')
        if is_dup:
            dedup_log.append({
                'Table': '01_Main_Keyword_Shortlist', 'OriginalSection': 'Broad Expansion',
                'RemovedKeyword': row['Keyword'], 'KeptKeyword': kept_kw,
                'DedupReason': reason, 'BalancedScore': row['BalancedScore'],
                'Note': 'Keyword remains in All Candidates pool'
            })
        else:
            selected_broad.append(add_to_shortlist(row, 'Broad Expansion'))
            
    # Broad Fallback
    if len(selected_broad) < 5:
        fallback_candidates = df_sorted[df_sorted['Bucket'].isin(['System Keywords', 'Game Keywords'])]
        for _, row in fallback_candidates.iterrows():
            if len(selected_broad) >= 5:
                break
            if not volume_eligible(row, 'Broad Expansion'):
                continue
            norm = normalize_text(row['Keyword'])
            if norm in selected_normalized:
                continue
            if row['RelevancyScore'] < 0.45:
                continue
            is_dup, reason, kept_kw = check_duplicate(row['Keyword'], 'Broad Expansion (Fallback)')
            if is_dup:
                dedup_log.append({
                    'Table': '01_Main_Keyword_Shortlist', 'RemovedKeyword': row['Keyword'],
                    'OriginalSection': 'Broad Expansion (Fallback)', 'KeptKeyword': kept_kw,
                    'DedupReason': reason, 'BalancedScore': row['BalancedScore'],
                    'Note': 'Keyword remains in All Candidates pool'
                })
            else:
                entry = add_to_shortlist(row, 'Broad Expansion')
                entry['QuotaStatus'] = 'FILLED'
                entry['FillSource'] = row['Bucket']
                entry['FillReason'] = 'Broad Expansion Quota Fallback'
                selected_broad.append(entry)

    # Consider (Quota: 10)
    consider_candidates = df_sorted[df_sorted['Bucket'] == 'Consider Keywords']
    for _, row in consider_candidates.iterrows():
        if len(selected_consider) >= 10:
            break
        if not volume_eligible(row, 'Consider Keywords'):
            continue
        is_dup, reason, kept_kw = check_duplicate(row['Keyword'], 'Consider Keywords')
        if is_dup:
            dedup_log.append({
                'Table': '01_Main_Keyword_Shortlist', 'RemovedKeyword': row['Keyword'],
                'OriginalSection': 'Consider Keywords', 'KeptKeyword': kept_kw,
                'DedupReason': reason, 'BalancedScore': row['BalancedScore'],
                'Note': 'Keyword remains in All Candidates pool'
            })
        else:
            selected_consider.append(add_to_shortlist(row, 'Consider Keywords'))
            
    # Consider Fallback
    if len(selected_consider) < 10:
        selected_kws = {item['Keyword'].lower() for item in selected_core + selected_broad}
        missed_opps = df_sorted[df_sorted['Bucket'].isin(['Core Intent Final', 'Broad Expansion']) & 
                                (~df_sorted['Keyword'].str.lower().isin(selected_kws))]
        for _, row in missed_opps.iterrows():
            if len(selected_consider) >= 10:
                break
            if not volume_eligible(row, 'Consider Keywords'):
                continue
            norm = normalize_text(row['Keyword'])
            if norm in selected_normalized:
                continue
            if row['RelevancyScore'] < 0.45:
                continue
            is_dup, reason, kept_kw = check_duplicate(row['Keyword'], 'Consider Keywords (Fallback)')
            if is_dup:
                dedup_log.append({
                    'Table': '01_Main_Keyword_Shortlist', 'RemovedKeyword': row['Keyword'],
                    'OriginalSection': 'Consider Keywords (Fallback)', 'KeptKeyword': kept_kw,
                    'DedupReason': reason, 'BalancedScore': row['BalancedScore'],
                    'Note': 'Keyword remains in All Candidates pool'
                })
            else:
                entry = add_to_shortlist(row, 'Consider Keywords')
                entry['QuotaStatus'] = 'FILLED'
                entry['FillSource'] = row['Bucket']
                entry['FillReason'] = 'Consider Keywords Quota Fallback (Missed Opportunity)'
                selected_consider.append(entry)

    return selected_core, selected_broad, selected_consider, dedup_log

selected_core, selected_broad, selected_consider, dedup_log_list = build_shortlist(df, config)

def build_curated_sheet(df_all, bucket_name, sheet_name):
    df_sorted, dedup_entries = _shared_text_dedup.prepare_dataframe(df_all[df_all['Bucket'] == bucket_name], sheet_name, config)
    df_sorted = df_sorted.sort_values(by=['BalancedScore', 'Rank_numeric', 'KEI', 'Difficulty'], ascending=[False, True, False, True]).copy()
    selected = []
    selected_normalized, selected_tokens = set(), set()
    
    for _, row in df_sorted.iterrows():
        if len(selected) >= 30:
            break
        norm = normalize_text(row['Keyword'])
        tokens = " ".join(sorted(norm.split()))
        
        is_dup = False
        reason = ""
        kept_kw = ""
        
        if norm in selected_normalized:
            is_dup = True
            for item in selected:
                if normalize_text(item['Keyword']) == norm:
                    kept_kw = item['Keyword']
                    break
            reason = f"Exact normalized duplicate of '{kept_kw}'"
        elif tokens in selected_tokens:
            is_dup = True
            for item in selected:
                t = " ".join(sorted(normalize_text(item['Keyword']).split()))
                if t == tokens:
                    kept_kw = item['Keyword']
                    break
            reason = f"Same normalized token set as '{kept_kw}'"
            
        if is_dup:
            dedup_entries.append({
                'Table': sheet_name, 'RemovedKeyword': row['Keyword'],
                'OriginalSection': bucket_name, 'KeptKeyword': kept_kw,
                'DedupReason': reason, 'BalancedScore': row['BalancedScore'],
                'Note': 'Keyword remains in All Candidates pool'
            })
        else:
            selected_normalized.add(norm)
            selected_tokens.add(tokens)
            entry = row.to_dict()
            entry['Section'] = bucket_name
            entry['QuotaStatus'] = 'EXACT'
            entry['FillSource'] = ''
            entry['FillReason'] = ''
            selected.append(entry)
            
    return selected, dedup_entries

# System Keywords (capped <=30, no fallback fill)
selected_feature, dedup_feat = build_curated_sheet(df, 'System Keywords', '02_System_Keywords')
# Game Keywords (capped <=30, no fallback fill)
selected_style, dedup_style = build_curated_sheet(df, 'Game Keywords', '03_Game_Keywords')

dedup_log_list.extend(dedup_feat)
dedup_log_list.extend(dedup_style)
df_dedup_log = pd.DataFrame(_shared_text_dedup.normalize_log_entries(dedup_log_list))

# Metadata assignment
print("[Step 9] Metadata slot assignment...")

# Start Interactive Selector Dashboard
selections_file = _shared_keyword_filter.selection_cache_path(os.path.dirname(__file__), config, INPUT_PATH, config.get("market", ""))
selection_cache_meta = _shared_keyword_filter.build_selection_cache_meta(INPUT_PATH, config.get("market", ""), config)

confirmed_selection = None

if os.path.exists(selections_file):
    print(f"\n[Step 9] Found existing keyword selections in {selections_file}. Checking cache metadata...")
    with open(selections_file, "r", encoding="utf-8") as f:
        cached_selection = json.load(f)
    if _shared_keyword_filter.is_selection_cache_valid(cached_selection, selection_cache_meta):
        confirmed_selection, _ = _shared_keyword_filter.unwrap_selection_payload(cached_selection)
        print("[Step 9] Selection cache matches current input. Loading cached selections...")
    else:
        print("[Step 9] Selection cache does not match current input/market. Ignoring cached selections.")
else:
    if args.interactive:
        confirmed_selection = start_interactive_server(df, config, app_profile)
        if confirmed_selection:
            payload = _shared_keyword_filter.wrap_selection_payload(confirmed_selection, selection_cache_meta)
            _shared_keyword_filter.atomic_write_json(selections_file, payload)
                
            print("\n" + "="*50)
            print("[SELECTION_CONFIRMED] Keyword selections successfully saved!")
            print("Selections saved to:", selections_file)
            print("="*50 + "\n")
    else:
        print("\n[Step 9] Run in non-interactive/headless mode. Using defaults from shortlist logic.")

if confirmed_selection:
    user_core = confirmed_selection.get("core_keywords", [])
    user_secondary = confirmed_selection.get("secondary_keywords", [])
    user_feature = confirmed_selection.get("feature_keywords", [])
    user_style = confirmed_selection.get("style_keywords", [])
    
    df_lookup = df.set_index('Keyword')
    
    selected_core = []
    for kw in user_core:
        if kw in df_lookup.index:
            row = df_lookup.loc[kw]
            if isinstance(row, pd.DataFrame):
                row = row.iloc[0]
            entry = row.to_dict()
            entry['Keyword'] = kw
            entry['Section'] = 'Core Intent Final'
            entry['QuotaStatus'] = 'EXACT'
            selected_core.append(entry)
            
    selected_broad = []
    selected_consider = []
    for idx, kw in enumerate(user_secondary):
        if kw in df_lookup.index:
            row = df_lookup.loc[kw]
            if isinstance(row, pd.DataFrame):
                row = row.iloc[0]
            entry = row.to_dict()
            entry['Keyword'] = kw
            if idx < 5:
                entry['Section'] = 'Broad Expansion'
                selected_broad.append(entry)
            else:
                entry['Section'] = 'Consider Keywords'
                selected_consider.append(entry)
                
    selected_feature = []
    for kw in user_feature:
        if kw in df_lookup.index:
            row = df_lookup.loc[kw]
            if isinstance(row, pd.DataFrame):
                row = row.iloc[0]
            entry = row.to_dict()
            entry['Keyword'] = kw
            entry['Section'] = 'System Keywords'
            selected_feature.append(entry)
            
    selected_style = []
    for kw in user_style:
        if kw in df_lookup.index:
            row = df_lookup.loc[kw]
            if isinstance(row, pd.DataFrame):
                row = row.iloc[0]
            entry = row.to_dict()
            entry['Keyword'] = kw
            entry['Section'] = 'Game Keywords'
            selected_style.append(entry)
            
    config["app_title_draft"] = confirmed_selection.get("title", "")
    config["short_desc_draft"] = confirmed_selection.get("short_description", "")
    config["full_desc_draft"] = confirmed_selection.get("full_description", "")

all_shortlist = selected_core + selected_broad + selected_consider
for idx, entry in enumerate(all_shortlist):
    sec = entry['Section']
    if sec == 'Core Intent Final':
        if idx < 2:
            entry['WhereToUse'] = '🏷️ Title'
        elif idx < 9:
            entry['WhereToUse'] = '📱 Short Description'
        else:
            entry['WhereToUse'] = '📄 Full Description'
    elif sec == 'Broad Expansion':
        entry['WhereToUse'] = '📄 Full Description'
    else:
        entry['WhereToUse'] = '🔍 Consider / Research Only'

df_shortlist = pd.DataFrame(all_shortlist)

# Stylization & Export
print("[Step 10] Exporting to stylized Excel Workbook...")
wb = Workbook()
wb.remove(wb.active)

def style_sheet(ws, title, is_report=False):
    ws.views.sheetView[0].showGridLines = True
    if not is_report and ws.max_row > 1:
        ws.freeze_panes = 'A2'
        
    navy_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    white_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin = Side(border_style="thin", color="D3D3D3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    if not is_report:
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = navy_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border
                
                col_name = ws.cell(row=1, column=col_idx).value
                if col_name in ['Volume', 'Max. Volume', 'Difficulty', 'Rank', 'MaximumReach']:
                    try: cell.value = int(float(cell.value))
                    except: pass
                elif col_name in ['BalancedScore', 'RelevancyScore', 'KEI', 'VolumeN', 'DifficultyN', 'KEIN', 'CurrentRankN', 'OpportunityRankGap', 'ExpansionValue', 'Traffic Stability']:
                    try:
                        cell.value = round(float(cell.value), 4)
                        if col_name == 'Traffic Stability':
                            cell.number_format = '0.00%'
                        else:
                            cell.number_format = '0.0000'
                    except: pass
                        
    for col in ws.columns:
        max_len = 0
        for cell in col:
            val_str = str(cell.value or '')
            if '\n' in val_str:
                val_str = max(val_str.split('\n'), key=len)
            max_len = max(max_len, len(val_str))
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
        # Hide Traffic Stability and Stability Class columns
        col_name = ws.cell(row=1, column=col[0].column).value
        if col_name in ['Traffic Stability', 'Stability Class']:
            ws.column_dimensions[col_letter].hidden = True

# --- 00_README_CONFIG ---
ws_readme = wb.create_sheet(title="00_README_CONFIG")
ws_readme.views.sheetView[0].showGridLines = True
ws_readme.cell(row=1, column=1, value="ASO Keyword Planner v3.6 - Configuration Summary").font = Font(size=14, bold=True)
configs = [
    ("Pipeline Version", "ASO Keyword Planner v3.6"),
    ("App Name", config["app_name"]),
    ("App ID", config["app_id"]),
    ("Category", config["category"]),
    ("Market", config["market"]),
    ("Platform", config["platform_mode"]),
    ("Profile Status", app_profile.get("ProfileStatus", "")),
    ("Profile Error", app_profile.get("ProfileError", "")),
    ("Run Date", pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"))
]
if "app_title_draft" in config:
    configs.append(("Custom Title Draft", config["app_title_draft"]))
if "short_desc_draft" in config:
    configs.append(("Custom Short Desc Draft", config["short_desc_draft"]))
if "full_desc_draft" in config:
    configs.append(("Custom Full Desc Draft", config["full_desc_draft"]))
for idx, (lbl, val) in enumerate(configs, 3):
    ws_readme.cell(row=idx, column=1, value=lbl).font = Font(bold=True)
    ws_readme.cell(row=idx, column=2, value=val)
    ws_readme.cell(row=idx, column=1).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws_readme.cell(row=idx, column=2).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
ws_readme.column_dimensions['A'].width = 25
ws_readme.column_dimensions['B'].width = 80

# --- 01_Main_Keyword_Shortlist ---
ws_shortlist = wb.create_sheet(title="01_Main_Keyword_Shortlist")
cols_shortlist = ['Keyword', 'EN', 'TranslationStatus', 'TranslationError', 'Volume', 'Max. Volume', 'MaximumReach', 'VolumeN', 'Difficulty', 'KEI', 'Rank', 'BalancedScore', 'Traffic Stability', 'Stability Class', 'Section', 'RelevancyScore', 'MergedVariants', 'ReviewVariants',
                  'CompetitorProven', 'ProvenDetails', 'DetectedLanguage', 'LanguageGroup', 'NaturalnessFlag', 'WhereToUse', 'QuotaStatus', 'FillSource', 'FillReason', 'Reason']
for col_idx, col in enumerate(cols_shortlist, 1):
    ws_shortlist.cell(row=1, column=col_idx, value=col)
for row_idx, entry in enumerate(all_shortlist, 2):
    for col_idx, col in enumerate(cols_shortlist, 1):
        ws_shortlist.cell(row=row_idx, column=col_idx, value=entry.get(col, ''))
style_sheet(ws_shortlist, "01_Main_Keyword_Shortlist")

# --- 02_System_Keywords ---
ws_system = wb.create_sheet(title="02_System_Keywords")
cols_curated = ['Keyword', 'EN', 'TranslationStatus', 'TranslationError', 'Volume', 'Max. Volume', 'MaximumReach', 'VolumeN', 'Difficulty', 'KEI', 'Rank', 'BalancedScore', 'Traffic Stability', 'Stability Class', 'Section', 'RelevancyScore', 'MergedVariants', 'ReviewVariants', 'Reason']
for col_idx, col in enumerate(cols_curated, 1):
    ws_system.cell(row=1, column=col_idx, value=col)
for row_idx, entry in enumerate(selected_feature, 2):
    for col_idx, col in enumerate(cols_curated, 1):
        ws_system.cell(row=row_idx, column=col_idx, value=entry.get(col, ''))
style_sheet(ws_system, "02_System_Keywords")

# --- 03_Game_Keywords ---
ws_game = wb.create_sheet(title="03_Game_Keywords")
for col_idx, col in enumerate(cols_curated, 1):
    ws_game.cell(row=1, column=col_idx, value=col)
for row_idx, entry in enumerate(selected_style, 2):
    for col_idx, col in enumerate(cols_curated, 1):
        ws_game.cell(row=row_idx, column=col_idx, value=entry.get(col, ''))
style_sheet(ws_game, "03_Game_Keywords")

# --- 04_Dropped_Audit ---
ws_dropped = wb.create_sheet(title="04_Dropped_Audit")
df_dropped = df[df['Bucket'] == 'Dropped'].sort_values(by=['BalancedScore', 'Rank_numeric', 'KEI', 'Difficulty'], ascending=[False, True, False, True])
cols_audit = ['Keyword', 'EN', 'TranslationStatus', 'TranslationError', 'Volume', 'Max. Volume', 'MaximumReach', 'VolumeN', 'Difficulty', 'KEI', 'Rank', 'BalancedScore', 'Traffic Stability', 'Stability Class', 'RelevancyScore', 'DecisionRule', 'Reason', 'HardFilterRule', 'HardFilterTerm', 'HardFilterSource', 'PolicyFlags']
for col_idx, col in enumerate(cols_audit, 1):
    ws_dropped.cell(row=1, column=col_idx, value=col)
for row_idx, (_, row) in enumerate(df_dropped.iterrows(), 2):
    for col_idx, col in enumerate(cols_audit, 1):
        ws_dropped.cell(row=row_idx, column=col_idx, value=row.get(col, ''))
style_sheet(ws_dropped, "04_Dropped_Audit")

# --- 05_Report_Summary ---
ws_report = wb.create_sheet(title="05_Report_Summary")
ws_report.views.sheetView[0].showGridLines = True
ws_report.cell(row=1, column=1, value="ASO Keyword Planner v3.6 - Report Summary").font = Font(size=14, bold=True)
ws_report.cell(row=3, column=1, value="Metric Summary").font = Font(size=12, bold=True)
metrics = [
    ("Total Raw Keywords", len(df)),
    ("Unique Keywords", df['Keyword'].nunique()),
    ("Dropped Keywords", len(df_dropped)),
    ("Core Intent Selected", len(selected_core)),
    ("Broad Expansion Selected", len(selected_broad)),
    ("Consider Selected", len(selected_consider)),
    ("System Keywords Curated (02)", len(selected_feature)),
    ("Game Keywords Curated (03)", len(selected_style)),
    ("Text Dedup Log Entries (PRUNED + REVIEW)", len(df_dedup_log))
]
for idx, (lbl, val) in enumerate(metrics, 4):
    ws_report.cell(row=idx, column=1, value=lbl).font = Font(bold=True)
    ws_report.cell(row=idx, column=2, value=val)

ws_report.cell(row=15, column=1, value="Language Summary").font = Font(size=12, bold=True)
lang_counts = df['LanguageGroup'].value_counts()
for idx, (lang_g, count) in enumerate(lang_counts.items(), 17):
    ws_report.cell(row=idx, column=1, value=lang_g).font = Font(bold=True)
    ws_report.cell(row=idx, column=2, value=count)

ws_report.cell(row=24, column=1, value="Naturalness Summary").font = Font(size=12, bold=True)
nat_counts = df['NaturalnessFlag'].value_counts()
for idx, (flag, count) in enumerate(nat_counts.items(), 26):
    ws_report.cell(row=idx, column=1, value=flag).font = Font(bold=True)
    ws_report.cell(row=idx, column=2, value=count)

ws_report.cell(row=3, column=4, value="Sheet Index").font = Font(size=12, bold=True)
sheets_info = [
    ("00_README_CONFIG", "App configuration parameters and run metadata"),
    ("01_Main_Keyword_Shortlist", "Top 25 Core + 5 Broad + 10 Consider shortlist for metadata allocation"),
    ("02_System_Keywords", "Curated system, console, and platform candidates (capped <= 30)"),
    ("03_Game_Keywords", "Curated game title and franchise candidates (capped <= 30, Research Only)"),
    ("04_Dropped_Audit", "Dropped keywords with detailed reasons"),
    ("05_Report_Summary", "Summary stats, language breakdowns, and sheet indices"),
    ("06_All_Candidates", "Full candidate pool with detailed score and policy values"),
    ("07_Language_Mismatch", "Audit sheet for keywords mismatching US_EN market language"),
    ("08_Generic_Style_Reserve", "Broad style-only keywords held back from metadata shortlist"),
    ("09_Manual_Review", "Audit sheet for keywords flagged with mixed or unknown languages"),
    ("10_Top_By_Score", "Candidates sorted by BalancedScore before diversity overlap filtering"),
    ("11_Secondary_Language", "Research candidates matching Spanish (Secondary Language)"),
    ("12_Text_Dedup_Log", "Log of text-level duplicates and review candidates during optimization")
]
for idx, (title, purpose) in enumerate(sheets_info, 5):
    ws_report.cell(row=idx, column=4, value=title).font = Font(bold=True)
    ws_report.cell(row=idx, column=5, value=purpose)

thin_border = Border(left=Side(style='thin', color='C0C0C0'), right=Side(style='thin', color='C0C0C0'), 
                     top=Side(style='thin', color='C0C0C0'), bottom=Side(style='thin', color='C0C0C0'))
for r in range(4, 13):
    ws_report.cell(row=r, column=1).border = thin_border
    ws_report.cell(row=r, column=2).border = thin_border
for r in range(17, 17 + len(lang_counts)):
    ws_report.cell(row=r, column=1).border = thin_border
    ws_report.cell(row=r, column=2).border = thin_border
for r in range(26, 26 + len(nat_counts)):
    ws_report.cell(row=r, column=1).border = thin_border
    ws_report.cell(row=r, column=2).border = thin_border
for r in range(5, 5 + len(sheets_info)):
    ws_report.cell(row=r, column=4).border = thin_border
    ws_report.cell(row=r, column=5).border = thin_border

ws_report.column_dimensions['A'].width = 30
ws_report.column_dimensions['B'].width = 15
ws_report.column_dimensions['D'].width = 25
ws_report.column_dimensions['E'].width = 65

# --- 06_All_Candidates ---
ws_all = wb.create_sheet(title="06_All_Candidates")
cols_all = ['Keyword', 'EN', 'TranslationStatus', 'TranslationError', 'Volume', 'Max. Volume', 'MaximumReach', 'VolumeN', 'Difficulty', 'KEI', 'Rank', 'BalancedScore', 'Traffic Stability', 'Stability Class', 'RelevancyScore', 'CompetitorProven', 'ProvenDetails', 'Bucket',
            'DetectedLanguage', 'LanguageGroup', 'NaturalnessFlag', 'Reason', 'HardFilterRule', 'HardFilterTerm', 'HardFilterSource', 'PolicyFlags']
for col_idx, col in enumerate(cols_all, 1):
    ws_all.cell(row=1, column=col_idx, value=col)
for row_idx, (_, row) in enumerate(df.sort_values(by=['BalancedScore', 'Rank_numeric', 'KEI', 'Difficulty'], ascending=[False, True, False, True]).iterrows(), 2):
    for col_idx, col in enumerate(cols_all, 1):
        ws_all.cell(row=row_idx, column=col_idx, value=row.get(col, ''))
style_sheet(ws_all, "06_All_Candidates")

# --- 07_Language_Mismatch ---
ws_lang_m = wb.create_sheet(title="07_Language_Mismatch")
df_lang_m = df[df['Bucket'] == 'Language Mismatch Audit'].sort_values(by=['BalancedScore', 'Rank_numeric', 'KEI', 'Difficulty'], ascending=[False, True, False, True])
for col_idx, col in enumerate(cols_curated, 1):
    ws_lang_m.cell(row=1, column=col_idx, value=col)
for row_idx, (_, row) in enumerate(df_lang_m.iterrows(), 2):
    for col_idx, col in enumerate(cols_curated, 1):
        ws_lang_m.cell(row=row_idx, column=col_idx, value=row.get(col, ''))
style_sheet(ws_lang_m, "07_Language_Mismatch")

# --- 08_Generic_Style_Reserve ---
ws_reserve = wb.create_sheet(title="08_Generic_Style_Reserve")
df_reserve = df[df['Bucket'] == 'Generic Style Reserve'].sort_values(by=['BalancedScore', 'Rank_numeric', 'KEI', 'Difficulty'], ascending=[False, True, False, True])
for col_idx, col in enumerate(cols_curated, 1):
    ws_reserve.cell(row=1, column=col_idx, value=col)
for row_idx, (_, row) in enumerate(df_reserve.iterrows(), 2):
    for col_idx, col in enumerate(cols_curated, 1):
        ws_reserve.cell(row=row_idx, column=col_idx, value=row.get(col, ''))
style_sheet(ws_reserve, "08_Generic_Style_Reserve")

# --- 09_Manual_Review ---
ws_mrev = wb.create_sheet(title="09_Manual_Review")
df_mrev = df[df['Bucket'] == 'Manual Review'].sort_values(by=['BalancedScore', 'Rank_numeric', 'KEI', 'Difficulty'], ascending=[False, True, False, True])
for col_idx, col in enumerate(cols_curated, 1):
    ws_mrev.cell(row=1, column=col_idx, value=col)
for row_idx, (_, row) in enumerate(df_mrev.iterrows(), 2):
    for col_idx, col in enumerate(cols_curated, 1):
        ws_mrev.cell(row=row_idx, column=col_idx, value=row.get(col, ''))
style_sheet(ws_mrev, "09_Manual_Review")

# --- 10_Top_By_Score ---
ws_tps = wb.create_sheet(title="10_Top_By_Score")
df_tps = df.sort_values(by=['BalancedScore', 'Rank_numeric', 'KEI', 'Difficulty'], ascending=[False, True, False, True]).head(50)
for col_idx, col in enumerate(cols_curated, 1):
    ws_tps.cell(row=1, column=col_idx, value=col)
for row_idx, (_, row) in enumerate(df_tps.iterrows(), 2):
    for col_idx, col in enumerate(cols_curated, 1):
        ws_tps.cell(row=row_idx, column=col_idx, value=row.get(col, ''))
style_sheet(ws_tps, "10_Top_By_Score")

# --- 11_Secondary_Language ---
ws_seclang = wb.create_sheet(title="11_Secondary_Language")
df_seclang = df[df['LanguageGroup'] == 'SECONDARY'].sort_values(by=['BalancedScore', 'Rank_numeric', 'KEI', 'Difficulty'], ascending=[False, True, False, True])
for col_idx, col in enumerate(cols_curated, 1):
    ws_seclang.cell(row=1, column=col_idx, value=col)
for row_idx, (_, row) in enumerate(df_seclang.iterrows(), 2):
    for col_idx, col in enumerate(cols_curated, 1):
        ws_seclang.cell(row=row_idx, column=col_idx, value=row.get(col, ''))
style_sheet(ws_seclang, "11_Secondary_Language")

# --- 12_Text_Dedup_Log ---
ws_dedup = wb.create_sheet(title="12_Text_Dedup_Log")
cols_dedup = _shared_text_dedup.TEXT_DEDUP_LOG_COLUMNS
for col_idx, col in enumerate(cols_dedup, 1):
    ws_dedup.cell(row=1, column=col_idx, value=col)
if not df_dedup_log.empty:
    for row_idx, (_, row) in enumerate(df_dedup_log.iterrows(), 2):
        for col_idx, col in enumerate(cols_dedup, 1):
            ws_dedup.cell(row=row_idx, column=col_idx, value=row.get(col, ''))
style_sheet(ws_dedup, "12_Text_Dedup_Log")

# Save
print(f"Saving stylized workbook to {OUTPUT_PATH}...")
try:
    wb.save(OUTPUT_PATH)
    print("Pipeline for Game Emulator complete!")
except PermissionError:
    alt_path = OUTPUT_PATH.replace(".xlsx", "_temp.xlsx")
    print(f"WARNING: Permission denied to write to {OUTPUT_PATH} (file is likely open in another program).")
    print(f"Saving to fallback path: {alt_path}")
    wb.save(alt_path)
    print("Pipeline complete (saved to fallback path)!")
