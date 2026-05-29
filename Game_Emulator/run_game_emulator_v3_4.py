import pandas as pd
import numpy as np
import re
import os
import unicodedata
import json
import urllib.request
import urllib.parse
import html as html_lib
import ssl
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import argparse

# Resolve script and project root directories
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)

# Parse arguments
parser = argparse.ArgumentParser(description="ASO Keyword Planner for Game Emulator")
parser.add_argument("--csv", type=str, default=None, help="Path to input CSV")
parser.add_argument("--market", type=str, default="US_EN", help="Market code (e.g. US_EN)")
parser.add_argument("--output", type=str, default="", help="Path to output Excel file")
parser.add_argument("--interactive", action="store_true", help="Run interactive Web UI selector")
args, unknown = parser.parse_known_args()

INPUT_PATH = args.csv
if args.output:
    OUTPUT_PATH = args.output
else:
    # Update OUTPUT_PATH dynamically
    csv_dir = os.path.dirname(os.path.abspath(INPUT_PATH))
    OUTPUT_PATH = os.path.join(csv_dir, "Game_Emulator", f"GameEmulator_{args.market.replace('_', '-')}_Output.xlsx")

# Game Emulator configuration
config = {
    "app_id": "com.game.emulator.gb4.retro.gameboy.collection",
    "app_name": "Game Emulator: GB4 Retro Games",
    "category": "Game Emulator",
    "market": args.market,
    "platform_mode": "google_play",
    "semantic_mode": "game_emulator",
    
    "intent_core_terms": [
        "game emulator", "retro game emulator", "retro games emulator", 
        "gba emulator", "gameboy emulator", "arcade emulator", "handheld emulator",
        "gb4 emulator", "gb4 emulador", "gb4", "gba4 emulator", "gba4 emulador", "gba4"
    ],
    
    "feature_terms": [
        # Nintendo
        'gameboy', 'gba', 'gbc', 'gb', 'game boy', 'game boy advance', 'gb advance',
        'nes', 'snes', 'n64', 'super nintendo',
        '3ds', 'ds', 'nds', 'nintendo ds',
        'gamecube', 'wii', 'wii u', 'nintendo',
        # Sony
        'ps2', 'ps3', 'ps4', 'ps5', 'playstation', 'playstation 2', 'playstation 3',
        'playstation 4', 'playstation 5',
        'psp', 'psx', 'ps vita',
        # Microsoft/Sega/Others
        'xbox', 'sega', 'dreamcast', 'arcade', 'fliperama',
        'handheld', 'console', 'portable',
        # Retro/Classic indicators
        'retro', 'retrô', 'classic', 'classicos', 'clássico', 'klasik',
        '8bit', '8-bit', '16bit', '16-bit', '32bit',
        'old', 'vintage', 'nostalgic', 'nostalgia', 'nostálgia',
        'jadul', 'lawas', 'advance', 'collection'
    ],
    
    "style_terms": [
        # Nintendo IP
        'pokemon', 'pokémon', 'pokemon red', 'pokemon blue', 'pokemon fire red',
        'pokemon emerald', 'pokemon yellow', 'pokemon gold', 'pokemon silver',
        'mario', 'super mario', 'mario kart', 'zelda', 'legend of zelda',
        'metroid', 'kirby', 'donkey kong', 'star fox', 'fire emblem',
        'animal crossing', 'smash bros', 'super smash bros',
        # Sega/Sonic
        'sonic', 'sonic the hedgehog', 'tails', 'knuckles',
        'golden axe', 'streets of rage', 'shinobi',
        # Capcom/Konami/Square
        'street fighter', 'mega man', 'rockman',
        'castlevania', 'metal gear', 'silent hill',
        'final fantasy', 'dragon quest', 'chrono trigger',
        'resident evil', 'devil may cry', 'monster hunter',
        # Namco/Others
        'pacman', 'pac-man', 'galaga', 'dig dug',
        'tekken', 'soulcalibur', 'ridge racer',
        'naruto', 'dragon ball', 'dbz', 'bleach', 'one piece',
        'tetris', 'puzzle',
        # Sony IP
        'god of war', 'crash', 'crash bandicoot', 'spyro',
        'gran turismo', 'twisted metal', 'parappa',
        # Fighting/Shooters
        'mortal kombat', 'mk', 'killer instinct',
        'doom', 'quake', 'wolfenstein', 'duke nukem',
        'contra', 'metal slug', 'gunstar heroes',
        # RPG/JRPG
        'earthbound', 'mother', 'undertale',
        'persona', 'shin megami tensei', 'smt',
        'suikoden', 'wild arms', 'vagrant story',
        # Platform/Adventure
        'banjo kazooie', 'conker', 'rareware',
        'banjo-tooie', 'perfect dark',
        # Misc classic
        'bomberman', 'ice climber', 'excitebike',
        'duck hunt', 'punch out', 'kid icarus',
        'wario', 'waluigi', 'yoshi', 'luigi'
    ],
    
    "competitor_brands": [
        # Major multi-system
        'ppsspp', 'ppsspp gold', 'dolphin', 'dolphin emulator',
        'retroarch', 'retroarch emulator',
        'delta', 'delta emulator', 'delta nintendo emulator',
        'citra', 'citra emulator', 'lime3ds', 'lime3ds emulator',
        'aethersx2', 'aethersx2 emulator',
        'lemuroid', 'lemuroid emulator',
        # GBA specific
        'my boy', 'my boy emulator', 'my boy free', 'my boy gba',
        'john gba', 'john gba lite', 'john gba emulator',
        'gameboid', 'gameboid emulator',
        'gba4ios', 'gba4ios emulator',
        'vgbanext', 'vgbanext emulator',
        'gamma emulator', 'gamma game emulator',
        # NDS/3DS
        'drastic', 'drastic emulator', 'drastic ds',
        'melonDS', 'melon ds', 'desmume',
        # NES/SNES
        'snes9x', 'snes9x ex', 'zsnes',
        'nestopia', 'fceux', 'quicknes',
        'super retro 16', 'super retro 16 plus',
        # PS1/PS2
        'epsxe', 'epsxe emulator',
        'fpse', 'fpse emulator',
        'pcsx2', 'pcsx2 emulator', 'damonps2', 'damon ps2',
        'play!', 'play emulator',
        # N64
        'mupen64plus', 'mupen64', 'mupen',
        'project64', 'project 64', 'n64oid',
        # Arcade/MAME
        'mame', 'mame4droid', 'mame4ios',
        'fba', 'fbneo', 'final burn alpha', 'final burn neo',
        'kawaks',
        # All-in-one
        'classicboy', 'classicboy lite',
        'super retro 16', 'retro game boy',
        'emu', 'emu games', 'emu paradise',
        'roms', 'romsmania', 'loveroms', 'romhustler',
        # Cloud/Remote
        'netboom', 'netboom cloud gaming',
        'airconsole', 'air console',
        'starparks', 'chikii', 'chikii cloud',
        'psplay', 'psplay remote play',
        'xbplay', 'xbplay remote play',
        # Misc
        'superpsx', 'super psx',
        'bitboy', 'bitboy emulator',
        'onecast', 'onecast xbox',
        'happy chick', 'happy chick emulator',
        'pizza emulator', 'pizza boy', 'pizza boy gba',
        'easy emu', 'mock emulator', 'lucky emulator',
        'gk emulator', 'gas emulator', 'folium emulator', 'jeans emulator',
        'emulsio', 'emulator guia', 'emulator md2',
        'emulator anak permainan', 'emulator juegos pro',
        'retro game master', 'retro game hub',
        # Browser/Non-emulator
        'dolphin browser', '870 fitness'
    ],
    
    "noise_terms": [
        'game', 'games', 'gaming', 'gamer', 'gamers', 'gameplay',
        'video game', 'videogame', 'video games', 'computer game',
        'android game', 'mobile game', 'phone game', 'tablet game',
        'play', 'playing', 'player', 'fun', 'entertainment',
        'download', 'free', 'gratis', 'grátis', 'premium', 'pro', 'lite',
        'best', 'top', 'new', 'old', 'latest', 'update', 'version',
        'android', 'ios', 'iphone', 'ipad', 'phone', 'mobile', 'tablet',
        'app', 'application', 'software', 'tool', 'utility', 'program',
        'device', 'system', 'platform', 'technology', 'digital',
        'emulator', 'emulador', 'emulation', 'emu', 'emulators', 'emuladores',
        'simulador', 'simulator', 'simulate', 'virtual', 'virtual machine',
        'rom', 'roms', 'iso', 'bios', 'cheat', 'cheats', 'hack', 'mod',
        'save', 'load', 'state', 'slot', 'backup', 'restore',
        'controller', 'control', 'controle', 'kontrol', 'kontroler',
        'gamepad', 'joypad', 'joystick', 'pad', 'button', 'buttons',
        'd-pad', 'dpad', 'analog', 'stick', 'trigger', 'bumper',
        'bluetooth controller', 'wireless controller', 'usb controller',
        'remote', 'remoto', 'remote play', 'second screen',
        'cloud', 'cloud gaming', 'streaming', 'stream', 'remote',
        'geforce now', 'xbox cloud', 'playstation now', 'stadia',
        'nvidia', 'shadow', 'boosteroid', 'blacknut',
        'screen', 'display', 'monitor', 'resolution', 'fps', 'hz',
        'battery', 'storage', 'memory', 'ram', 'cpu', 'gpu',
        'speed', 'fast', 'slow', 'lag', 'latency', 'ping',
        'online', 'offline', 'multiplayer', 'coop', 'pvp', 'pve',
        'wifi', 'internet', 'network', 'connection', 'server',
        'account', 'login', 'register', 'profile', 'avatar',
        'chat', 'message', 'friend', 'social', 'community', 'forum',
        'rate', 'review', 'feedback', 'support', 'help', 'faq',
        'guide', 'tutorial', 'walkthrough', 'tips', 'tricks',
        'news', 'blog', 'update', 'patch', 'dlc', 'expansion',
        'skin', 'theme', 'wallpaper', 'icon', 'font', 'sound',
        'music', 'song', 'audio', 'soundtrack', 'ost', 'bgm',
        'record', 'recording', 'screenshot', 'capture', 'clip',
        'share', 'export', 'import', 'sync', 'backup', 'transfer',
        'the', 'a', 'an', 'and', 'or', 'but', 'for', 'with', 'without',
        'in', 'on', 'at', 'to', 'from', 'by', 'of', 'about', 'into',
        'through', 'during', 'before', 'after', 'above', 'below',
        'between', 'among', 'within', 'against', 'under', 'over',
        'good', 'great', 'awesome', 'amazing', 'excellent', 'perfect',
        'bad', 'terrible', 'awful', 'horrible', 'worst',
        'big', 'small', 'huge', 'tiny', 'large', 'mini',
        'easy', 'hard', 'difficult', 'simple', 'complex',
        'first', 'last', 'next', 'previous', 'final',
        '1990', '1995', '2000', '2005', '2010', '2015', '2020',
        '90s', '80s', '00s', 'year', 'years', 'decade',
        'one', 'two', 'three', 'four', 'five', 'six', 'seven', 'eight', 'nine', 'ten'
    ],
    
    "typo_blacklist": [
        'gretro', 'restro', 'gaem', 'gam emulador', 'imulator', 'gbã', 'gbã emulator', 'emulsio',
        '0s5', 'pspusado', 'ps4ps5', 'ps ps ps', 'ps5ps4', 'pxp', 'pps', 'pc5', 'eio', 'psdp', 'pspp', 'ppsp', 'ssip', 'ds3',
        'pintasan', '870 fitness', 'maldives', 'dolphin browser', 'restaurați poza', 'memperlambat',
        'ukuran panjang', 'seperti apa', 'seperti apa itu', 'tujuan', 'posisi', 'tercepat',
        'calculator', 'calendar', 'weather', 'clock', 'alarm', 'reminder',
        'note', 'notes', 'file manager', 'gallery', 'camera', 'video player',
        'music player', 'audio player', 'podcast', 'radio', 'news', 'magazine',
        'shopping', 'delivery', 'food', 'restaurant', 'hotel', 'travel',
        'booking', 'ticket', 'flight', 'train', 'bus', 'map', 'navigation',
        'gps', 'tracker', 'fitness', 'health', 'medical', 'diet', 'yoga',
        'meditation', 'sleep', 'water', 'step', 'calorie', 'workout',
        'education', 'learning', 'course', 'lesson', 'quiz', 'exam', 'study',
        'language', 'dictionary', 'translator', 'keyboard', 'input method',
        'launcher', 'home screen', 'lock screen', 'live wallpaper', 'widget',
        'keyboard theme', 'icon pack', 'font', 'ringtone', 'notification sound'
    ],
    
    "risky_platform_terms": [
        "ios", "iphone", "apple", "os 17", "os 18", "os17", "os18", "ipad"
    ],
    
    "balanced_weights": {
        "VolumeN": 0.20,
        "DifficultyN": 0.15,
        "KEIN": 0.15,
        "RelevancyScore": 0.30,
        "CurrentRankN": 0.10,
        "ExpansionValue": 0.10
    }
}

# --- Google Play Scraper & Competitor Profile Builder ---
def get_app_profile(config, seed_query):
    profile_path = os.path.join(os.path.dirname(OUTPUT_PATH), "App_Profile.json")
    
    # Check if cached profile is fresh (< 14 days)
    if os.path.exists(profile_path):
        try:
            with open(profile_path, "r", encoding="utf-8") as f:
                profile = json.load(f)
            last_checked = datetime.fromisoformat(profile.get("last_checked", "2000-01-01"))
            if datetime.now() - last_checked < timedelta(days=14):
                print(f"Loaded fresh App Profile from {profile_path} (Last checked: {last_checked})")
                return profile
        except Exception as e:
            print(f"Error reading profile cache: {e}. Re-fetching...")
            
    print("App Profile is missing or older than 14 days. Fetching from Google Play Store...")
    profile = {
        "app_id": config["app_id"],
        "last_checked": datetime.now().isoformat(),
        "title": "",
        "short_description": "",
        "full_description": "",
        "competitors": []
    }
    
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    
    def fetch_url(url):
        req = urllib.request.Request(
            url, 
            headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36'}
        )
        try:
            with urllib.request.urlopen(req, context=ctx, timeout=10) as response:
                return response.read().decode('utf-8')
        except Exception as e:
            print(f"Error fetching URL {url}: {e}")
            return None

    # 1. Fetch own app metadata
    own_url = f"https://play.google.com/store/apps/details?id={config['app_id']}&hl=en&gl=US"
    own_html = fetch_url(own_url)
    if own_html:
        t_match = re.search(r'<meta property="og:title" content="([^"]+)"', own_html)
        if t_match:
            title = html_lib.unescape(t_match.group(1))
            profile["title"] = re.sub(r'\s*-\s*Apps on Google Play$', '', title)
        s_match = re.search(r'<meta name="description" content="([^"]+)"', own_html)
        if not s_match:
            s_match = re.search(r'<meta name="twitter:description" content="([^"]+)"', own_html)
        if s_match:
            profile["short_description"] = html_lib.unescape(s_match.group(1))
        f_match = re.search(r'data-g-id="description"[^>]*>(.+?)</div>', own_html, re.DOTALL)
        if f_match:
            f_desc = re.sub(r'<[^>]+>', '\n', f_match.group(1))
            profile["full_description"] = html_lib.unescape(re.sub(r'\n+', '\n', f_desc).strip())
        else:
            j_match = re.search(r'\[\[null,"([^"]{50,})"\s*\]\]', own_html)
            if j_match:
                f_desc = j_match.group(1).replace('\\u003cbr\\u003e', '\n').replace('\\n', '\n')
                profile["full_description"] = html_lib.unescape(f_desc)
                
    # 2. Search for competitors
    search_url = f"https://play.google.com/store/search?q={urllib.parse.quote_plus(seed_query)}&c=apps&hl=en&gl=US"
    search_html = fetch_url(search_url)
    competitor_pids = []
    if search_html:
        raw_pids = re.findall(r'href="/store/apps/details\?id=([a-zA-Z0-9._]+)"', search_html)
        seen = set([config['app_id']])
        for pid in raw_pids:
            if pid not in seen:
                seen.add(pid)
                competitor_pids.append(pid)
                if len(competitor_pids) >= 3:
                    break
                    
    # 3. Fetch competitor details
    for pid in competitor_pids:
        comp_url = f"https://play.google.com/store/apps/details?id={pid}&hl=en&gl=US"
        comp_html = fetch_url(comp_url)
        comp_data = {
            "package_id": pid,
            "title": "",
            "short_description": "",
            "desc200": ""
        }
        if comp_html:
            t_match = re.search(r'<meta property="og:title" content="([^"]+)"', comp_html)
            if t_match:
                title = html_lib.unescape(t_match.group(1))
                comp_data["title"] = re.sub(r'\s*-\s*Apps on Google Play$', '', title)
            s_match = re.search(r'<meta name="description" content="([^"]+)"', comp_html)
            if not s_match:
                s_match = re.search(r'<meta name="twitter:description" content="([^"]+)"', comp_html)
            if s_match:
                comp_data["short_description"] = html_lib.unescape(s_match.group(1))
            f_match = re.search(r'data-g-id="description"[^>]*>(.+?)</div>', comp_html, re.DOTALL)
            f_desc_str = ""
            if f_match:
                f_desc_str = re.sub(r'<[^>]+>', '\n', f_match.group(1))
                f_desc_str = html_lib.unescape(re.sub(r'\n+', '\n', f_desc_str).strip())
            else:
                j_match = re.search(r'\[\[null,"([^"]{50,})"\s*\]\]', comp_html)
                if j_match:
                    f_desc_str = html_lib.unescape(j_match.group(1).replace('\\u003cbr\\u003e', '\n').replace('\\n', '\n'))
            comp_data["desc200"] = f_desc_str[:200]
        profile["competitors"].append(comp_data)
        
    try:
        os.makedirs(os.path.dirname(profile_path), exist_ok=True)
        with open(profile_path, "w", encoding="utf-8") as f:
            json.dump(profile, f, indent=4, ensure_ascii=False)
        print(f"Saved fresh App Profile to {profile_path}")
    except Exception as e:
        print(f"Error saving profile: {e}")
        
    return profile

# Build or load App Profile using seed query 'Game Emulator'
app_profile = get_app_profile(config, "Game Emulator")

# --- Local HTTP Server for Selection & ASO Dashboard ---
def start_interactive_server(df, config, app_profile):
    import http.server
    import socketserver
    import webbrowser
    import threading
    import time
    
    keywords_data = []
    for idx, row in df.iterrows():
        keywords_data.append({
            "Keyword": row['Keyword'],
            "Volume": int(row['Volume']),
            "Difficulty": int(row['Difficulty']),
            "KEI": float(row['KEI']),
            "RelevancyScore": float(row['RelevancyScore']),
            "BalancedScore": float(row['BalancedScore']),
            "Bucket": row.get('Bucket', 'Unclassified'),
            "Rank": str(row.get('Rank', '')),
            "CompetitorProven": row.get('CompetitorProven', 'No'),
            "ProvenDetails": row.get('ProvenDetails', '')
        })
        
    data_payload = {
        "app_name": config["app_name"],
        "app_id": config["app_id"],
        "category": config["category"],
        "competitors": app_profile.get("competitors", []),
        "keywords": keywords_data
    }
    
    server_data = {"confirmed_payload": None}
    
    class SelectionHandler(http.server.BaseHTTPRequestHandler):
        def log_message(self, format, *args):
            return
            
        def do_GET(self):
            if self.path == '/':
                self.send_response(200)
                self.send_header('Content-Type', 'text/html; charset=utf-8')
                self.end_headers()
                html_path = os.path.join(os.path.dirname(__file__), 'interactive_optimizer.html')
                with open(html_path, 'r', encoding='utf-8') as f:
                    self.wfile.write(f.read().encode('utf-8'))
            elif self.path == '/data':
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps(data_payload, ensure_ascii=False).encode('utf-8'))
            else:
                self.send_response(404)
                self.end_headers()
                
        def do_POST(self):
            if self.path == '/confirm':
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                payload = json.loads(post_data.decode('utf-8'))
                server_data["confirmed_payload"] = payload
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(b'{"status": "ok"}')
                
                def shutdown_soon():
                    time.sleep(0.5)
                    server.shutdown()
                threading.Thread(target=shutdown_soon).start()
                
    port = 8000
    for p in range(8000, 8050):
        try:
            server = socketserver.TCPServer(("", p), SelectionHandler)
            port = p
            break
        except OSError:
            continue
            
    print(f"\n[INTERACTIVE SELECTOR] Starting local server at http://localhost:{port}...")
    print("Opening browser automatically... Please select keywords and write ASO descriptions in the dashboard, then click Confirm.")
    
    webbrowser.open(f"http://localhost:{port}/")
    server.serve_forever()
    server.server_close()
    print("[INTERACTIVE SELECTOR] Configuration received. Proceeding with Excel generation...")
    return server_data["confirmed_payload"]



# Step 1: Load and Clean
print("[Step 1] Loading raw candidates...")
df_raw = pd.read_csv(INPUT_PATH, encoding="utf-8-sig")

df = pd.DataFrame()
raw_cols = ['Keyword', 'Volume', 'Difficulty', 'KEI', 'Rank', 'CurrentRank', 'RankStatus', 'MaximumReach']
for col in raw_cols:
    if col in df_raw.columns:
        df[col] = df_raw[col]
    elif col == 'CurrentRank' and 'Rank' in df_raw.columns:
        df['CurrentRank'] = df_raw['Rank']
    elif col == 'Rank' and 'CurrentRank' in df_raw.columns:
        df['Rank'] = df_raw['CurrentRank']
    elif col == 'RankStatus' and 'Rank Status' in df_raw.columns:
        df['RankStatus'] = df_raw['Rank Status']
    elif col == 'MaximumReach' and 'Maximum Reach' in df_raw.columns:
        df['MaximumReach'] = df_raw['Maximum Reach']

if 'Volume' not in df.columns:
    df['Volume'] = 0
if 'Difficulty' not in df.columns:
    df['Difficulty'] = 0
if 'KEI' not in df.columns:
    df['KEI'] = 0
if 'Rank' not in df.columns:
    df['Rank'] = 'Unranked'

df['Volume'] = pd.to_numeric(df['Volume'], errors='coerce').fillna(0).astype(int)
df['Difficulty'] = pd.to_numeric(df['Difficulty'], errors='coerce').fillna(0).astype(int)
df['KEI'] = pd.to_numeric(df['KEI'], errors='coerce').fillna(0).astype(float)

# Load Max. Volume
max_vol_col = None
for col in df_raw.columns:
    if col.lower().replace('.', '').strip() == 'max volume':
        max_vol_col = col
        break

if max_vol_col is not None:
    df['Max. Volume'] = df_raw[max_vol_col]
else:
    df['Max. Volume'] = df['Volume']

df['Max. Volume'] = pd.to_numeric(df['Max. Volume'], errors='coerce').fillna(df['Volume']).astype(int)
df['Traffic Stability'] = (df['Volume'] / df['Max. Volume']).fillna(1.0).clip(0.0, 1.0)

def get_stability_class(ratio):
    if ratio >= 0.85:
        return 'Stable'
    elif ratio >= 0.50:
        return 'Moderate'
    else:
        return 'Volatile'

df['Stability Class'] = df['Traffic Stability'].apply(get_stability_class)
df['Rank_numeric'] = pd.to_numeric(df['Rank'], errors='coerce').fillna(999)

# Singularization / Normalization
def normalize_text(text):
    if not isinstance(text, str):
        return ""
    text = text.lower().strip()
    text = "".join(c for c in unicodedata.normalize('NFD', text) if unicodedata.category(c) != 'Mn')
    text = re.sub(r'[-_]', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()
    words = text.split()
    normalized_words = []
    for w in words:
        if w.endswith('s') and len(w) > 3:
            if w in ['themes', 'widgets', 'panels', 'settings', 'shortcuts', 'icons', 'styles', 'games', 'emulators', 'consoles']:
                w = w[:-1]
        normalized_words.append(w)
    return " ".join(normalized_words)

df['keyword_normalized'] = df['Keyword'].apply(normalize_text)

# Step 2: Language Detection
print("[Step 2] Language classification...")

try:
    from langdetect import detect_langs, DetectorFactory
    DetectorFactory.seed = 0
    HAS_LANGDETECT = True
except ImportError:
    HAS_LANGDETECT = False

def load_english_vocab():
    vocab = set()
    path = os.path.join(PROJECT_ROOT, "Docs_and_Templates", "english_words_10k.txt")
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                for line in f:
                    vocab.add(line.strip().lower())
        except Exception:
            pass
    return vocab

english_vocab = load_english_vocab()

def get_root_word(w):
    w = w.lower()
    if len(w) > 4:
        if w.endswith('ing'):
            return w[:-3]
        if w.endswith('ed'):
            return w[:-2]
        if w.endswith('es'):
            return w[:-2]
        if w.endswith('er'):
            return w[:-2]
        if w.endswith('est'):
            return w[:-3]
    if len(w) > 3:
        if w.endswith('s'):
            return w[:-1]
    return w

def detect_keyword_language(kw, market_lang, config):
    kw_lower = str(kw).lower().strip()
    if not kw_lower:
        return market_lang.split("_")[1].lower() if "_" in market_lang else "en", 'PRIMARY'
        
    primary_lang = market_lang.split("_")[1].lower() if "_" in market_lang else "en"
    
    # Whitelist of common English words/ASO-related terms that might not be in 10k list
    eng_words = {
        'ar', 'fyp', 'app', 'apps', 'free', 'download', 'android', 'new', 'best', 'top',
        'doggy', 'dogy', 'shrek', 'diy', 'pro', 'lite', 'tiktok', 'snapchat', 'instagram',
        'youtube', 'facebook', 'whatsapp', 'messenger', 'pinterest', 'google', 'play',
        '3d', 'arstudio', 'augmented', 'virtual', 'scanning', 'scanner', 'doge', 'da', 'doin'
    }
    
    # Add words from configuration terms
    for key in ['intent_core_words', 'intent_core_terms', 'feature_terms', 'style_terms', 'visual_terms', 'noise_terms']:
        if key in config:
            for term in config[key]:
                for w in str(term).lower().split():
                    eng_words.add(w)
                    
    # Clean words in keyword
    words = [re.sub(r'[^a-z0-9]', '', w) for w in kw_lower.split()]
    words = [w for w in words if w]
    
    if not words:
        return primary_lang, 'PRIMARY'
        
    # If all words in the keyword are in the English whitelist OR the 10k vocabulary, it's Primary!
    is_primary = True
    for w in words:
        root = get_root_word(w)
        if w not in eng_words and w not in english_vocab and root not in eng_words and root not in english_vocab:
            is_primary = False
            break
            
    if is_primary:
        return primary_lang, 'PRIMARY'
        
    # If not all words are in the whitelists, use langdetect if available
    if HAS_LANGDETECT:
        try:
            langs = detect_langs(kw_lower)
            best_lang = langs[0].lang
            prob = langs[0].prob
            
            if best_lang == primary_lang:
                return best_lang, 'PRIMARY'
                
            if prob > 0.7:
                # langdetect tends to misclassify short English words as Norwegian or Danish
                if best_lang in ['no', 'da'] and primary_lang == 'en':
                    return 'en', 'PRIMARY'
                
                # Check for secondary languages from policy
                secondary_langs = [l.split('-')[0].lower() for l in config.get('market_language_policy', {}).get('secondary_languages', [])]
                if best_lang in secondary_langs:
                    return best_lang, 'SECONDARY'
                    
                return best_lang, 'FOREIGN'
        except Exception:
            pass
            
    return primary_lang, 'PRIMARY'

# Populate language columns in df
detected_langs = []
lang_groups = []

for idx, row in df.iterrows():
    # If columns already exist in raw data, use them
    raw_lang = df_raw.loc[idx, 'DetectedLanguage'] if 'DetectedLanguage' in df_raw.columns else None
    raw_group = df_raw.loc[idx, 'LanguageGroup'] if 'LanguageGroup' in df_raw.columns else None
    
    if pd.notna(raw_lang) and pd.notna(raw_group):
        detected_langs.append(raw_lang)
        lang_groups.append(raw_group)
    else:
        lang, group = detect_keyword_language(row['Keyword'], config.get('market', 'US_EN'), config)
        detected_langs.append(lang)
        lang_groups.append(group)

df['DetectedLanguage'] = detected_langs
df['LanguageGroup'] = lang_groups

# Step 3: Hard filters
print("[Step 3] Hard filters...")
df['is_competitor'] = df['Keyword'].apply(
    lambda x: any(re.search(r'\b' + re.escape(brand.lower()) + r'\b', str(x).lower()) 
                  for brand in config['competitor_brands'])
)
df['is_typo'] = df['Keyword'].apply(
    lambda x: any(re.search(r'\b' + re.escape(typo.lower()) + r'\b', str(x).lower()) 
                  for typo in config['typo_blacklist'])
)
df['is_irrelevant'] = False

def is_noise_only(kw, config):
    kw_lower = str(kw).lower().strip()
    has_core = any(term.lower() in kw_lower for term in config['intent_core_terms'])
    if has_core:
        return False
    words = kw_lower.split()
    if len(words) > 1:
        has_feat = any(f.lower() in kw_lower for f in config['feature_terms'])
        has_sty = any(s.lower() in kw_lower for s in config['style_terms'])
        if has_feat or has_sty:
            return False
    noise_words = set(t.lower() for t in config['noise_terms'])
    if all(w in noise_words for w in words):
        return True
    if len(words) == 1:
        w = words[0]
        if w in noise_words:
            return True
    return False

df['is_noise'] = df['Keyword'].apply(lambda x: is_noise_only(x, config))

# Naturalness Filter
print("[Step 4] Naturalness checking...")
def check_naturalness(kw, config):
    kw_lower = str(kw).lower()
    words = kw_lower.split()
    if len(words) == 0:
        return 'UNNATURAL', 'Empty keyword'
    word_counts = {}
    for w in words:
        word_counts[w] = word_counts.get(w, 0) + 1
    max_repeat = max(word_counts.values()) if word_counts else 0
    ratio = max_repeat / len(words) if words else 0
    if ratio > 0.5 and len(words) > 2:
        return 'STUFFING', 'Too many repeated words'
    has_core = any(term.lower() in kw_lower for term in config['intent_core_terms'])
    if len(words) > 6:
        if has_core:
            pass
        else:
            return 'TOO_LONG', f'Keyword has too many words ({len(words)})'
    grammar_patterns = [
        r"\b(game game|emulator emulator|play play)\b",
        r"\b(what is|how to|why do|when is|where is)\b"
    ]
    for pat in grammar_patterns:
        if re.search(pat, kw_lower):
            return 'UNNATURAL', 'Fails structural validation'
    for char in kw_lower:
        if ord(char) > 127 and char not in 'áéíóúüñ¿¡íóú':
            return 'LANGUAGE_BLEED', 'Foreign script character detected'
    return 'OK', 'Natural enough for keyword research'

naturalness = df['Keyword'].apply(lambda x: check_naturalness(x, config))
df['NaturalnessFlag'] = [n[0] for n in naturalness]
df['NaturalnessReason'] = [n[1] for n in naturalness]

# Scoring Logic (Game Emulator spec formulas)
print("[Step 5] Relevancy Scoring...")

# Pre-calculate competitor-proven keywords and score boost
def normalize_match_text(text):
    if not isinstance(text, str):
        return ""
    text = text.lower()
    text = re.sub(r'[^a-z0-9\s]', ' ', text)
    return re.sub(r'\s+', ' ', text).strip()

def check_keyword_in_text(kw_norm, text_norm):
    if not kw_norm or not text_norm:
        return False
    pattern = r'\b' + re.escape(kw_norm) + r'\b'
    return bool(re.search(pattern, text_norm))

competitor_proven_list = []
proven_details_list = []
competitor_boost_list = []

for idx, row in df.iterrows():
    kw_norm = normalize_match_text(row['Keyword'])
    
    comp_title_matches = []
    comp_short_matches = []
    comp_desc_matches = []
    
    for comp in app_profile.get("competitors", []):
        comp_title_norm = normalize_match_text(comp.get("title", ""))
        comp_short_norm = normalize_match_text(comp.get("short_description", ""))
        comp_desc_norm = normalize_match_text(comp.get("desc200", ""))
        comp_name = comp.get("title", comp.get("package_id", ""))
        
        if check_keyword_in_text(kw_norm, comp_title_norm):
            comp_title_matches.append(comp_name)
        if check_keyword_in_text(kw_norm, comp_short_norm):
            comp_short_matches.append(comp_name)
        if check_keyword_in_text(kw_norm, comp_desc_norm):
            comp_desc_matches.append(comp_name)
            
    boost = 0.0
    details = []
    if comp_title_matches:
        boost += 0.12
        details.append(f"Title: {', '.join(comp_title_matches)}")
    if comp_short_matches:
        boost += 0.08
        details.append(f"ShortDesc: {', '.join(comp_short_matches)}")
    if comp_desc_matches:
        boost += 0.05
        details.append(f"Desc200: {', '.join(comp_desc_matches)}")
        
    boost = min(boost, 0.15)
    
    if boost > 0:
        competitor_proven_list.append("Yes")
        proven_details_list.append("; ".join(details))
        competitor_boost_list.append(boost)
    else:
        competitor_proven_list.append("No")
        proven_details_list.append("")
        competitor_boost_list.append(0.0)

df['CompetitorProven'] = competitor_proven_list
df['ProvenDetails'] = proven_details_list
df['CompetitorBoost'] = competitor_boost_list

def calculate_relevancy(row, config):
    kw = str(row['Keyword']).lower()
    score = 0.3  # Base score
    
    # +0.40: Core emulator intent
    if 'emulator' in kw or 'emulador' in kw:
        score += 0.40
        
    # +0.15: Specific Console/Feature
    feature_match = [
        'gameboy', 'gba', 'gbc', 'nes', 'snes', 'n64',
        'ps2', 'psp', '3ds', 'sega', 'arcade', 'gamecube', 'dreamcast', 'fliperama',
        'gb4', 'gba4'
    ]
    if any(c in kw for c in feature_match):
        score += 0.15
        
    # +0.10: Retro/Classic style
    retro_match = [
        'retro', 'retrô', 'classic', 'klasik', 'clássico',
        '8bit', '8-bit', '16bit', '16-bit',
        'old', 'nostalgic', 'nostalgia', 'nostálgia', 'vintage', 'jadul', 'lawas'
    ]
    if any(r in kw for r in retro_match):
        score += 0.10
        
    # +0.05: Game title IP
    ip_match = [
        'pokemon', 'pokémon', 'mario', 'zelda', 'naruto', 'sonic',
        'tetris', 'pacman', 'metroid', 'street fighter', 'crash',
        'god of war', 'gta', 'tekken', 'wwe', 'super smash bros'
    ]
    if any(g in kw for g in ip_match):
        score += 0.05
        
    # Penalties
    if row['is_competitor']:
        score -= 0.25
    if row['is_noise']:
        score -= 0.20
    if row['LanguageGroup'] == 'FOREIGN':
        score -= 0.35
    if row['NaturalnessFlag'] != 'OK':
        score -= 0.35
    # Competitor Boost
    score += row.get('CompetitorBoost', 0.0)
        
    return max(0.0, min(1.0, score))

df['RelevancyScore'] = df.apply(lambda r: calculate_relevancy(r, config), axis=1)

# Normalization & Balanced Score
print("[Step 6] Balanced Score Normalization...")
max_vol = df['Max. Volume'].max()
max_kei = df['KEI'].max()

df['VolumeN'] = np.log1p(df['Max. Volume']) / np.log1p(max_vol) if max_vol > 0 else 0
df['DifficultyN'] = 1.0 - (df['Difficulty'].clip(0, 100) / 100.0)
df['KEIN'] = np.log1p(df['KEI']) / np.log1p(max_kei) if max_kei > 0 else 0

def get_rank_n(rank_val):
    try:
        if pd.isna(rank_val) or rank_val in ['Unranked', 'unranked', 'empty', '']:
            return 0.0
        r = float(rank_val)
        if r <= 10: return 1.0
        elif r <= 30: return 0.8
        elif r <= 50: return 0.6
        elif r <= 100: return 0.3
        else: return 0.1
    except:
        return 0.0
df['CurrentRankN'] = df['Rank'].apply(get_rank_n)

def calculate_expansion(row):
    kw = str(row['Keyword']).lower()
    words = kw.split()
    n = len(words)
    if n == 1:
        score = 0.9
    elif n == 2:
        score = 0.7
    elif n == 3:
        score = 0.5
    else:
        score = 0.3
    if 'emulator' in kw or 'emulador' in kw:
        score += 0.1
    if row['is_competitor']:
        score = 0.1
    # Check if keyword has style/game IP
    has_style = any(re.search(r'\b' + re.escape(s.lower()) + r'\b', kw) for s in config['style_terms'])
    if has_style:
        score = min(score, 0.3)
    return max(0.0, min(1.0, score))

df['ExpansionValue'] = df.apply(calculate_expansion, axis=1)

bw = config['balanced_weights']
df['BalancedScore'] = (
    bw['VolumeN'] * df['VolumeN'] +
    bw['DifficultyN'] * df['DifficultyN'] +
    bw['KEIN'] * df['KEIN'] +
    bw['RelevancyScore'] * df['RelevancyScore'] +
    bw['CurrentRankN'] * df['CurrentRankN'] +
    bw['ExpansionValue'] * df['ExpansionValue']
)

# Add a small bonus based on LanguageGroup: +0.02 for PRIMARY, +0.01 for SECONDARY
def get_language_bonus(row):
    lg = str(row.get('LanguageGroup', 'PRIMARY')).upper()
    if lg == 'PRIMARY':
        return 0.02
    elif lg == 'SECONDARY':
        return 0.01
    return 0.0

df['BalancedScore'] = (df['BalancedScore'] + df.apply(get_language_bonus, axis=1)).round(4)
df['RelevancyScore'] = df['RelevancyScore'].round(4)

# Bucket classification (Game Emulator Mode)
print("[Step 7] Bucket classification...")
def classify_keyword(row, config):
    kw = str(row['Keyword']).lower()
    
    # Hard drops
    if row['is_competitor']:
        return 'Dropped', 'competitor_brand', 'Dropped: Competitor brand'
    if row['is_typo']:
        return 'Dropped', 'typo_truncated_broken', 'Dropped: Typo, truncated, or broken'
    if row['is_noise']:
        return 'Dropped', 'noise_only', 'Dropped: Noise-only generic term'
    if row['NaturalnessFlag'] != 'OK':
        return 'Dropped', 'unnatural', f"Dropped: Unnatural phrase ({row['NaturalnessReason']})"
        
    # Language Policy
    if row['LanguageGroup'] == 'FOREIGN':
        return 'Language Mismatch Audit', 'foreign_language_mismatch', 'Foreign language mismatch'
    if row['LanguageGroup'] in ['MIXED', 'UNKNOWN']:
        return 'Manual Review', 'manual_review', 'Mixed or unknown language'
    if row['LanguageGroup'] == 'SECONDARY':
        return 'Consider Keywords', 'secondary_language_handling', 'Spanish keyword for US_EN (Secondary language)'
        
    # Platform Risk
    has_platform_risk = any(term in kw for term in config['risky_platform_terms'])
    if has_platform_risk:
        return 'Consider Keywords', 'platform_style_risk', 'Platform-style risk; keep for review only'
        
    # Check console/feature and game titles
    has_core = any(term in kw for term in config['intent_core_terms'])
    has_feature = any(re.search(r'\b' + re.escape(f.lower()) + r'\b', kw) for f in config['feature_terms'])
    has_style = any(re.search(r'\b' + re.escape(s.lower()) + r'\b', kw) for s in config['style_terms'])
    
    if has_core:
        return 'Core Intent Final', 'core_intent_final', 'Strong core game emulator search intent'
        
    if has_style:
        generic_game_terms = ["retro games", "classic games", "gba games", "arcade games", "jogos retrô", "jogos gba", "game emulator"]
        if any(term in kw for term in generic_game_terms):
            return 'Broad Expansion', 'broad_expansion', 'Generic game/emulator variant'
        else:
            return 'Game Keywords', 'game_keywords', 'Game Title/Franchise candidate (Research Only)'
            
    if has_feature:
        return 'System Keywords', 'system_keywords', 'System/Console candidate'
        
    if row['RelevancyScore'] < 0.45:
        return 'Dropped', 'dropped', 'Dropped: Weak app intent after scoring'
        
    return 'Broad Expansion', 'broad_expansion', 'Moderately relevant emulator expansion'

classifications = df.apply(lambda r: classify_keyword(r, config), axis=1)
df['Bucket'] = [c[0] for c in classifications]
df['DecisionRule'] = [c[1] for c in classifications]
df['Reason'] = [c[2] for c in classifications]

# Shortlist building & duplicate checking
print("[Step 8] Near-Duplicate Cleanup & Shortlist building...")
def build_shortlist(df_all, config):
    df_sorted = df_all.sort_values(by=['BalancedScore', 'Rank_numeric', 'KEI', 'Difficulty'], ascending=[False, True, False, True]).copy()
    selected_core, selected_broad, selected_consider = [], [], []
    selected_normalized, selected_tokens = set(), set()
    dedup_log = []
    
    def check_duplicate(kw, original_bucket):
        norm = normalize_text(kw)
        tokens = " ".join(sorted(norm.split()))
        if not norm:
            return True, "Empty normalized keyword", ""
        if norm in selected_normalized:
            all_selected = selected_core + selected_broad + selected_consider
            kept_kw = ""
            for item in all_selected:
                if normalize_text(item['Keyword']) == norm:
                    kept_kw = item['Keyword']
                    break
            return True, f"Exact normalized duplicate of '{kept_kw}'", kept_kw
        if tokens in selected_tokens:
            all_selected = selected_core + selected_broad + selected_consider
            kept_kw = ""
            for item in all_selected:
                t = " ".join(sorted(normalize_text(item['Keyword']).split()))
                if t == tokens:
                    kept_kw = item['Keyword']
                    break
            return True, f"Same normalized token set as '{kept_kw}'", kept_kw
        return False, "", ""
        
    def add_to_shortlist(item, section):
        norm = normalize_text(item['Keyword'])
        tokens = " ".join(sorted(norm.split()))
        selected_normalized.add(norm)
        selected_tokens.add(tokens)
        entry = item.to_dict()
        entry['Section'] = section
        entry['QuotaStatus'] = 'EXACT'
        entry['FillSource'] = ''
        entry['FillReason'] = ''
        return entry

    # Core (Quota: 25)
    core_candidates = df_sorted[df_sorted['Bucket'] == 'Core Intent Final']
    for _, row in core_candidates.iterrows():
        if len(selected_core) >= 25:
            break
        is_dup, reason, kept_kw = check_duplicate(row['Keyword'], 'Core Intent Final')
        if is_dup:
            dedup_log.append({
                'Table': '01_Main_Keyword_Shortlist', 'RemovedKeyword': row['Keyword'],
                'OriginalSection': 'Core Intent Final', 'KeptKeyword': kept_kw,
                'DedupReason': reason, 'BalancedScore': row['BalancedScore'],
                'Note': 'Keyword remains in All Candidates pool'
            })
        else:
            selected_core.append(add_to_shortlist(row, 'Core Intent Final'))
            
    # Core Fallback
    if len(selected_core) < 25:
        fallback_candidates = df_sorted[df_sorted['Bucket'].isin(['System Keywords', 'Broad Expansion'])]
        for _, row in fallback_candidates.iterrows():
            if len(selected_core) >= 25:
                break
            norm = normalize_text(row['Keyword'])
            if norm in selected_normalized:
                continue
            if row['RelevancyScore'] < 0.45:
                continue
            is_dup, reason, kept_kw = check_duplicate(row['Keyword'], 'Core Intent Final (Fallback)')
            if is_dup:
                dedup_log.append({
                    'Table': '01_Main_Keyword_Shortlist', 'RemovedKeyword': row['Keyword'],
                    'OriginalSection': 'Core Intent Final (Fallback)', 'KeptKeyword': kept_kw,
                    'DedupReason': reason, 'BalancedScore': row['BalancedScore'],
                    'Note': 'Keyword remains in All Candidates pool'
                })
            else:
                entry = add_to_shortlist(row, 'Core Intent Final')
                entry['QuotaStatus'] = 'FILLED'
                entry['FillSource'] = row['Bucket']
                entry['FillReason'] = 'Core Intent Quota Fallback'
                selected_core.append(entry)

    # Broad (Quota: 5)
    broad_candidates = df_sorted[df_sorted['Bucket'] == 'Broad Expansion']
    for _, row in broad_candidates.iterrows():
        if len(selected_broad) >= 5:
            break
        is_dup, reason, kept_kw = check_duplicate(row['Keyword'], 'Broad Expansion')
        if is_dup:
            dedup_log.append({
                'Table': '01_Main_Keyword_Shortlist', 'OriginalSection': 'Broad Expansion',
                'RemovedKeyword': row['Keyword'], 'KeptKeyword': kept_kw,
                'DedupReason': reason, 'BalancedScore': row['BalancedScore'],
                'Note': 'Keyword remains in All Candidates pool'
            })
        else:
            selected_broad.append(add_to_shortlist(row, 'Broad Expansion'))
            
    # Broad Fallback
    if len(selected_broad) < 5:
        fallback_candidates = df_sorted[df_sorted['Bucket'].isin(['System Keywords', 'Game Keywords'])]
        for _, row in fallback_candidates.iterrows():
            if len(selected_broad) >= 5:
                break
            norm = normalize_text(row['Keyword'])
            if norm in selected_normalized:
                continue
            if row['RelevancyScore'] < 0.45:
                continue
            is_dup, reason, kept_kw = check_duplicate(row['Keyword'], 'Broad Expansion (Fallback)')
            if is_dup:
                dedup_log.append({
                    'Table': '01_Main_Keyword_Shortlist', 'RemovedKeyword': row['Keyword'],
                    'OriginalSection': 'Broad Expansion (Fallback)', 'KeptKeyword': kept_kw,
                    'DedupReason': reason, 'BalancedScore': row['BalancedScore'],
                    'Note': 'Keyword remains in All Candidates pool'
                })
            else:
                entry = add_to_shortlist(row, 'Broad Expansion')
                entry['QuotaStatus'] = 'FILLED'
                entry['FillSource'] = row['Bucket']
                entry['FillReason'] = 'Broad Expansion Quota Fallback'
                selected_broad.append(entry)

    # Consider (Quota: 10)
    consider_candidates = df_sorted[df_sorted['Bucket'] == 'Consider Keywords']
    for _, row in consider_candidates.iterrows():
        if len(selected_consider) >= 10:
            break
        is_dup, reason, kept_kw = check_duplicate(row['Keyword'], 'Consider Keywords')
        if is_dup:
            dedup_log.append({
                'Table': '01_Main_Keyword_Shortlist', 'RemovedKeyword': row['Keyword'],
                'OriginalSection': 'Consider Keywords', 'KeptKeyword': kept_kw,
                'DedupReason': reason, 'BalancedScore': row['BalancedScore'],
                'Note': 'Keyword remains in All Candidates pool'
            })
        else:
            selected_consider.append(add_to_shortlist(row, 'Consider Keywords'))
            
    # Consider Fallback
    if len(selected_consider) < 10:
        selected_kws = {item['Keyword'].lower() for item in selected_core + selected_broad}
        missed_opps = df_sorted[df_sorted['Bucket'].isin(['Core Intent Final', 'Broad Expansion']) & 
                                (~df_sorted['Keyword'].str.lower().isin(selected_kws))]
        for _, row in missed_opps.iterrows():
            if len(selected_consider) >= 10:
                break
            norm = normalize_text(row['Keyword'])
            if norm in selected_normalized:
                continue
            if row['RelevancyScore'] < 0.45:
                continue
            is_dup, reason, kept_kw = check_duplicate(row['Keyword'], 'Consider Keywords (Fallback)')
            if is_dup:
                dedup_log.append({
                    'Table': '01_Main_Keyword_Shortlist', 'RemovedKeyword': row['Keyword'],
                    'OriginalSection': 'Consider Keywords (Fallback)', 'KeptKeyword': kept_kw,
                    'DedupReason': reason, 'BalancedScore': row['BalancedScore'],
                    'Note': 'Keyword remains in All Candidates pool'
                })
            else:
                entry = add_to_shortlist(row, 'Consider Keywords')
                entry['QuotaStatus'] = 'FILLED'
                entry['FillSource'] = row['Bucket']
                entry['FillReason'] = 'Consider Keywords Quota Fallback (Missed Opportunity)'
                selected_consider.append(entry)

    return selected_core, selected_broad, selected_consider, dedup_log

selected_core, selected_broad, selected_consider, dedup_log_list = build_shortlist(df, config)

def build_curated_sheet(df_all, bucket_name, sheet_name):
    df_sorted = df_all[df_all['Bucket'] == bucket_name].sort_values(by=['BalancedScore', 'Rank_numeric', 'KEI', 'Difficulty'], ascending=[False, True, False, True]).copy()
    selected = []
    selected_normalized, selected_tokens = set(), set()
    dedup_entries = []
    
    for _, row in df_sorted.iterrows():
        if len(selected) >= 30:
            break
        norm = normalize_text(row['Keyword'])
        tokens = " ".join(sorted(norm.split()))
        
        is_dup = False
        reason = ""
        kept_kw = ""
        
        if norm in selected_normalized:
            is_dup = True
            for item in selected:
                if normalize_text(item['Keyword']) == norm:
                    kept_kw = item['Keyword']
                    break
            reason = f"Exact normalized duplicate of '{kept_kw}'"
        elif tokens in selected_tokens:
            is_dup = True
            for item in selected:
                t = " ".join(sorted(normalize_text(item['Keyword']).split()))
                if t == tokens:
                    kept_kw = item['Keyword']
                    break
            reason = f"Same normalized token set as '{kept_kw}'"
            
        if is_dup:
            dedup_entries.append({
                'Table': sheet_name, 'RemovedKeyword': row['Keyword'],
                'OriginalSection': bucket_name, 'KeptKeyword': kept_kw,
                'DedupReason': reason, 'BalancedScore': row['BalancedScore'],
                'Note': 'Keyword remains in All Candidates pool'
            })
        else:
            selected_normalized.add(norm)
            selected_tokens.add(tokens)
            entry = row.to_dict()
            entry['Section'] = bucket_name
            entry['QuotaStatus'] = 'EXACT'
            entry['FillSource'] = ''
            entry['FillReason'] = ''
            selected.append(entry)
            
    return selected, dedup_entries

# System Keywords (capped <=30, no fallback fill)
selected_feature, dedup_feat = build_curated_sheet(df, 'System Keywords', '02_System_Keywords')
# Game Keywords (capped <=30, no fallback fill)
selected_style, dedup_style = build_curated_sheet(df, 'Game Keywords', '03_Game_Keywords')

dedup_log_list.extend(dedup_feat)
dedup_log_list.extend(dedup_style)
df_dedup_log = pd.DataFrame(dedup_log_list)

# Metadata assignment
print("[Step 9] Metadata slot assignment...")

# Start Interactive Selector Dashboard
selections_file = os.path.join(os.path.dirname(__file__), "selected_keywords.json")

confirmed_selection = None

if os.path.exists(selections_file):
    print(f"\n[Step 9] Found existing keyword selections in {selections_file}. Loading...")
    with open(selections_file, "r", encoding="utf-8") as f:
        confirmed_selection = json.load(f)
else:
    if args.interactive:
        confirmed_selection = start_interactive_server(df, config, app_profile)
        if confirmed_selection:
            with open(selections_file, "w", encoding="utf-8") as f:
                json.dump(confirmed_selection, f, indent=4, ensure_ascii=False)
                
            print("\n" + "="*50)
            print("[SELECTION_CONFIRMED] Keyword selections successfully saved!")
            print("Selections saved to:", selections_file)
            print("="*50 + "\n")
    else:
        print("\n[Step 9] Run in non-interactive/headless mode. Using defaults from shortlist logic.")

if confirmed_selection:
    user_core = confirmed_selection.get("core_keywords", [])
    user_secondary = confirmed_selection.get("secondary_keywords", [])
    user_feature = confirmed_selection.get("feature_keywords", [])
    user_style = confirmed_selection.get("style_keywords", [])
    
    df_lookup = df.set_index('Keyword')
    
    selected_core = []
    for kw in user_core:
        if kw in df_lookup.index:
            row = df_lookup.loc[kw]
            if isinstance(row, pd.DataFrame):
                row = row.iloc[0]
            entry = row.to_dict()
            entry['Keyword'] = kw
            entry['Section'] = 'Core Intent Final'
            entry['QuotaStatus'] = 'EXACT'
            selected_core.append(entry)
            
    selected_broad = []
    selected_consider = []
    for idx, kw in enumerate(user_secondary):
        if kw in df_lookup.index:
            row = df_lookup.loc[kw]
            if isinstance(row, pd.DataFrame):
                row = row.iloc[0]
            entry = row.to_dict()
            entry['Keyword'] = kw
            if idx < 5:
                entry['Section'] = 'Broad Expansion'
                selected_broad.append(entry)
            else:
                entry['Section'] = 'Consider Keywords'
                selected_consider.append(entry)
                
    selected_feature = []
    for kw in user_feature:
        if kw in df_lookup.index:
            row = df_lookup.loc[kw]
            if isinstance(row, pd.DataFrame):
                row = row.iloc[0]
            entry = row.to_dict()
            entry['Keyword'] = kw
            entry['Section'] = 'System Keywords'
            selected_feature.append(entry)
            
    selected_style = []
    for kw in user_style:
        if kw in df_lookup.index:
            row = df_lookup.loc[kw]
            if isinstance(row, pd.DataFrame):
                row = row.iloc[0]
            entry = row.to_dict()
            entry['Keyword'] = kw
            entry['Section'] = 'Game Keywords'
            selected_style.append(entry)
            
    config["app_title_draft"] = confirmed_selection.get("title", "")
    config["short_desc_draft"] = confirmed_selection.get("short_description", "")
    config["full_desc_draft"] = confirmed_selection.get("full_description", "")

all_shortlist = selected_core + selected_broad + selected_consider
for idx, entry in enumerate(all_shortlist):
    sec = entry['Section']
    if sec == 'Core Intent Final':
        if idx < 2:
            entry['WhereToUse'] = '🏷️ Title'
        elif idx < 9:
            entry['WhereToUse'] = '📱 Short Description'
        else:
            entry['WhereToUse'] = '📄 Full Description'
    elif sec == 'Broad Expansion':
        entry['WhereToUse'] = '📄 Full Description'
    else:
        entry['WhereToUse'] = '🔍 Consider / Research Only'

df_shortlist = pd.DataFrame(all_shortlist)

# Stylization & Export
print("[Step 10] Exporting to stylized Excel Workbook...")
wb = Workbook()
wb.remove(wb.active)

def style_sheet(ws, title, is_report=False):
    ws.views.sheetView[0].showGridLines = True
    if not is_report and ws.max_row > 1:
        ws.freeze_panes = 'A2'
        
    navy_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    white_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin = Side(border_style="thin", color="D3D3D3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    if not is_report:
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = navy_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border
                
                col_name = ws.cell(row=1, column=col_idx).value
                if col_name in ['Volume', 'Max. Volume', 'Difficulty', 'Rank', 'MaximumReach']:
                    try: cell.value = int(float(cell.value))
                    except: pass
                elif col_name in ['BalancedScore', 'RelevancyScore', 'KEI', 'VolumeN', 'DifficultyN', 'KEIN', 'CurrentRankN', 'OpportunityRankGap', 'ExpansionValue', 'Traffic Stability']:
                    try:
                        cell.value = round(float(cell.value), 4)
                        if col_name == 'Traffic Stability':
                            cell.number_format = '0.00%'
                        else:
                            cell.number_format = '0.0000'
                    except: pass
                        
    for col in ws.columns:
        max_len = 0
        for cell in col:
            val_str = str(cell.value or '')
            if '\n' in val_str:
                val_str = max(val_str.split('\n'), key=len)
            max_len = max(max_len, len(val_str))
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

# --- 00_README_CONFIG ---
ws_readme = wb.create_sheet(title="00_README_CONFIG")
ws_readme.views.sheetView[0].showGridLines = True
ws_readme.cell(row=1, column=1, value="ASO Keyword Planner v3.4 - Configuration Summary").font = Font(size=14, bold=True)
configs = [
    ("Pipeline Version", "ASO Keyword Planner v3.4"),
    ("App Name", config["app_name"]),
    ("App ID", config["app_id"]),
    ("Category", config["category"]),
    ("Market", config["market"]),
    ("Platform", config["platform_mode"]),
    ("Run Date", pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"))
]
if "app_title_draft" in config:
    configs.append(("Custom Title Draft", config["app_title_draft"]))
if "short_desc_draft" in config:
    configs.append(("Custom Short Desc Draft", config["short_desc_draft"]))
if "full_desc_draft" in config:
    configs.append(("Custom Full Desc Draft", config["full_desc_draft"]))
for idx, (lbl, val) in enumerate(configs, 3):
    ws_readme.cell(row=idx, column=1, value=lbl).font = Font(bold=True)
    ws_readme.cell(row=idx, column=2, value=val)
    ws_readme.cell(row=idx, column=1).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws_readme.cell(row=idx, column=2).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
ws_readme.column_dimensions['A'].width = 25
ws_readme.column_dimensions['B'].width = 80

# --- 01_Main_Keyword_Shortlist ---
ws_shortlist = wb.create_sheet(title="01_Main_Keyword_Shortlist")
cols_shortlist = ['Keyword', 'Volume', 'Max. Volume', 'Difficulty', 'KEI', 'Rank', 'Traffic Stability', 'Stability Class', 'Section', 'BalancedScore', 'RelevancyScore', 
                  'CompetitorProven', 'ProvenDetails', 'DetectedLanguage', 'LanguageGroup', 'NaturalnessFlag', 'WhereToUse', 'QuotaStatus', 'FillSource', 'FillReason', 'Reason']
for col_idx, col in enumerate(cols_shortlist, 1):
    ws_shortlist.cell(row=1, column=col_idx, value=col)
for row_idx, entry in enumerate(all_shortlist, 2):
    for col_idx, col in enumerate(cols_shortlist, 1):
        ws_shortlist.cell(row=row_idx, column=col_idx, value=entry.get(col, ''))
style_sheet(ws_shortlist, "01_Main_Keyword_Shortlist")

# --- 02_System_Keywords ---
ws_system = wb.create_sheet(title="02_System_Keywords")
cols_curated = ['Keyword', 'Volume', 'Max. Volume', 'Difficulty', 'KEI', 'Rank', 'Traffic Stability', 'Stability Class', 'Section', 'BalancedScore', 'RelevancyScore', 'Reason']
for col_idx, col in enumerate(cols_curated, 1):
    ws_system.cell(row=1, column=col_idx, value=col)
for row_idx, entry in enumerate(selected_feature, 2):
    for col_idx, col in enumerate(cols_curated, 1):
        ws_system.cell(row=row_idx, column=col_idx, value=entry.get(col, ''))
style_sheet(ws_system, "02_System_Keywords")

# --- 03_Game_Keywords ---
ws_game = wb.create_sheet(title="03_Game_Keywords")
for col_idx, col in enumerate(cols_curated, 1):
    ws_game.cell(row=1, column=col_idx, value=col)
for row_idx, entry in enumerate(selected_style, 2):
    for col_idx, col in enumerate(cols_curated, 1):
        ws_game.cell(row=row_idx, column=col_idx, value=entry.get(col, ''))
style_sheet(ws_game, "03_Game_Keywords")

# --- 04_Dropped_Audit ---
ws_dropped = wb.create_sheet(title="04_Dropped_Audit")
df_dropped = df[df['Bucket'] == 'Dropped'].sort_values(by=['BalancedScore', 'Rank_numeric', 'KEI', 'Difficulty'], ascending=[False, True, False, True])
cols_audit = ['Keyword', 'Volume', 'Max. Volume', 'Difficulty', 'KEI', 'Rank', 'Traffic Stability', 'Stability Class', 'BalancedScore', 'RelevancyScore', 'DecisionRule', 'Reason']
for col_idx, col in enumerate(cols_audit, 1):
    ws_dropped.cell(row=1, column=col_idx, value=col)
for row_idx, (_, row) in enumerate(df_dropped.iterrows(), 2):
    for col_idx, col in enumerate(cols_audit, 1):
        ws_dropped.cell(row=row_idx, column=col_idx, value=row.get(col, ''))
style_sheet(ws_dropped, "04_Dropped_Audit")

# --- 05_Report_Summary ---
ws_report = wb.create_sheet(title="05_Report_Summary")
ws_report.views.sheetView[0].showGridLines = True
ws_report.cell(row=1, column=1, value="ASO Keyword Planner v3.4 - Report Summary").font = Font(size=14, bold=True)
ws_report.cell(row=3, column=1, value="Metric Summary").font = Font(size=12, bold=True)
metrics = [
    ("Total Raw Keywords", len(df)),
    ("Unique Keywords", df['Keyword'].nunique()),
    ("Dropped Keywords", len(df_dropped)),
    ("Core Intent Selected", len(selected_core)),
    ("Broad Expansion Selected", len(selected_broad)),
    ("Consider Selected", len(selected_consider)),
    ("System Keywords Curated (02)", len(selected_feature)),
    ("Game Keywords Curated (03)", len(selected_style)),
    ("Duplicates Filtered (Dedup Log)", len(df_dedup_log))
]
for idx, (lbl, val) in enumerate(metrics, 4):
    ws_report.cell(row=idx, column=1, value=lbl).font = Font(bold=True)
    ws_report.cell(row=idx, column=2, value=val)

ws_report.cell(row=15, column=1, value="Language Summary").font = Font(size=12, bold=True)
lang_counts = df['LanguageGroup'].value_counts()
for idx, (lang_g, count) in enumerate(lang_counts.items(), 17):
    ws_report.cell(row=idx, column=1, value=lang_g).font = Font(bold=True)
    ws_report.cell(row=idx, column=2, value=count)

ws_report.cell(row=24, column=1, value="Naturalness Summary").font = Font(size=12, bold=True)
nat_counts = df['NaturalnessFlag'].value_counts()
for idx, (flag, count) in enumerate(nat_counts.items(), 26):
    ws_report.cell(row=idx, column=1, value=flag).font = Font(bold=True)
    ws_report.cell(row=idx, column=2, value=count)

ws_report.cell(row=3, column=4, value="Sheet Index").font = Font(size=12, bold=True)
sheets_info = [
    ("00_README_CONFIG", "App configuration parameters and run metadata"),
    ("01_Main_Keyword_Shortlist", "Top 25 Core + 5 Broad + 10 Consider shortlist for metadata allocation"),
    ("02_System_Keywords", "Curated system, console, and platform candidates (capped <= 30)"),
    ("03_Game_Keywords", "Curated game title and franchise candidates (capped <= 30, Research Only)"),
    ("04_Dropped_Audit", "Dropped keywords with detailed reasons"),
    ("05_Report_Summary", "Summary stats, language breakdowns, and sheet indices"),
    ("06_All_Candidates", "Full candidate pool with detailed score and policy values"),
    ("07_Language_Mismatch", "Audit sheet for keywords mismatching US_EN market language"),
    ("08_Generic_Style_Reserve", "Broad style-only keywords held back from metadata shortlist"),
    ("09_Manual_Review", "Audit sheet for keywords flagged with mixed or unknown languages"),
    ("10_Top_By_Score", "Candidates sorted by BalancedScore before diversity overlap filtering"),
    ("11_Secondary_Language", "Research candidates matching Spanish (Secondary Language)"),
    ("12_Text_Dedup_Log", "Log of text-level duplicates and variants pruned during optimization")
]
for idx, (title, purpose) in enumerate(sheets_info, 5):
    ws_report.cell(row=idx, column=4, value=title).font = Font(bold=True)
    ws_report.cell(row=idx, column=5, value=purpose)

thin_border = Border(left=Side(style='thin', color='C0C0C0'), right=Side(style='thin', color='C0C0C0'), 
                     top=Side(style='thin', color='C0C0C0'), bottom=Side(style='thin', color='C0C0C0'))
for r in range(4, 13):
    ws_report.cell(row=r, column=1).border = thin_border
    ws_report.cell(row=r, column=2).border = thin_border
for r in range(17, 17 + len(lang_counts)):
    ws_report.cell(row=r, column=1).border = thin_border
    ws_report.cell(row=r, column=2).border = thin_border
for r in range(26, 26 + len(nat_counts)):
    ws_report.cell(row=r, column=1).border = thin_border
    ws_report.cell(row=r, column=2).border = thin_border
for r in range(5, 5 + len(sheets_info)):
    ws_report.cell(row=r, column=4).border = thin_border
    ws_report.cell(row=r, column=5).border = thin_border

ws_report.column_dimensions['A'].width = 30
ws_report.column_dimensions['B'].width = 15
ws_report.column_dimensions['D'].width = 25
ws_report.column_dimensions['E'].width = 65

# --- 06_All_Candidates ---
ws_all = wb.create_sheet(title="06_All_Candidates")
cols_all = ['Keyword', 'Volume', 'Max. Volume', 'Difficulty', 'KEI', 'Rank', 'Traffic Stability', 'Stability Class', 'BalancedScore', 'RelevancyScore', 'CompetitorProven', 'ProvenDetails', 'Bucket', 
            'DetectedLanguage', 'LanguageGroup', 'NaturalnessFlag', 'Reason']
for col_idx, col in enumerate(cols_all, 1):
    ws_all.cell(row=1, column=col_idx, value=col)
for row_idx, (_, row) in enumerate(df.sort_values(by=['BalancedScore', 'Rank_numeric', 'KEI', 'Difficulty'], ascending=[False, True, False, True]).iterrows(), 2):
    for col_idx, col in enumerate(cols_all, 1):
        ws_all.cell(row=row_idx, column=col_idx, value=row.get(col, ''))
style_sheet(ws_all, "06_All_Candidates")

# --- 07_Language_Mismatch ---
ws_lang_m = wb.create_sheet(title="07_Language_Mismatch")
df_lang_m = df[df['Bucket'] == 'Language Mismatch Audit'].sort_values(by=['BalancedScore', 'Rank_numeric', 'KEI', 'Difficulty'], ascending=[False, True, False, True])
for col_idx, col in enumerate(cols_curated, 1):
    ws_lang_m.cell(row=1, column=col_idx, value=col)
for row_idx, (_, row) in enumerate(df_lang_m.iterrows(), 2):
    for col_idx, col in enumerate(cols_curated, 1):
        ws_lang_m.cell(row=row_idx, column=col_idx, value=row.get(col, ''))
style_sheet(ws_lang_m, "07_Language_Mismatch")

# --- 08_Generic_Style_Reserve ---
ws_reserve = wb.create_sheet(title="08_Generic_Style_Reserve")
df_reserve = df[df['Bucket'] == 'Generic Style Reserve'].sort_values(by=['BalancedScore', 'Rank_numeric', 'KEI', 'Difficulty'], ascending=[False, True, False, True])
for col_idx, col in enumerate(cols_curated, 1):
    ws_reserve.cell(row=1, column=col_idx, value=col)
for row_idx, (_, row) in enumerate(df_reserve.iterrows(), 2):
    for col_idx, col in enumerate(cols_curated, 1):
        ws_reserve.cell(row=row_idx, column=col_idx, value=row.get(col, ''))
style_sheet(ws_reserve, "08_Generic_Style_Reserve")

# --- 09_Manual_Review ---
ws_mrev = wb.create_sheet(title="09_Manual_Review")
df_mrev = df[df['Bucket'] == 'Manual Review'].sort_values(by=['BalancedScore', 'Rank_numeric', 'KEI', 'Difficulty'], ascending=[False, True, False, True])
for col_idx, col in enumerate(cols_curated, 1):
    ws_mrev.cell(row=1, column=col_idx, value=col)
for row_idx, (_, row) in enumerate(df_mrev.iterrows(), 2):
    for col_idx, col in enumerate(cols_curated, 1):
        ws_mrev.cell(row=row_idx, column=col_idx, value=row.get(col, ''))
style_sheet(ws_mrev, "09_Manual_Review")

# --- 10_Top_By_Score ---
ws_tps = wb.create_sheet(title="10_Top_By_Score")
df_tps = df.sort_values(by=['BalancedScore', 'Rank_numeric', 'KEI', 'Difficulty'], ascending=[False, True, False, True]).head(50)
for col_idx, col in enumerate(cols_curated, 1):
    ws_tps.cell(row=1, column=col_idx, value=col)
for row_idx, (_, row) in enumerate(df_tps.iterrows(), 2):
    for col_idx, col in enumerate(cols_curated, 1):
        ws_tps.cell(row=row_idx, column=col_idx, value=row.get(col, ''))
style_sheet(ws_tps, "10_Top_By_Score")

# --- 11_Secondary_Language ---
ws_seclang = wb.create_sheet(title="11_Secondary_Language")
df_seclang = df[df['LanguageGroup'] == 'SECONDARY'].sort_values(by=['BalancedScore', 'Rank_numeric', 'KEI', 'Difficulty'], ascending=[False, True, False, True])
for col_idx, col in enumerate(cols_curated, 1):
    ws_seclang.cell(row=1, column=col_idx, value=col)
for row_idx, (_, row) in enumerate(df_seclang.iterrows(), 2):
    for col_idx, col in enumerate(cols_curated, 1):
        ws_seclang.cell(row=row_idx, column=col_idx, value=row.get(col, ''))
style_sheet(ws_seclang, "11_Secondary_Language")

# --- 12_Text_Dedup_Log ---
ws_dedup = wb.create_sheet(title="12_Text_Dedup_Log")
cols_dedup = ['Table', 'RemovedKeyword', 'OriginalSection', 'KeptKeyword', 'DedupReason', 'BalancedScore', 'Note']
for col_idx, col in enumerate(cols_dedup, 1):
    ws_dedup.cell(row=1, column=col_idx, value=col)
if not df_dedup_log.empty:
    for row_idx, (_, row) in enumerate(df_dedup_log.iterrows(), 2):
        for col_idx, col in enumerate(cols_dedup, 1):
            ws_dedup.cell(row=row_idx, column=col_idx, value=row.get(col, ''))
style_sheet(ws_dedup, "12_Text_Dedup_Log")

# Save
print(f"Saving stylized workbook to {OUTPUT_PATH}...")
try:
    wb.save(OUTPUT_PATH)
    print("Pipeline for Game Emulator complete!")
except PermissionError:
    alt_path = OUTPUT_PATH.replace(".xlsx", "_temp.xlsx")
    print(f"WARNING: Permission denied to write to {OUTPUT_PATH} (file is likely open in another program).")
    print(f"Saving to fallback path: {alt_path}")
    wb.save(alt_path)
    print("Pipeline complete (saved to fallback path)!")
