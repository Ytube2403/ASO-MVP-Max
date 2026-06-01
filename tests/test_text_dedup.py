import os
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

from shared import text_dedup


class TextDedupTests(unittest.TestCase):
    def test_nfkc_casefold_matches_width_and_case_variants(self):
        self.assertEqual(text_dedup.normalize_text("STRASSE"), text_dedup.normalize_text("stra\u00dfe"))
        self.assertEqual(text_dedup.normalize_text("\uff76\uff92\uff97"), text_dedup.normalize_text("\u30ab\u30e1\u30e9"))

    def test_meaningful_marks_are_preserved(self):
        self.assertNotEqual(text_dedup.normalize_text("m\u00e1"), text_dedup.normalize_text("ma"))
        self.assertNotEqual(
            text_dedup.normalize_text("\u092b\u093c\u093f\u0932\u094d\u091f\u0930"),
            text_dedup.normalize_text("\u092b\u093f\u0932\u094d\u091f\u0930"),
        )
        self.assertNotEqual(
            text_dedup.normalize_text("\u062a\u064e\u0637\u0628\u064a\u0642"),
            text_dedup.normalize_text("\u062a\u0637\u0628\u064a\u0642"),
        )

    def test_accent_fold_variants_remain_independent_by_default(self):
        result = text_dedup.deduplicate_candidates(
            [
                {"Keyword": "retro", "Volume": 10, "Bucket": "Core Intent Final"},
                {"Keyword": "retr\u00f4", "Volume": 8, "Bucket": "Core Intent Final"},
            ],
            "sheet",
            "pt",
        )
        self.assertEqual(len(result.records), 2)
        self.assertEqual(result.log_entries, [])
        self.assertTrue(all("ReviewVariants" not in record for record in result.records))

    def test_accent_fold_can_be_enabled_for_explicit_locale_allowlist(self):
        result = text_dedup.deduplicate_candidates(
            [
                {"Keyword": "retro", "Volume": 10, "Bucket": "Core Intent Final"},
                {"Keyword": "retr\u00f4", "Volume": 8, "Bucket": "Core Intent Final"},
            ],
            "sheet",
            "pt-BR",
            {"accent_fold_auto_merge_locales": ["pt"]},
        )
        self.assertEqual([record["Keyword"] for record in result.records], ["retro"])
        self.assertEqual(result.log_entries[0]["DedupRule"], "accent_fold_allowlist")

    def test_mixed_latin_cyrillic_keyword_remains_independent(self):
        result = text_dedup.deduplicate_candidates(
            [{"Keyword": "\u0430r filter", "Volume": 10, "Bucket": "Core Intent Final"}],
            "sheet",
            "en",
        )
        self.assertEqual(len(result.records), 1)
        self.assertEqual(result.log_entries, [])

    def test_cluster_keeps_current_volume_winner_across_buckets(self):
        result = text_dedup.deduplicate_candidates(
            [
                {"Keyword": "AR FILTER", "Volume": 5, "Max. Volume": 50, "BalancedScore": 0.9, "Bucket": "Core Intent Final"},
                {"Keyword": "ar-filter", "Volume": 12, "Max. Volume": 12, "BalancedScore": 0.5, "Bucket": "Broad Expansion"},
            ],
            "sheet",
            "en",
        )
        self.assertEqual([record["Keyword"] for record in result.records], ["ar-filter"])
        self.assertEqual(result.records[0]["MergedVariants"], "AR FILTER")
        self.assertEqual(result.log_entries[0]["Action"], "PRUNED")
        self.assertEqual(result.log_entries[0]["KeptVolume"], 12)

    def test_quota_selection_can_refill_from_remaining_unique_candidates(self):
        result = text_dedup.deduplicate_candidates(
            [
                {"Keyword": "ar filter", "Volume": 30, "Bucket": "Core Intent Final"},
                {"Keyword": "AR-FILTER", "Volume": 20, "Bucket": "Core Intent Final"},
                {"Keyword": "camera effects", "Volume": 10, "Bucket": "Feature Keywords"},
            ],
            "sheet",
            "en",
        )
        quota = result.records[:2]
        self.assertEqual([record["Keyword"] for record in quota], ["ar filter", "camera effects"])

    def test_log_entries_round_trip_through_workbook_schema(self):
        result = text_dedup.deduplicate_candidates(
            [
                {"Keyword": "AR FILTER", "Volume": 5, "Bucket": "Core Intent Final"},
                {"Keyword": "ar-filter", "Volume": 12, "Bucket": "Broad Expansion"},
            ],
            "sheet",
            "en",
        )
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "text_dedup_log.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "12_Text_Dedup_Log"
            sheet.append(text_dedup.TEXT_DEDUP_LOG_COLUMNS)
            for entry in result.log_entries:
                sheet.append([entry.get(column, "") for column in text_dedup.TEXT_DEDUP_LOG_COLUMNS])
            workbook.save(path)

            loaded = load_workbook(path, read_only=True, data_only=True)
            rows = list(loaded["12_Text_Dedup_Log"].iter_rows(values_only=True))
            loaded.close()
        self.assertEqual(list(rows[0]), text_dedup.TEXT_DEDUP_LOG_COLUMNS)
        self.assertEqual(rows[1][1], "PRUNED")

    def test_dictionary_segmentation_scripts_do_not_fake_token_bags(self):
        keys = text_dedup.build_dedup_keys("\u30ab\u30e1\u30e9\u30a2\u30d7\u30ea", "ja")
        self.assertFalse(keys.tokenization_trusted)
        self.assertEqual(keys.stemmed_token_bag_key, ())

    @unittest.skipUnless(text_dedup.snowballstemmer, "snowballstemmer is optional")
    def test_snowball_stemming_merges_locale_variants(self):
        result = text_dedup.deduplicate_candidates(
            [
                {"Keyword": "filters app", "Volume": 5, "Bucket": "Broad Expansion"},
                {"Keyword": "filter app", "Volume": 9, "Bucket": "Broad Expansion"},
            ],
            "sheet",
            "en",
        )
        self.assertEqual([record["Keyword"] for record in result.records], ["filter app"])
        self.assertEqual(result.log_entries[0]["DedupRule"], "stemmed_sequence_key")


if __name__ == "__main__":
    unittest.main()
