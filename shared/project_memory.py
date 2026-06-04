import ast
import importlib.util
import json
import os
from datetime import datetime

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from shared.keyword_filter.matcher import normalize_filter_text
from shared.profile_service import adapt_custom_profile, empty_profile


TERM_GROUPS = {
    "core_terms": "intent_core_terms",
    "core_words": "intent_core_words",
    "feature_terms": "feature_terms",
    "style_terms": "style_terms",
    "visual_terms": "visual_terms",
}
RISK_GROUPS = {
    "competitor_brands": "competitor_brands",
    "noise_terms": "noise_terms",
    "typo_blacklist": "typo_blacklist",
    "irrelevant_intent_terms": "irrelevant_intent_terms",
    "risky_platform_terms": "risky_platform_terms",
    "risky_ip_terms": "risky_ip_terms",
    "ambiguous_brand_terms": "ambiguous_brand_terms",
    "platform_affiliation_terms": "platform_affiliation_terms",
}


def _read_json(path):
    with open(path, "r", encoding="utf-8") as handle:
        return json.load(handle)


def _load_module(path, module_name):
    spec = importlib.util.spec_from_file_location(module_name, path)
    if spec is None or spec.loader is None:
        raise ImportError(f"Cannot import {path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _safe_literal(node):
    if isinstance(node, ast.Constant):
        return node.value
    if isinstance(node, ast.List):
        return [_safe_literal(item) for item in node.elts]
    if isinstance(node, ast.Tuple):
        return tuple(_safe_literal(item) for item in node.elts)
    if isinstance(node, ast.Dict):
        output = {}
        for key_node, value_node in zip(node.keys, node.values):
            key = _safe_literal(key_node)
            if key is None:
                continue
            output[key] = _safe_literal(value_node)
        return output
    if isinstance(node, ast.UnaryOp) and isinstance(node.op, ast.USub):
        value = _safe_literal(node.operand)
        return -value if isinstance(value, (int, float)) else ""
    return ""


def _literal_config_from_runner(runner_path):
    if not runner_path or not os.path.exists(runner_path):
        return {}, ""
    with open(runner_path, "r", encoding="utf-8") as handle:
        tree = ast.parse(handle.read(), filename=runner_path)
    for node in tree.body:
        if isinstance(node, ast.Assign):
            for target in node.targets:
                if isinstance(target, ast.Name) and target.id == "config":
                    value = _safe_literal(node.value)
                    if isinstance(value, dict):
                        return value, runner_path
                    return {}, "runner config is not a dict"
    return {}, "runner config assignment not found"


def load_config_from_app(app_folder, runner_path=None):
    """Load setup config without running the pipeline."""
    app_folder = os.path.abspath(app_folder)
    config_path = os.path.join(app_folder, "app_config.py")
    config = {}
    source = ""
    notes = []

    if os.path.exists(config_path):
        module = _load_module(config_path, f"project_memory_config_{abs(hash(config_path))}")
        if hasattr(module, "APP_CONFIG") and isinstance(module.APP_CONFIG, dict):
            config = dict(module.APP_CONFIG)
            source = config_path
        else:
            config, source = _literal_config_from_runner(runner_path)
            if hasattr(module, "FILTER_POLICY") and isinstance(module.FILTER_POLICY, dict):
                config.update(module.FILTER_POLICY)
                notes.append("FILTER_POLICY overlaid from app_config.py")
            if not config and hasattr(module, "FILTER_POLICY") and isinstance(module.FILTER_POLICY, dict):
                config = dict(module.FILTER_POLICY)
                source = config_path
                notes.append("Only FILTER_POLICY was available")
    else:
        config, source = _literal_config_from_runner(runner_path)
        notes.append("app_config.py was not found")

    return config, source, notes


def load_profile_from_app(app_folder, config):
    profile_path = os.path.join(os.path.abspath(app_folder), "App_Profile.json")
    if not os.path.exists(profile_path):
        return empty_profile(config, error="App_Profile.json not found"), {}, profile_path
    raw_profile = _read_json(profile_path)
    return adapt_custom_profile(raw_profile, config), raw_profile, profile_path


def _terms(config, key):
    return [str(term) for term in config.get(key, []) or [] if str(term).strip()]


def _policy_items(policy):
    return [
        {"name": key, "action": value}
        for key, value in sorted((policy or {}).items())
        if key.endswith("_action") or key == "core_intent_override"
    ]


def _overlap_warnings(config):
    warnings = []
    competitors = {normalize_filter_text(term) for term in _terms(config, "competitor_brands")}
    for key in ("feature_terms", "risky_platform_terms", "ambiguous_brand_terms", "platform_affiliation_terms"):
        overlap = sorted(competitors & {normalize_filter_text(term) for term in _terms(config, key)})
        if overlap:
            warnings.append(f"competitor_brands overlaps {key}: {', '.join(overlap)}")
    return warnings


def _setup_warnings(config, profile, raw_profile, notes):
    warnings = list(notes or [])
    competitors = profile.get("competitors", []) or []
    if not competitors:
        warnings.append("No competitor apps found in App_Profile.json")
    if profile.get("ProfileError"):
        warnings.append(str(profile.get("ProfileError")))
    if not _terms(config, "competitor_brands"):
        warnings.append("No competitor_brands configured for hard-drop protection")
    for key, label, minimum in (
        ("intent_core_terms", "core terms", 3),
        ("feature_terms", "feature terms", 3),
        ("style_terms", "style terms", 3),
    ):
        if len(_terms(config, key)) < minimum:
            warnings.append(f"Only {len(_terms(config, key))} {label} configured")
    config_app_id = str(config.get("app_id", "")).strip()
    profile_app_id = str(profile.get("app_id", "")).strip()
    if config_app_id and profile_app_id and config_app_id != profile_app_id:
        warnings.append(f"Profile app_id differs from config app_id: {profile_app_id} != {config_app_id}")
    if raw_profile and "competitor_strategy" in raw_profile:
        suggested = raw_profile.get("competitor_strategy", {}).get("suggested_competitors", []) or []
        if len(suggested) != len(competitors):
            warnings.append("Some suggested competitors could not be adapted")
    warnings.extend(_overlap_warnings(config))
    return warnings


def build_project_memory(config, app_profile=None, raw_profile=None, app_folder="", config_source="", notes=None):
    app_profile = app_profile or {}
    raw_profile = raw_profile or {}
    corrected = raw_profile.get("corrected_positioning", {}) if isinstance(raw_profile, dict) else {}
    memory = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "app_folder": os.path.abspath(app_folder) if app_folder else "",
        "config_source": config_source,
        "identity": {
            "app_name": config.get("app_name", ""),
            "app_id": config.get("app_id", ""),
            "category": config.get("category", ""),
            "market": config.get("market") or "Runtime-selected",
            "platform_mode": config.get("platform_mode", ""),
            "semantic_mode": config.get("semantic_mode", ""),
        },
        "positioning": {
            "store_title": app_profile.get("title", ""),
            "short_description": app_profile.get("short_description", ""),
            "full_description_summary": app_profile.get("full_description", ""),
            "primary_positioning": corrected.get("primary_positioning", ""),
            "not_the_right_positioning": corrected.get("not_the_right_positioning", []) or [],
            "strongest_differentiators": corrected.get("strongest_differentiators", []) or [],
            "profile_status": app_profile.get("ProfileStatus", ""),
            "profile_error": app_profile.get("ProfileError", ""),
        },
        "keyword_setup": {name: _terms(config, key) for name, key in TERM_GROUPS.items()},
        "competitor_setup": {
            "competitor_apps": app_profile.get("competitors", []) or [],
            "blocked_brands": _terms(config, "competitor_brands"),
        },
        "risk_setup": {
            "groups": {name: _terms(config, key) for name, key in RISK_GROUPS.items()},
            "risk_policy": _policy_items(config.get("risk_policy", {}) or {}),
        },
        "user_overrides": config.get("user_overrides", {}) or {},
        "keyword_quota": config.get("keyword_quota", {}) or {},
    }
    memory["warnings"] = _setup_warnings(config, app_profile, raw_profile, notes)
    return memory


def build_project_memory_for_app(app_folder, runner_path=None):
    config, source, notes = load_config_from_app(app_folder, runner_path=runner_path)
    profile, raw_profile, _ = load_profile_from_app(app_folder, config)
    return build_project_memory(config, profile, raw_profile, app_folder, source, notes)


def _join_values(values, limit=20):
    values = [str(value) for value in values or [] if str(value).strip()]
    if not values:
        return ""
    visible = values[:limit]
    suffix = f" (+{len(values) - limit} more)" if len(values) > limit else ""
    return ", ".join(visible) + suffix


def add_project_memory_sheet(wb, memory):
    ws = wb.create_sheet(title="00_Project_Memory", index=0)
    ws.views.sheetView[0].showGridLines = False
    section_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
    warning_fill = PatternFill(start_color="FEF7E0", end_color="FEF7E0", fill_type="solid")
    title_font = Font(name="Calibri", size=15, bold=True, color="1A73E8")
    section_font = Font(name="Calibri", size=11, bold=True, color="202124")
    thin = Side(border_style="thin", color="DADCE0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    row = 1
    ws.cell(row=row, column=1, value="Project Memory / Setup Overview").font = title_font
    ws.cell(row=row, column=2, value=memory.get("generated_at", ""))
    row += 2

    def section(title):
        nonlocal row
        ws.cell(row=row, column=1, value=title)
        ws.cell(row=row, column=1).fill = section_fill
        ws.cell(row=row, column=1).font = section_font
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1

    def pair(label, value):
        nonlocal row
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=1).font = section_font
        row += 1

    section("Identity")
    for key, value in memory.get("identity", {}).items():
        pair(key, value)

    section("Keyword Setup")
    for key, values in memory.get("keyword_setup", {}).items():
        pair(key, _join_values(values))

    section("Competitors")
    pair("Blocked brands", _join_values(memory.get("competitor_setup", {}).get("blocked_brands", [])))
    for comp in memory.get("competitor_setup", {}).get("competitor_apps", []) or []:
        pair(comp.get("title") or comp.get("package_id", "Competitor"), comp.get("short_description", ""))

    section("Risk Policy")
    for item in memory.get("risk_setup", {}).get("risk_policy", []):
        pair(item["name"], item["action"])

    section("Warnings")
    warnings = memory.get("warnings", []) or ["No setup warnings"]
    for warning in warnings:
        pair("warning", warning)
        ws.cell(row=row - 1, column=2).fill = warning_fill

    for col_idx, width in enumerate([28, 90, 18], 1):
        ws.column_dimensions[chr(64 + col_idx)].width = width
    for cells in ws.iter_rows():
        for cell in cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row > 2:
                cell.border = border
    return ws


def render_project_memory_markdown(memory):
    identity = memory.get("identity", {})
    competitor_setup = memory.get("competitor_setup", {})
    lines = [
        "# Project Memory",
        "",
        f"Generated: {memory.get('generated_at', '')}",
        "",
        "## Identity",
        f"- App: {identity.get('app_name', '')}",
        f"- App ID: {identity.get('app_id', '')}",
        f"- Category: {identity.get('category', '')}",
        f"- Market: {identity.get('market', '')}",
        f"- Semantic mode: {identity.get('semantic_mode', '')}",
        "",
        "## Positioning",
    ]
    positioning = memory.get("positioning", {})
    for key in ("store_title", "short_description", "primary_positioning"):
        value = positioning.get(key, "")
        if value:
            lines.append(f"- {key}: {value}")
    for key in ("not_the_right_positioning", "strongest_differentiators"):
        values = positioning.get(key, []) or []
        if values:
            lines.append(f"- {key}: {_join_values(values, limit=30)}")
    lines.extend(["", "## Keyword Setup"])
    for key, values in memory.get("keyword_setup", {}).items():
        lines.append(f"- {key}: {_join_values(values, limit=40)}")
    lines.extend(["", "## Competitors"])
    lines.append(f"- Blocked brands: {_join_values(competitor_setup.get('blocked_brands', []), limit=40)}")
    for comp in competitor_setup.get("competitor_apps", []) or []:
        title = comp.get("title") or comp.get("package_id", "")
        package_id = comp.get("package_id", "")
        lines.append(f"- {title} ({package_id})")
    lines.extend(["", "## Risk Rules"])
    for item in memory.get("risk_setup", {}).get("risk_policy", []):
        lines.append(f"- {item['name']}: {item['action']}")
    lines.extend(["", "## Warnings"])
    warnings = memory.get("warnings", []) or ["No setup warnings"]
    lines.extend(f"- {warning}" for warning in warnings)
    lines.append("")
    return "\n".join(lines)


def write_project_memory_markdown(app_folder, memory):
    path = os.path.join(os.path.abspath(app_folder), "PROJECT_MEMORY.md")
    with open(path, "w", encoding="utf-8", newline="\n") as handle:
        handle.write(render_project_memory_markdown(memory))
    return path
