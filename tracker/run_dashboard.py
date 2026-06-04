import os
import sys
import webbrowser
import threading
import time
import io
from flask import Flask, jsonify, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from . import data_scanner, db_manager
except ImportError:
    import data_scanner
    import db_manager

app = Flask(__name__, static_folder="static", static_url_path="")
DASHBOARD_HOST = os.environ.get("ASO_DASHBOARD_HOST", "127.0.0.1")
DASHBOARD_PORT = int(os.environ.get("ASO_DASHBOARD_PORT", "5101"))

ASO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ASO_ROOT not in sys.path:
    sys.path.insert(0, ASO_ROOT)

from shared.app_registry import resolve_app
from shared.project_memory import build_project_memory_for_app

# Ensure DB initialized and initial scan performed on startup
print("Initializing database and scanning directories...")
db_manager.init_db()
data_scanner.scan_and_import()

@app.route("/")
def index():
    return app.send_static_file("index.html")

@app.route("/api/apps")
def get_apps():
    try:
        apps = db_manager.get_available_apps()
        return jsonify({"success": True, "apps": apps})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/setup/<app_name>")
def get_setup(app_name):
    try:
        try:
            registered_app = resolve_app(app_name, ASO_ROOT)
            app_folder = os.path.join(ASO_ROOT, *registered_app["folder"].split("/"))
            runner_path = registered_app.get("runner_path")
        except KeyError:
            app_folder = os.path.join(ASO_ROOT, "apps", app_name)
            runner_path = None
        if not os.path.isdir(app_folder):
            return jsonify({"success": False, "error": f"Unknown app setup folder: {app_name}"}), 404
        memory = build_project_memory_for_app(app_folder, runner_path=runner_path)
        return jsonify({"success": True, "data": memory})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/locales/<app_name>")
def get_locales(app_name):
    try:
        locales = db_manager.get_available_locales(app_name)
        return jsonify({"success": True, "locales": locales})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/months/<app_name>/<locale>")
def get_months(app_name, locale):
    try:
        months = db_manager.get_available_months(app_name, locale)
        return jsonify({"success": True, "months": months})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/overview/<app_name>/<locale>")
def get_overview(app_name, locale):
    try:
        overview_data = db_manager.get_overview_data(app_name, locale)
        return jsonify({"success": True, "data": overview_data})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/keywords/<app_name>/<locale>")
def get_keywords(app_name, locale):
    try:
        month_a = request.args.get("month_a", "")
        month_b = request.args.get("month_b", "")
        if not month_b:
            return jsonify({"success": False, "error": "month_b is required"}), 400
            
        data = db_manager.get_comparison_data_v2(app_name, locale, month_a, month_b)
        return jsonify({"success": True, "data": data})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/trend/<app_name>/<locale>/<path:keyword>")
def get_trend(app_name, locale, keyword):
    try:
        trend_data = db_manager.get_keyword_trend(app_name, locale, keyword)
        return jsonify({"success": True, "data": trend_data})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/movers/<app_name>/<locale>")
def get_movers(app_name, locale):
    try:
        month_a = request.args.get("month_a", "")
        month_b = request.args.get("month_b", "")
        if not month_b:
            return jsonify({"success": False, "error": "month_b is required"}), 400
            
        movers = db_manager.get_movers(app_name, locale, month_a, month_b)
        return jsonify({"success": True, "data": movers})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/refresh", methods=["POST"])
def refresh_data():
    try:
        files, rows = data_scanner.scan_and_import()
        return jsonify({"success": True, "files_imported": files, "rows_imported": rows})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/export/<app_name>/<locale>")
def export_excel(app_name, locale):
    try:
        month_a = request.args.get("month_a", "")
        month_b = request.args.get("month_b", "")
        
        if not month_b:
            return "month_b parameter is required", 400
            
        wb = Workbook()
        
        # Style definitions
        font_family = "Segoe UI"
        title_font = Font(name=font_family, size=16, bold=True, color="1A73E8")
        header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        sub_header_font = Font(name=font_family, size=12, bold=True, color="202124")
        bold_font = Font(name=font_family, size=11, bold=True)
        regular_font = Font(name=font_family, size=11)
        italic_font = Font(name=font_family, size=10, italic=True)
        
        header_fill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
        zebra_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        
        # Alert fills
        green_fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
        red_fill = PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid")
        yellow_fill = PatternFill(start_color="FEF7E0", end_color="FEF7E0", fill_type="solid")
        
        green_text = Font(name=font_family, size=11, color="137333")
        red_text = Font(name=font_family, size=11, color="C5221F")
        yellow_text = Font(name=font_family, size=11, color="B06000")
        
        thin_border_side = Side(border_style="thin", color="DADCE0")
        thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
        
        align_left = Alignment(horizontal="left", vertical="center")
        align_center = Alignment(horizontal="center", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        
        # =========================================================================
        # SHEET 1: Overview
        # =========================================================================
        ws_overview = wb.active
        ws_overview.title = "Overview"
        ws_overview.views.sheetView[0].showGridLines = True
        
        ws_overview.append([])
        ws_overview.cell(row=2, column=2, value=f"ASO Performance Overview - {app_name} ({locale})").font = title_font
        ws_overview.cell(row=3, column=2, value=f"Generated on: {time.strftime('%Y-%m-%d %H:%M:%S')}").font = italic_font
        
        overview_data = db_manager.get_overview_data(app_name, locale)
        
        # Headers for Overview table
        ov_headers = ["Month", "ASO Power Score", "Total Keywords", "Top 1", "Top 2-3", "Top 4-10", "Top 11-30", "Top 31-100", "Unranked"]
        start_row = 5
        
        for col_idx, text in enumerate(ov_headers, start=2):
            cell = ws_overview.cell(row=start_row, column=col_idx, value=text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = thin_border
            
        for row_idx, month_data in enumerate(overview_data, start=start_row + 1):
            m = month_data["month"]
            score = month_data["aso_power_score"]
            t = month_data["tiers"]
            
            row_data = [
                m, score, t["total_keywords"], t["top_1"], t["top_2_3"],
                t["top_4_10"], t["top_11_30"], t["top_31_100"], t["unranked"]
            ]
            
            for col_idx, val in enumerate(row_data, start=2):
                cell = ws_overview.cell(row=row_idx, column=col_idx, value=val)
                cell.font = regular_font
                cell.border = thin_border
                
                # Zebra striping
                if row_idx % 2 == 0:
                    cell.fill = zebra_fill
                    
                if col_idx == 2:  # Month
                    cell.alignment = align_center
                elif col_idx == 3:  # Power Score
                    cell.font = bold_font
                    cell.alignment = align_right
                else:
                    cell.alignment = align_right
                    
        # =========================================================================
        # SHEET 2: Keyword Comparison
        # =========================================================================
        ws_comp = wb.create_sheet(title="Keyword Comparison")
        ws_comp.views.sheetView[0].showGridLines = True
        
        comp_headers = [
            "Keyword", "Brand", 
            f"Volume ({month_a or 'N/A'})", f"Volume ({month_b})", "Volume Delta",
            f"Difficulty ({month_a or 'N/A'})", f"Difficulty ({month_b})", "Difficulty Delta",
            f"Rank ({month_a or 'N/A'})", f"Rank ({month_b})", "Rank Delta", "Status"
        ]
        
        ws_comp.append([])
        ws_comp.cell(row=2, column=2, value=f"Keyword Comparison: {month_a or 'N/A'} vs {month_b}").font = title_font
        
        start_row = 4
        for col_idx, text in enumerate(comp_headers, start=2):
            cell = ws_comp.cell(row=start_row, column=col_idx, value=text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = thin_border
            
        comparison_data = db_manager.get_comparison_data_v2(app_name, locale, month_a, month_b)
        
        for row_idx, r in enumerate(comparison_data, start=start_row + 1):
            kw = r["keyword"]
            brand = "Yes" if r["brand"] else "No"
            vol_a = r["volume_a"]
            vol_b = r["volume_b"]
            diff_vol = r["diff_volume"]
            diff_a = r["difficulty_a"]
            diff_b = r["difficulty_b"]
            diff_diff = r["diff_difficulty"]
            rank_a = r["rank_a"] if r["rank_a"] > 0 else "Unranked"
            rank_b = r["rank_b"] if r["rank_b"] > 0 else "Unranked"
            diff_rank = r["diff_rank"]
            status = r["status"]
            
            row_data = [
                kw, brand, vol_a, vol_b, diff_vol,
                diff_a, diff_b, diff_diff,
                rank_a, rank_b, diff_rank, status
            ]
            
            for col_idx, val in enumerate(row_data, start=2):
                cell = ws_comp.cell(row=row_idx, column=col_idx, value=val)
                cell.font = regular_font
                cell.border = thin_border
                
                # Zebra striping
                if row_idx % 2 == 0:
                    cell.fill = zebra_fill
                
                # Alignments
                if col_idx == 2:  # Keyword
                    cell.alignment = align_left
                elif col_idx in (3, 13):  # Brand, Status
                    cell.alignment = align_center
                elif col_idx in (10, 11):  # Rank A, Rank B
                    cell.alignment = align_center
                else:  # Numeric columns
                    cell.alignment = align_right
                    
                # Highlighting Volume Delta (Col 6)
                if col_idx == 6:
                    if val > 0:
                        cell.fill = green_fill
                        cell.font = green_text
                    elif val < 0:
                        cell.fill = red_fill
                        cell.font = red_text
                        
                # Highlighting Rank Delta (Col 12)
                if col_idx == 12:
                    # For Rank Delta, negative is GOOD (rank improved)
                    if val < 0:
                        cell.fill = green_fill
                        cell.font = green_text
                        cell.value = f"↑ {abs(val)}"
                    elif val > 0:
                        cell.fill = red_fill
                        cell.font = red_text
                        cell.value = f"↓ {val}"
                    else:
                        cell.value = "—"
                        
                # Status cell color
                if col_idx == 13:
                    if val == "New":
                        cell.fill = green_fill
                        cell.font = green_text
                    elif val == "Lost":
                        cell.fill = red_fill
                        cell.font = red_text
                    elif val == "↑":
                        cell.font = green_text
                    elif val == "↓":
                        cell.font = red_text
                        
        # =========================================================================
        # SHEET 3: Movers
        # =========================================================================
        ws_movers = wb.create_sheet(title="Movers")
        ws_movers.views.sheetView[0].showGridLines = True
        
        ws_movers.append([])
        ws_movers.cell(row=2, column=2, value=f"Top Movers Summary: {month_a or 'N/A'} vs {month_b}").font = title_font
        
        movers = db_manager.get_movers(app_name, locale, month_a, month_b)
        
        def write_movers_table(ws, title, data_list, start_c, headers, is_gainer=True):
            # Write Title
            ws.cell(row=4, column=start_c, value=title).font = sub_header_font
            
            # Write Headers
            for col_idx, h in enumerate(headers):
                cell = ws.cell(row=5, column=start_c + col_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = align_center
                cell.border = thin_border
                
            # Write Data
            for row_idx, r in enumerate(data_list):
                curr_row = 6 + row_idx
                
                kw = r["keyword"]
                vol = r["volume_b"]
                diff = r["difficulty_b"]
                
                if title in ("Top Gainers", "Top Losers"):
                    r_a = r["rank_a"] if r["rank_a"] > 0 else "Unranked"
                    r_b = r["rank_b"] if r["rank_b"] > 0 else "Unranked"
                    delta = abs(r["diff_rank"])
                    delta_str = f"↑ {delta}" if is_gainer else f"↓ {delta}"
                    row_data = [kw, r_a, r_b, delta_str, vol]
                elif title == "New Keywords":
                    rank = r["rank_b"] if r["rank_b"] > 0 else "Unranked"
                    row_data = [kw, vol, diff, rank]
                else:  # Lost Keywords
                    vol_a = r["volume_a"]
                    rank_a = r["rank_a"] if r["rank_a"] > 0 else "Unranked"
                    row_data = [kw, vol_a, rank_a]
                    
                for col_idx, val in enumerate(row_data):
                    cell = ws.cell(row=curr_row, column=start_c + col_idx, value=val)
                    cell.font = regular_font
                    cell.border = thin_border
                    
                    if curr_row % 2 == 0:
                        cell.fill = zebra_fill
                        
                    if col_idx == 0:
                        cell.alignment = align_left
                    else:
                        cell.alignment = align_center
                        
                    # Highlight diffs
                    if title in ("Top Gainers", "Top Losers") and col_idx == 3:
                        cell.font = bold_font
                        if is_gainer:
                            cell.fill = green_fill
                            cell.font = green_text
                        else:
                            cell.fill = red_fill
                            cell.font = red_text
                            
                    # Highlight new/lost status colors
                    if title == "New Keywords" and col_idx == 3:
                        cell.fill = green_fill
                        cell.font = green_text
                    if title == "Lost Keywords" and col_idx == 2:
                        cell.fill = red_fill
                        cell.font = red_text
                        
        # 1. Gainers (Cols 2-6)
        gain_headers = ["Keyword", "Old Rank", "New Rank", "Delta", "Volume"]
        write_movers_table(ws_movers, "Top Gainers", movers["gainers"], 2, gain_headers, is_gainer=True)
        
        # 2. Losers (Cols 8-12)
        lose_headers = ["Keyword", "Old Rank", "New Rank", "Delta", "Volume"]
        write_movers_table(ws_movers, "Top Losers", movers["losers"], 8, lose_headers, is_gainer=False)
        
        # 3. New (Cols 14-17)
        new_headers = ["Keyword", "Volume", "Difficulty", "Rank"]
        write_movers_table(ws_movers, "New Keywords", movers["new"], 14, new_headers)
        
        # 4. Lost (Cols 19-21)
        lost_headers = ["Keyword", "Last Volume", "Last Rank"]
        write_movers_table(ws_movers, "Lost Keywords", movers["lost"], 19, lost_headers)
        
        # Autofit columns for all sheets
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                
                # Check column size
                for cell in col:
                    if cell.value:
                        # Handle title text span
                        if cell.row in (2, 3, 4) and col_letter in ('B', 'H', 'N', 'S'):
                            continue
                        max_len = max(max_len, len(str(cell.value)))
                        
                if max_len > 0:
                    sheet.column_dimensions[col_letter].width = max(max_len + 3, 10)
                    
        # Save workbook to memory
        excel_stream = io.BytesIO()
        wb.save(excel_stream)
        excel_stream.seek(0)
        
        filename = f"ASO_Comparison_{app_name}_{locale}_{month_b}.xlsx"
        return send_file(
            excel_stream,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500

def open_browser():
    # Wait for flask to start
    time.sleep(1.5)
    url = f"http://{DASHBOARD_HOST}:{DASHBOARD_PORT}"
    print(f"Opening dashboard in browser at {url}...")
    webbrowser.open(url)

if __name__ == "__main__":
    # Start browser-opening thread
    threading.Thread(target=open_browser, daemon=True).start()
    
    # Run server
    app.run(host=DASHBOARD_HOST, port=DASHBOARD_PORT, debug=False)
