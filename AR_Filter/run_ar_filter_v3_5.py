import pandas as pd
import numpy as np
import re
import os
import unicodedata
import json
import urllib.request
import urllib.parse
import html as html_lib
import ssl
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import argparse

# Parse arguments
parser = argparse.ArgumentParser(description="ASO Keyword Planner for AR Filter")
parser.add_argument("--csv", type=str, default=None, help="Path to input CSV")
parser.add_argument("--market", type=str, default="US_EN", help="Market code (e.g. US_EN)")
parser.add_argument("--output", type=str, default="", help="Path to output Excel file")
parser.add_argument("--interactive", action="store_true", help="Run interactive Web UI selector")
args, unknown = parser.parse_known_args()

INPUT_PATH = args.csv
if args.output:
    OUTPUT_PATH = args.output
else:
    # Update OUTPUT_PATH dynamically
    csv_dir = os.path.dirname(os.path.abspath(INPUT_PATH))
    OUTPUT_PATH = os.path.join(csv_dir, "AR_Filter", f"ARFilter_{args.market.replace('_', '-')}_Output.xlsx")

# AR Filter configuration
config = {
    "app_id": "com.filter.ar.effect.camera3d.fyp.meme",
    "app_name": "AR Filter: FYP Dogy Filter",
    "category": "Entertainment / Camera Filter",
    "market": args.market,
    "platform_mode": "google_play",
    "semantic_mode": "ar_filter",
    
    "intent_core_terms": [
        "ar filter", "ar filters", "ar camera", "face filter", "face filters",
        "face effect", "face effects", "funny filter", "funny filters",
        "funny face filter", "funny face filters", "weird face filter",
        "weird face filters", "funny face effect", "funny face effects",
        "meme filter", "meme filters", "dogy filter", "dogy filters",
        "ar dogy filter", "dog filter", "dog filters", "ar effect", "ar effects",
        "3d filter", "3d filters", "3d ar filter", "3d ar filters"
    ],
    
    "intent_core_words": [
        "filter", "filters", "effect", "effects", "camera", "cam", "lens", "lenses"
    ],
    
    "feature_terms": [
        "3d characters", "3d character", "ar characters", "ar character",
        "funny video", "funny videos", "ar video", "ar videos", "short video",
        "short videos", "dance with", "interactive", "time warp scan",
        "time warp", "warp face", "face warp", "crying filter", "crying filters",
        "crying face", "bald filter", "bald filters", "bald head", "dog face",
        "puppy filter", "puppy face", "distort face", "distortion filter",
        "funny face", "silly face", "weird face", "ugly face", "ugly filter",
        "ugly face filter", "face morph", "morphing filter",
        "face transformation", "face tune", "funny camera", "meme camera",
        "funny movements", "character actions", "character control"
    ],
    
    "style_terms": [
        "fyp", "meme", "memes", "tiktok trend", "tiktok trends", "snapchat",
        "instagram", "tiktok", "dogy", "dogy dance", "funny dance",
        "silly dance", "crying", "bald", "dog", "puppy", "cartoon", "anime",
        "avatar", "weird", "funny", "crazy", "hilarious", "playful"
    ],
    
    "visual_terms": [
        "camera", "video", "clip", "clips", "recording", "recorder", "lens",
        "lenses", "selfie", "selfies", "photo", "photos", "picture", "pictures"
    ],
    
    "competitor_brands": [
        "snapchat", "tiktok", "instagram", "b612", "snow", "faceapp",
        "youcam makeup", "youcam", "beautyplus", "beauty plus", "banuba",
        "faceplay", "face play", "sweetsnap", "sweet snap", "facelab",
        "faceline", "faceover", "facetime", "reface", "wombo", "camera360",
        "camera 360", "retrica", "picsart", "pics art", "facetune", "lensa",
        "loopsie", "face warp", "face morph", "faceapp free", "snapchat filter"
    ],
    
    "typo_blacklist": [
        "camra", "camara", "fliter", "filte", "efect", "effets", "effct",
        "fliters", "camear", "arflter", "arfliters", "dogyf", "dogyy", "doggyy",
        "snapcht", "instgrm", "tik tok free", "tictok"
    ],
    
    "irrelevant_intent_terms": [
        "makeup tutorial", "makeup editor", "makeup games", "beauty salon",
        "acne remover", "teeth whitening", "hair color changer",
        "virtual makeup", "widget", "widgets", "emulator", "emulators",
        "retro games", "gba emulator", "nes emulator", "wallpaper pack",
        "keyboard themes", "launcher theme", "remote control", "tv remote",
        "smart remote", "gamepad", "controller", "hotspot"
    ],
    
    "risky_platform_terms": [
        "snapchat", "tiktok", "instagram", "facebook", "messenger", "whatsapp",
        "facetime", "iphone", "ios", "ipad", "apple", "android"
    ],
    
    "user_overrides": {
        "force_top30_terms": [],
        "force_consider_terms": [],
        "force_drop_terms": []
    },
    
    "balanced_weights": {
        "VolumeN": 0.20,
        "DifficultyN": 0.15,
        "KEIN": 0.15,
        "RelevancyScore": 0.30,
        "CurrentRankN": 0.10,
        "ExpansionValue": 0.10
    }
}

# Add "doggy" (double g spelling variant) to base config
config["intent_core_terms"].extend(["doggy filter", "doggy filters", "ar doggy filter"])
config["style_terms"].append("doggy")

# Save English-only config terms before localization merge
_base_config_terms = {k: list(config.get(k, [])) for k in ['intent_core_words', 'intent_core_terms', 'feature_terms', 'style_terms', 'visual_terms', 'noise_terms']}

# Apply Market-Specific Localization to Config Dictionaries
market_lang = config["market"].split("_")[1].upper() if "_" in config["market"] else "EN"

localized_data = {
    "ES": {
        "intent_core_words": ["filtro", "filtros", "cámara", "camara", "efecto", "efectos", "lente", "lentes"],
        "intent_core_terms": ["filtro ar", "filtros ar", "camara ar", "cámara ar", "filtro de cara", "filtros de cara", "filtro de rostro", "filtros de rostro", "efeito ar", "efecto ar", "efectos ar", "filtro divertido", "filtros divertidos", "filtro gracioso", "filtros graciosos", "filtro de perro", "filtros de perro", "filtro de perrito", "filtros de perrito", "filtro meme", "filtros de memes", "filtro facial", "filtros faciales"],
        "feature_terms": ["personaje 3d", "personajes 3d", "personaje ar", "personajes ar", "video divertido", "videos divertidos", "video ar", "videos ar", "broma de filtro", "filtro de broma", "deformar cara", "cara de perro", "cara de perrito", "filtro feo", "filtro de llanto", "filtro calvo", "efecto calvo", "muñeco 3d", "muñeco animado", "avatar animado", "personaje animado", "personajes animados"],
        "style_terms": ["divertido", "gracioso", "perro", "perrito", "mascota", "mascotas", "broma", "bromas", "animado", "realidad aumentada", "virtual"],
        "visual_terms": ["foto", "fotos", "video", "videos", "cámara", "camara", "selfie", "selfies", "imagen", "imágenes"]
    },
    "PT": {
        "intent_core_words": ["filtro", "filtros", "câmera", "camera", "efeito", "efeitos", "lente", "lentes"],
        "intent_core_terms": ["filtro ar", "filtros ar", "camera ar", "câmera ar", "filtro de cara", "filtros de cara", "filtro de rosto", "filtros de rosto", "efeito ar", "efeitos ar", "filtro divertido", "filtros divertidos", "filtro engraçado", "filtros engraçados", "filtro de cachorro", "filtros de cachorro", "filtro de cão", "filtro de pet", "filtro meme", "filtros de memes", "filtro facial", "filtros faciais"],
        "feature_terms": ["personagem 3d", "personagens 3d", "personagem ar", "personagens ar", "video divertido", "videos divertidos", "video ar", "videos ar", "piada de filtro", "filtro de piada", "deformar rosto", "cara de cachorro", "filtro feio", "filtro de choro", "filtro careca", "efeito careca", "boneco 3d", "boneco animado", "avatar animado", "personagem animado", "personagens animados"],
        "style_terms": ["divertido", "engraçado", "cachorro", "cão", "pet", "pets", "piada", "piadas", "animado", "realidade aumentada", "virtual"],
        "visual_terms": ["foto", "fotos", "video", "videos", "câmera", "camera", "selfie", "selfies", "imagem", "imagens"]
    },
    "ID": {
        "intent_core_words": ["filter", "kamera", "efek", "lensa"],
        "intent_core_terms": ["filter ar", "kamera ar", "efek ar", "filter wajah", "efek wajah", "filter lucu", "efek lucu", "filter meme", "filter anjing", "filter 3d"],
        "feature_terms": ["karakter 3d", "karakter ar", "video lucu", "video ar", "video pendek", "filter nangis", "filter botak", "muka anjing", "muka jelek", "kamera lucu", "kamera meme"],
        "style_terms": ["fyp", "meme", "lucu", "tren tiktok", "anjing", "kartun", "anime", "avatar", "aneh"],
        "visual_terms": ["kamera", "video", "rekaman", "selfie", "foto", "gambar"]
    }
}

if market_lang in localized_data:
    for key, words in localized_data[market_lang].items():
        if key in config:
            config[key].extend(words)
            config[key] = list(set(config[key]))

# --- Google Play Scraper & Competitor Profile Builder ---
def get_app_profile(config, seed_query):
    # Check for custom User-supplied App Profile first
    script_dir = os.path.dirname(os.path.abspath(__file__))
    user_profile_path = os.path.join(script_dir, "App_Profile.json")
    if os.path.exists(user_profile_path):
        try:
            with open(user_profile_path, "r", encoding="utf-8") as f:
                u_profile = json.load(f)
            print(f"Loaded custom User App Profile from {user_profile_path}")
            
            # If Schema 2.0 format, adapt to version 1.0 format expected by the script
            if "schema_version" in u_profile or "competitor_strategy" in u_profile:
                print("Adapting Schema 2.0 Profile to pipeline compatibility format...")
                adapted = {
                    "app_id": u_profile.get("app_identity", {}).get("app_id", config["app_id"]),
                    "title": u_profile.get("app_identity", {}).get("title", ""),
                    "short_description": u_profile.get("live_store_metadata", {}).get("short_description", ""),
                    "full_description": u_profile.get("live_store_metadata", {}).get("full_description_digest", {}).get("one_sentence_summary", ""),
                    "competitors": [],
                    "last_checked": datetime.now().isoformat()
                }
                
                # Fetch competitors from competitor_strategy -> suggested_competitors
                sug_comps = u_profile.get("competitor_strategy", {}).get("suggested_competitors", [])
                for sc in sug_comps:
                    # Check for explicit short_description and desc200
                    s_desc = sc.get("short_description")
                    d_200 = sc.get("desc200")
                    
                    if not s_desc or not d_200:
                        # Fallback to old behavior
                        desc_text = " ".join(sc.get("overlap_keywords", []))
                        if sc.get("why_relevant"):
                            desc_text += " " + sc["why_relevant"]
                        if not s_desc:
                            s_desc = sc.get("why_relevant", "")
                        if not d_200:
                            d_200 = desc_text[:200]
                        
                    adapted["competitors"].append({
                        "package_id": sc.get("package_id", ""),
                        "title": sc.get("title", ""),
                        "short_description": s_desc,
                        "desc200": d_200
                    })
                u_profile = adapted
                
            return u_profile
        except Exception as e:
            print(f"Error reading custom user profile: {e}. Falling back to default path...")

    profile_path = os.path.join(os.path.dirname(OUTPUT_PATH), "App_Profile.json")
    
    # Check if cached profile is fresh (< 14 days)
    if os.path.exists(profile_path):
        try:
            with open(profile_path, "r", encoding="utf-8") as f:
                profile = json.load(f)
            last_checked = datetime.fromisoformat(profile.get("last_checked", "2000-01-01"))
            if datetime.now() - last_checked < timedelta(days=14):
                print(f"Loaded fresh App Profile from {profile_path} (Last checked: {last_checked})")
                return profile
        except Exception as e:
            print(f"Error reading profile cache: {e}. Re-fetching...")
            
    print("App Profile is missing or older than 14 days. Fetching from Google Play Store...")
    profile = {
        "app_id": config["app_id"],
        "last_checked": datetime.now().isoformat(),
        "title": "",
        "short_description": "",
        "full_description": "",
        "competitors": []
    }
    
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    
    def fetch_url(url):
        req = urllib.request.Request(
            url, 
            headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36'}
        )
        try:
            with urllib.request.urlopen(req, context=ctx, timeout=10) as response:
                return response.read().decode('utf-8')
        except Exception as e:
            print(f"Error fetching URL {url}: {e}")
            return None

    # 1. Fetch own app metadata
    own_url = f"https://play.google.com/store/apps/details?id={config['app_id']}&hl=en&gl=US"
    own_html = fetch_url(own_url)
    if own_html:
        t_match = re.search(r'<meta property="og:title" content="([^"]+)"', own_html)
        if t_match:
            title = html_lib.unescape(t_match.group(1))
            profile["title"] = re.sub(r'\s*-\s*Apps on Google Play$', '', title)
        s_match = re.search(r'<meta name="description" content="([^"]+)"', own_html)
        if not s_match:
            s_match = re.search(r'<meta name="twitter:description" content="([^"]+)"', own_html)
        if s_match:
            profile["short_description"] = html_lib.unescape(s_match.group(1))
        f_match = re.search(r'data-g-id="description"[^>]*>(.+?)</div>', own_html, re.DOTALL)
        if f_match:
            f_desc = re.sub(r'<[^>]+>', '\n', f_match.group(1))
            profile["full_description"] = html_lib.unescape(re.sub(r'\n+', '\n', f_desc).strip())
        else:
            j_match = re.search(r'\[\[null,"([^"]{50,})"\s*\]\]', own_html)
            if j_match:
                f_desc = j_match.group(1).replace('\\u003cbr\\u003e', '\n').replace('\\n', '\n')
                profile["full_description"] = html_lib.unescape(f_desc)
                
    # 2. Search for competitors
    search_url = f"https://play.google.com/store/search?q={urllib.parse.quote_plus(seed_query)}&c=apps&hl=en&gl=US"
    search_html = fetch_url(search_url)
    competitor_pids = []
    if search_html:
        raw_pids = re.findall(r'href="/store/apps/details\?id=([a-zA-Z0-9._]+)"', search_html)
        seen = set([config['app_id']])
        for pid in raw_pids:
            if pid not in seen:
                seen.add(pid)
                competitor_pids.append(pid)
                if len(competitor_pids) >= 3:
                    break
                    
    # 3. Fetch competitor details
    for pid in competitor_pids:
        comp_url = f"https://play.google.com/store/apps/details?id={pid}&hl=en&gl=US"
        comp_html = fetch_url(comp_url)
        comp_data = {
            "package_id": pid,
            "title": "",
            "short_description": "",
            "desc200": ""
        }
        if comp_html:
            t_match = re.search(r'<meta property="og:title" content="([^"]+)"', comp_html)
            if t_match:
                title = html_lib.unescape(t_match.group(1))
                comp_data["title"] = re.sub(r'\s*-\s*Apps on Google Play$', '', title)
            s_match = re.search(r'<meta name="description" content="([^"]+)"', comp_html)
            if not s_match:
                s_match = re.search(r'<meta name="twitter:description" content="([^"]+)"', comp_html)
            if s_match:
                comp_data["short_description"] = html_lib.unescape(s_match.group(1))
            f_match = re.search(r'data-g-id="description"[^>]*>(.+?)</div>', comp_html, re.DOTALL)
            f_desc_str = ""
            if f_match:
                f_desc_str = re.sub(r'<[^>]+>', '\n', f_match.group(1))
                f_desc_str = html_lib.unescape(re.sub(r'\n+', '\n', f_desc_str).strip())
            else:
                j_match = re.search(r'\[\[null,"([^"]{50,})"\s*\]\]', comp_html)
                if j_match:
                    f_desc_str = html_lib.unescape(j_match.group(1).replace('\\u003cbr\\u003e', '\n').replace('\\n', '\n'))
            comp_data["desc200"] = f_desc_str[:200]
        profile["competitors"].append(comp_data)
        
    try:
        os.makedirs(os.path.dirname(profile_path), exist_ok=True)
        with open(profile_path, "w", encoding="utf-8") as f:
            json.dump(profile, f, indent=4, ensure_ascii=False)
        print(f"Saved fresh App Profile to {profile_path}")
    except Exception as e:
        print(f"Error saving profile: {e}")
        
    return profile

# Build or load App Profile using seed query 'AR Filter'
app_profile = get_app_profile(config, "AR Filter")

# --- Local HTTP Server for Selection & ASO Dashboard ---
def start_interactive_server(df, config, app_profile):
    import http.server
    import socketserver
    import webbrowser
    import threading
    import time
    
    keywords_data = []
    for idx, row in df.iterrows():
        keywords_data.append({
            "Keyword": row['Keyword'],
            "Volume": int(row['Volume']),
            "Difficulty": int(row['Difficulty']),
            "KEI": float(row['KEI']),
            "RelevancyScore": float(row['RelevancyScore']),
            "BalancedScore": float(row['BalancedScore']),
            "Bucket": row.get('Bucket', 'Unclassified'),
            "Rank": str(row.get('Rank', '')),
            "CompetitorProven": row.get('CompetitorProven', 'No'),
            "ProvenDetails": row.get('ProvenDetails', '')
        })
        
    data_payload = {
        "app_name": config["app_name"],
        "app_id": config["app_id"],
        "category": config["category"],
        "competitors": app_profile.get("competitors", []),
        "keywords": keywords_data
    }
    
    server_data = {"confirmed_payload": None}
    
    class SelectionHandler(http.server.BaseHTTPRequestHandler):
        def log_message(self, format, *args):
            return
            
        def do_GET(self):
            if self.path == '/':
                self.send_response(200)
                self.send_header('Content-Type', 'text/html; charset=utf-8')
                self.end_headers()
                html_path = os.path.join(os.path.dirname(__file__), 'interactive_optimizer.html')
                with open(html_path, 'r', encoding='utf-8') as f:
                    self.wfile.write(f.read().encode('utf-8'))
            elif self.path == '/data':
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps(data_payload, ensure_ascii=False).encode('utf-8'))
            else:
                self.send_response(404)
                self.end_headers()
                
        def do_POST(self):
            if self.path == '/confirm':
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                payload = json.loads(post_data.decode('utf-8'))
                server_data["confirmed_payload"] = payload
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(b'{"status": "ok"}')
                
                def shutdown_soon():
                    time.sleep(0.5)
                    server.shutdown()
                threading.Thread(target=shutdown_soon).start()
                
    port = 8000
    for p in range(8000, 8050):
        try:
            server = socketserver.TCPServer(("", p), SelectionHandler)
            port = p
            break
        except OSError:
            continue
            
    print(f"\n[INTERACTIVE SELECTOR] Starting local server at http://localhost:{port}...")
    print("Opening browser automatically... Please select keywords and write ASO descriptions in the dashboard, then click Confirm.")
    
    webbrowser.open(f"http://localhost:{port}/")
    server.serve_forever()
    server.server_close()
    print("[INTERACTIVE SELECTOR] Configuration received. Proceeding with Excel generation...")
    return server_data["confirmed_payload"]



# Load and Clean
print("[Step 1] Loading candidates...")
df_raw = pd.read_csv(INPUT_PATH, encoding="utf-8-sig")

df = pd.DataFrame()
raw_cols = ['Keyword', 'Volume', 'Difficulty', 'KEI', 'Rank', 'CurrentRank', 'RankStatus', 'MaximumReach']
for col in raw_cols:
    if col in df_raw.columns:
        df[col] = df_raw[col]
    elif col == 'CurrentRank' and 'Rank' in df_raw.columns:
        df['CurrentRank'] = df_raw['Rank']
    elif col == 'Rank' and 'CurrentRank' in df_raw.columns:
        df['Rank'] = df_raw['CurrentRank']
    elif col == 'RankStatus' and 'Rank Status' in df_raw.columns:
        df['RankStatus'] = df_raw['Rank Status']
    elif col == 'MaximumReach' and 'Maximum Reach' in df_raw.columns:
        df['MaximumReach'] = df_raw['Maximum Reach']

if 'Volume' not in df.columns:
    df['Volume'] = 0
if 'Difficulty' not in df.columns:
    df['Difficulty'] = 0
if 'KEI' not in df.columns:
    df['KEI'] = 0
if 'Rank' not in df.columns:
    df['Rank'] = 'Unranked'

df['Volume'] = pd.to_numeric(df['Volume'], errors='coerce').fillna(0).astype(int)
df['Difficulty'] = pd.to_numeric(df['Difficulty'], errors='coerce').fillna(0).astype(int)
df['KEI'] = pd.to_numeric(df['KEI'], errors='coerce').fillna(0).astype(float)

# Load Max. Volume
max_vol_col = None
for col in df_raw.columns:
    if col.lower().replace('.', '').strip() == 'max volume':
        max_vol_col = col
        break

if max_vol_col is not None:
    df['Max. Volume'] = df_raw[max_vol_col]
else:
    df['Max. Volume'] = df['Volume']

df['Max. Volume'] = pd.to_numeric(df['Max. Volume'], errors='coerce').fillna(df['Volume']).astype(int)
df['Traffic Stability'] = (df['Volume'] / df['Max. Volume']).fillna(1.0).clip(0.0, 1.0)

def get_stability_class(ratio):
    if ratio >= 0.85:
        return 'Stable'
    elif ratio >= 0.50:
        return 'Moderate'
    else:
        return 'Volatile'

df['Stability Class'] = df['Traffic Stability'].apply(get_stability_class)
df['Rank_numeric'] = pd.to_numeric(df['Rank'], errors='coerce').fillna(999)

# Singularization / Normalization
def normalize_text(text):
    if not isinstance(text, str):
        return ""
    text = text.lower().strip()
    text = "".join(c for c in unicodedata.normalize('NFD', text) if unicodedata.category(c) != 'Mn')
    text = re.sub(r'[-_]', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()
    words = text.split()
    normalized_words = []
    for w in words:
        if w.endswith('s') and len(w) > 3:
            if w in ['themes', 'widgets', 'panels', 'settings', 'shortcuts', 'icons', 'styles', 'games', 'emulators', 'consoles']:
                w = w[:-1]
        normalized_words.append(w)
    return " ".join(normalized_words)

df['keyword_normalized'] = df['Keyword'].apply(normalize_text)

# Language classification
print("[Step 2] Language classification...")

try:
    from langdetect import detect_langs, DetectorFactory
    DetectorFactory.seed = 0
    HAS_LANGDETECT = True
except ImportError:
    HAS_LANGDETECT = False

def load_english_vocab():
    vocab = set()
    # Use relative path from project root for portability
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(script_dir)
    path = os.path.join(project_root, "Docs_and_Templates", "english_words_10k.txt")
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                for line in f:
                    vocab.add(line.strip().lower())
        except Exception:
            pass
    return vocab

english_vocab = load_english_vocab()

def get_root_word(w):
    w = w.lower()
    if len(w) > 4:
        if w.endswith('ing'):
            return w[:-3]
        if w.endswith('ed'):
            return w[:-2]
        if w.endswith('es'):
            return w[:-2]
        if w.endswith('er'):
            return w[:-2]
        if w.endswith('est'):
            return w[:-3]
    if len(w) > 3:
        if w.endswith('s'):
            return w[:-1]
    return w

# --- Dynamic Google Play Country-Language Mapping Loader ---
_COUNTRY_LANG_MAP = None

def _load_country_language_map():
    global _COUNTRY_LANG_MAP
    if _COUNTRY_LANG_MAP is not None:
        return _COUNTRY_LANG_MAP
    mapping = {}
    try:
        import openpyxl
        script_dir = os.path.dirname(os.path.abspath(__file__))
        map_path = os.path.join(script_dir, "google_play_country_language_map.xlsx")
        if not os.path.exists(map_path):
            map_path = os.path.join(os.path.dirname(script_dir), "google_play_country_language_map.xlsx")
        
        if os.path.exists(map_path):
            wb = openpyxl.load_workbook(map_path, read_only=True)
            if 'Country-Language Map' in wb.sheetnames:
                ws = wb['Country-Language Map']
                for row in ws.iter_rows(min_row=5, values_only=True):
                    if len(row) >= 7:
                        country = str(row[3] or '').strip().upper()
                        primary_locale = str(row[4] or '').strip()
                        secondary_locale = str(row[6] or '').strip()
                        if not country:
                            continue
                        
                        p_langs = [primary_locale.split('-')[0].lower()] if primary_locale else []
                        s_langs = []
                        if secondary_locale and secondary_locale.lower() != 'none':
                            for s in secondary_locale.split(';'):
                                s = s.strip()
                                if s:
                                    s_langs.append(s.split('-')[0].lower())
                        mapping[country] = {'primary': p_langs, 'secondary': s_langs}
            wb.close()
    except Exception as e:
        print(f"Warning loading country-language map: {e}")
    _COUNTRY_LANG_MAP = mapping
    return _COUNTRY_LANG_MAP

def _get_language_policy(config, primary_lang):
    """Get or auto-derive market language policy, incorporating the spreadsheet mapping."""
    policy = config.get('market_language_policy', {})
    
    # 1. If explicit policy exists in the configuration, use it
    if policy.get('primary_languages') or policy.get('secondary_languages'):
        policy_primary = [l.split('-')[0].lower() for l in policy.get('primary_languages', [])]
        secondary_langs = [l.split('-')[0].lower() for l in policy.get('secondary_languages', [])]
        if not policy_primary:
            policy_primary = [primary_lang]
        return policy_primary, secondary_langs
    
    # 2. Derive policy dynamically from the Country-Language map
    market = config.get('market', 'US_EN')
    if "_" in market:
        parts = market.split("_")
        country_code = parts[0].upper()
        target_lang = parts[1].lower()
    else:
        country_code = market.upper()
        target_lang = primary_lang
        
    cmap = _load_country_language_map()
    if cmap and country_code in cmap:
        sheet_primary = cmap[country_code]['primary']
        sheet_secondary = cmap[country_code]['secondary']
        
        # If the target language is part of the spreadsheet's primary languages
        if target_lang in sheet_primary:
            policy_primary = [target_lang]
            secondary_langs = [l for l in sheet_secondary if l != target_lang]
        else:
            policy_primary = [target_lang]
            # Include the spreadsheet's primary and other secondary languages as secondary for this run
            secondary_langs = [l for l in (sheet_primary + sheet_secondary) if l != target_lang]
            
        # Ensure secondary_langs is deduplicated and doesn't contain target_lang
        secondary_langs = list(dict.fromkeys(secondary_langs))
    else:
        # Fallback to standard derivation if spreadsheet is not loaded or country not found
        policy_primary = [target_lang]
        secondary_langs = ['en'] if target_lang != 'en' else []
        
    return policy_primary, secondary_langs

def _build_eng_words_only(base_terms):
    """Build English-only whitelist from config terms that were defined in English.
    Only uses the BASE config keys, not localized extensions."""
    eng_words = {
        'fyp', 'app', 'apps', 'free', 'download', 'android', 'new', 'best', 'top',
        'doggy', 'dogy', 'shrek', 'diy', 'pro', 'lite', 'tiktok', 'snapchat', 'instagram',
        'youtube', 'facebook', 'whatsapp', 'messenger', 'pinterest', 'google', 'play',
        '3d', 'arstudio', 'augmented', 'virtual', 'scanning', 'scanner', 'doge'
    }
    # Add words from configuration terms (these are typically English in the base config)
    for key in ['intent_core_words', 'intent_core_terms', 'feature_terms', 'style_terms', 'visual_terms', 'noise_terms']:
        if key in base_terms:
            for term in base_terms[key]:
                for w in str(term).lower().split():
                    # Skip words that look non-ASCII (likely localized terms)
                    if all(c.isascii() for c in w):
                        eng_words.add(w)
    return eng_words

# Pre-build the English whitelist once
_eng_words_cache = _build_eng_words_only(_base_config_terms)

# langdetect confusion matrix: known misclassifications for short text
# Maps (detected_lang) -> list of (likely_actual_lang) for correction
_LANGDETECT_CONFUSION = {
    'no': ['en'],        # Norwegian often = English
    'da': ['en'],        # Danish often = English
    'it': ['en', 'es'],  # Italian often = English or Spanish (short words)
    'ro': ['en'],        # Romanian often = English
    'sl': ['en'],        # Slovenian often = English
    'so': ['en'],        # Somali often = English
    'tl': ['es'],        # Tagalog often = Spanish
    'pt': ['es'],        # Portuguese often = Spanish (and vice versa)
    'id': ['en'],        # Indonesian often = English
    'tr': ['es'],        # Turkish often = Spanish for single words
    'af': ['en'],        # Afrikaans often = English
    'cy': ['en'],        # Welsh often = English
    'sw': ['en'],        # Swahili often = English
}

def detect_keyword_language(kw, market_lang, config):
    def lang_match(l1, l2):
        l1, l2 = str(l1).lower(), str(l2).lower()
        return l1 == l2 or (l1 == 'fil' and l2 == 'tl') or (l1 == 'tl' and l2 == 'fil')

    kw_lower = str(kw).lower().strip()
    primary_lang = market_lang.split("_")[1].lower() if "_" in market_lang else "en"
    
    if not kw_lower:
        return primary_lang, 'PRIMARY'
    
    policy_primary, secondary_langs = _get_language_policy(config, primary_lang)
    
    # Clean words in keyword
    words = [re.sub(r'[^a-z0-9]', '', w) for w in kw_lower.split()]
    words = [w for w in words if w]
    
    if not words:
        return primary_lang, 'PRIMARY'
    
    # --- Tier 1: Check if keyword is entirely English ---
    all_english = True
    for w in words:
        root = get_root_word(w)
        if w not in _eng_words_cache and w not in english_vocab and root not in _eng_words_cache and root not in english_vocab:
            all_english = False
            break
    
    if all_english:
        # Classify English based on market language policy
        if any(lang_match('en', p) for p in policy_primary):
            return 'en', 'PRIMARY'
        elif any(lang_match('en', s) for s in secondary_langs):
            return 'en', 'SECONDARY'
        else:
            return 'en', 'FOREIGN'
    
    # --- Tier 2: Use langdetect with guardrails ---
    if HAS_LANGDETECT:
        try:
            langs = detect_langs(kw_lower)
            best_lang = langs[0].lang
            prob = langs[0].prob
            
            # Apply confusion matrix corrections for short text
            word_count = len(words)
            confusion_corrected = False
            if word_count <= 3 and best_lang in _LANGDETECT_CONFUSION:
                likely_langs = _LANGDETECT_CONFUSION[best_lang]
                # If primary_lang is in the confusion list, prefer it
                if any(lang_match(primary_lang, l) for l in likely_langs):
                    best_lang = primary_lang
                    confusion_corrected = True
                # If any secondary lang is in the confusion list
                elif any(any(lang_match(s, l) for l in likely_langs) for s in secondary_langs):
                    for s in secondary_langs:
                        if any(lang_match(s, l) for l in likely_langs):
                            best_lang = s
                            confusion_corrected = True
                            break
                # If 'en' is in confusion list and primary is English
                elif any(lang_match('en', l) for l in likely_langs) and any(lang_match('en', p) for p in policy_primary):
                    best_lang = 'en'
                    confusion_corrected = True
                # For very short keywords (1-2 words), default to primary
                elif word_count <= 2:
                    best_lang = primary_lang
                    confusion_corrected = True
            
            # Classify the detected language
            if lang_match(best_lang, primary_lang) or any(lang_match(best_lang, p) for p in policy_primary):
                return best_lang, 'PRIMARY'
            
            # For confusion-corrected results, trust the correction directly
            if confusion_corrected:
                if any(lang_match(best_lang, s) for s in secondary_langs):
                    return best_lang, 'SECONDARY'
                return best_lang, 'FOREIGN'
            
            # For non-corrected results, require higher confidence for short keywords
            min_prob = 0.85 if word_count <= 2 else 0.7 if word_count <= 3 else 0.6
            if prob >= min_prob:
                if any(lang_match(best_lang, s) for s in secondary_langs):
                    return best_lang, 'SECONDARY'
                # Only mark as FOREIGN with sufficient confidence
                return best_lang, 'FOREIGN'
            
        except Exception:
            pass
    
    # --- Tier 3: Fallback to market primary language ---
    return primary_lang, 'PRIMARY'

# Populate language columns in df
detected_langs = []
lang_groups = []

for idx, row in df.iterrows():
    # If columns already exist in raw data, use them
    raw_lang = df_raw.loc[idx, 'DetectedLanguage'] if 'DetectedLanguage' in df_raw.columns else None
    raw_group = df_raw.loc[idx, 'LanguageGroup'] if 'LanguageGroup' in df_raw.columns else None
    
    if pd.notna(raw_lang) and pd.notna(raw_group):
        detected_langs.append(raw_lang)
        lang_groups.append(raw_group)
    else:
        lang, group = detect_keyword_language(row['Keyword'], config.get('market', 'US_EN'), config)
        detected_langs.append(lang)
        lang_groups.append(group)

df['DetectedLanguage'] = detected_langs
df['LanguageGroup'] = lang_groups

# Translate non-English keywords to English
print("[Step 2.5] Translating non-English keywords to English...")
from concurrent.futures import ThreadPoolExecutor

def translate_to_english(keyword, lang):
    if str(lang).lower() == 'en':
        return keyword
    try:
        import urllib.request
        import urllib.parse
        import json
        q = urllib.parse.quote(str(keyword))
        url = f"https://translate.googleapis.com/translate_a/single?client=gtx&sl=auto&tl=en&dt=t&q={q}"
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        ctx = ssl._create_unverified_context() if hasattr(ssl, '_create_unverified_context') else None
        with urllib.request.urlopen(req, timeout=3, context=ctx) as resp:
            data = json.loads(resp.read().decode('utf-8'))
            return data[0][0][0]
    except Exception:
        return keyword

def translate_keywords_parallel(df_in):
    unique_translations = {}
    to_translate = []
    
    for idx, row in df_in.iterrows():
        kw = row['Keyword']
        lang = row['DetectedLanguage']
        if str(lang).lower() == 'en':
            unique_translations[kw] = kw
        else:
            unique_translations[kw] = None
            to_translate.append((kw, lang))
            
    to_translate_unique = list(set(to_translate))
    
    if to_translate_unique:
        def worker(item):
            kw, lang = item
            return kw, translate_to_english(kw, lang)
            
        with ThreadPoolExecutor(max_workers=20) as executor:
            results = executor.map(worker, to_translate_unique)
            for kw, translated in results:
                unique_translations[kw] = translated
                
    return df_in['Keyword'].map(unique_translations).fillna(df_in['Keyword']).tolist()

df['EN'] = translate_keywords_parallel(df)

# Hard filters
print("[Step 3] Hard filters...")
df['is_competitor'] = df['Keyword'].apply(
    lambda x: any(re.search(r'\b' + re.escape(brand.lower()) + r'\b', str(x).lower()) 
                  for brand in config['competitor_brands'])
)
df['is_typo'] = df['Keyword'].apply(
    lambda x: any(re.search(r'\b' + re.escape(typo.lower()) + r'\b', str(x).lower()) 
                  for typo in config['typo_blacklist'])
)
df['is_irrelevant'] = df['Keyword'].apply(
    lambda x: any(re.search(r'\b' + re.escape(term.lower()) + r'\b', str(x).lower()) 
                  for term in config['irrelevant_intent_terms'])
)

def is_noise_only(kw, config):
    kw_lower = str(kw).lower().strip()
    has_core = any(term.lower() in kw_lower for term in config['intent_core_terms'])
    if has_core:
        return False
    words = kw_lower.split()
    if len(words) > 1:
        has_feat = any(f.lower() in kw_lower for f in config['feature_terms'])
        has_sty = any(s.lower() in kw_lower for s in config['style_terms'])
        if has_feat or has_sty:
            return False
    noise_words = set(t.lower() for t in config['noise_terms'])
    if all(w in noise_words for w in words):
        return True
    if len(words) == 1:
        w = words[0]
        if w in noise_words:
            return True
    return False

# Setup default noise if missing in script
if 'noise_terms' not in config:
    config['noise_terms'] = ['app', 'apps', 'free', 'download', 'android', 'for android', 'new', 'best', 'top', '2026', '2025']
df['is_noise'] = df['Keyword'].apply(lambda x: is_noise_only(x, config))

# Naturalness Filter
print("[Step 4] Naturalness checking...")
def check_naturalness(kw, config):
    kw_lower = str(kw).lower()
    words = kw_lower.split()
    if len(words) == 0:
        return 'UNNATURAL', 'Empty keyword'
    word_counts = {}
    for w in words:
        word_counts[w] = word_counts.get(w, 0) + 1
    max_repeat = max(word_counts.values()) if word_counts else 0
    ratio = max_repeat / len(words) if words else 0
    if ratio > 0.5 and len(words) > 2:
        return 'STUFFING', 'Too many repeated words'
    has_core = any(term.lower() in kw_lower for term in config['intent_core_terms'])
    if len(words) > 6:
        if has_core:
            pass
        else:
            return 'TOO_LONG', f'Keyword has too many words ({len(words)})'
    grammar_patterns = [
        r"\b(game game|widget widget|theme theme)\b",
        r"\b(what is|how to|why do|when is|where is)\b"
    ]
    for pat in grammar_patterns:
        if re.search(pat, kw_lower):
            return 'UNNATURAL', 'Fails structural validation'
    for char in kw_lower:
        if ord(char) > 127 and char not in 'áéíóúüñ¿¡íóú':
            return 'LANGUAGE_BLEED', 'Foreign script character detected'
    return 'OK', 'Natural enough for keyword research'

if 'NaturalnessFlag' in df_raw.columns:
    df['NaturalnessFlag'] = df_raw['NaturalnessFlag'].fillna('OK')
    df['NaturalnessReason'] = df_raw.get('NaturalnessReason', 'Natural enough for keyword research')
else:
    naturalness = df['Keyword'].apply(lambda x: check_naturalness(x, config))
    df['NaturalnessFlag'] = [n[0] for n in naturalness]
    df['NaturalnessReason'] = [n[1] for n in naturalness]

# Scoring Logic
print("[Step 5] Relevancy Scoring...")

# Pre-calculate competitor-proven keywords and score boost
def normalize_match_text(text):
    if not isinstance(text, str):
        return ""
    text = text.lower()
    text = re.sub(r'[^a-z0-9\s]', ' ', text)
    return re.sub(r'\s+', ' ', text).strip()

def check_keyword_in_text(kw_norm, text_norm):
    if not kw_norm or not text_norm:
        return False
    pattern = r'\b' + re.escape(kw_norm) + r'\b'
    return bool(re.search(pattern, text_norm))

competitor_proven_list = []
proven_details_list = []
competitor_boost_list = []

for idx, row in df.iterrows():
    kw_norm = normalize_match_text(row['Keyword'])
    
    comp_title_matches = []
    comp_short_matches = []
    comp_desc_matches = []
    
    for comp in app_profile.get("competitors", []):
        comp_title_norm = normalize_match_text(comp.get("title", ""))
        comp_short_norm = normalize_match_text(comp.get("short_description", ""))
        comp_desc_norm = normalize_match_text(comp.get("desc200", ""))
        comp_name = comp.get("title", comp.get("package_id", ""))
        
        if check_keyword_in_text(kw_norm, comp_title_norm):
            comp_title_matches.append(comp_name)
        if check_keyword_in_text(kw_norm, comp_short_norm):
            comp_short_matches.append(comp_name)
        if check_keyword_in_text(kw_norm, comp_desc_norm):
            comp_desc_matches.append(comp_name)
            
    boost = 0.0
    details = []
    if comp_title_matches:
        boost += 0.12
        details.append(f"Title: {', '.join(comp_title_matches)}")
    if comp_short_matches:
        boost += 0.08
        details.append(f"ShortDesc: {', '.join(comp_short_matches)}")
    if comp_desc_matches:
        boost += 0.05
        details.append(f"Desc200: {', '.join(comp_desc_matches)}")
        
    boost = min(boost, 0.15)
    
    if boost > 0:
        competitor_proven_list.append("Yes")
        proven_details_list.append("; ".join(details))
        competitor_boost_list.append(boost)
    else:
        competitor_proven_list.append("No")
        proven_details_list.append("")
        competitor_boost_list.append(0.0)

df['CompetitorProven'] = competitor_proven_list
df['ProvenDetails'] = proven_details_list
df['CompetitorBoost'] = competitor_boost_list

def calculate_relevancy(row, config):
    kw = str(row['Keyword']).lower()
    score = 0.30 # baseline
    
    # Core intent or core words
    has_core_term = any(term in kw for term in config['intent_core_terms'])
    has_core_word = any(re.search(r'\b' + re.escape(w.lower()) + r'\b', kw) for w in config.get('intent_core_words', []))
    
    # Feature, Style, and Visual matches
    has_feature = any(re.search(r'\b' + re.escape(f.lower()) + r'\b', kw) for f in config['feature_terms'])
    has_style = any(re.search(r'\b' + re.escape(s.lower()) + r'\b', kw) for s in config['style_terms'])
    has_visual = any(re.search(r'\b' + re.escape(v.lower()) + r'\b', kw) for v in config.get('visual_terms', []))
    
    if has_core_term:
        score += 0.40
    elif has_core_word:
        # Check if accompanied by relevant context (feature, style, or visual)
        if has_feature or has_style or has_visual:
            score += 0.40
        else:
            score += 0.10  # Weak bonus for generic core words without context
        
    # Feature match
    if has_feature:
        score += 0.15
        
    # Style match
    if has_style:
        score += 0.10
        
    # Visual match
    if has_visual:
        score += 0.05
        
    # Penalties
    if row['is_competitor']:
        score -= 0.20
    if row['is_irrelevant']:
        score -= 0.25
    if row['LanguageGroup'] == 'FOREIGN':
        score -= 0.30
        
    # Competitor Boost
    score += row.get('CompetitorBoost', 0.0)
        
    return max(0.0, min(1.0, score))

if 'RelevancyScore' in df_raw.columns:
    df['RelevancyScore'] = df_raw['RelevancyScore'].fillna(0.3).astype(float) + df['CompetitorBoost']
    df['RelevancyScore'] = df['RelevancyScore'].clip(0.0, 1.0)
else:
    df['RelevancyScore'] = df.apply(lambda r: calculate_relevancy(r, config), axis=1)

# Normalization & Balanced Score
print("[Step 6] Balanced Score Normalization...")
max_vol = df['Max. Volume'].max()
max_kei = df['KEI'].max()

df['VolumeN'] = np.log1p(df['Max. Volume']) / np.log1p(max_vol) if max_vol > 0 else 0
df['DifficultyN'] = 1.0 - (df['Difficulty'].clip(0, 100) / 100.0)
df['KEIN'] = np.log1p(df['KEI']) / np.log1p(max_kei) if max_kei > 0 else 0

def get_rank_n(rank_val):
    try:
        if pd.isna(rank_val) or rank_val in ['Unranked', 'unranked', 'empty', '']:
            return 0.0
        r = float(rank_val)
        if r <= 10: return 1.0
        elif r <= 30: return 0.8
        elif r <= 50: return 0.6
        elif r <= 100: return 0.3
        else: return 0.1
    except:
        return 0.0
df['CurrentRankN'] = df['Rank'].apply(get_rank_n)

def calculate_expansion(row, config):
    kw = str(row['Keyword']).lower()
    words = kw.split()
    n = len(words)
    if n == 1:
        score = 0.9
    elif n == 2:
        score = 0.7
    elif n == 3:
        score = 0.5
    else:
        score = 0.3
        
    if 'widget' in kw or 'control' in kw:
        score += 0.1
    if row['is_competitor']:
        score = 0.1
    return max(0.0, min(1.0, score))

df['ExpansionValue'] = df.apply(lambda r: calculate_expansion(r, config), axis=1)

bw = config['balanced_weights']
df['BalancedScore'] = (
    bw['VolumeN'] * df['VolumeN'] +
    bw['DifficultyN'] * df['DifficultyN'] +
    bw['KEIN'] * df['KEIN'] +
    bw['RelevancyScore'] * df['RelevancyScore'] +
    bw['CurrentRankN'] * df['CurrentRankN'] +
    bw['ExpansionValue'] * df['ExpansionValue']
)

# Add a small bonus based on LanguageGroup: +0.02 for PRIMARY, +0.01 for SECONDARY
def get_language_bonus(row):
    lg = str(row.get('LanguageGroup', 'PRIMARY')).upper()
    if lg == 'PRIMARY':
        return 0.02
    elif lg == 'SECONDARY':
        return 0.01
    return 0.0

df['BalancedScore'] = (df['BalancedScore'] + df.apply(get_language_bonus, axis=1)).round(4)
df['RelevancyScore'] = df['RelevancyScore'].round(4)

# Bucket classification
print("[Step 7] Bucket classification...")
def classify_keyword(row, config):
    kw = str(row['Keyword']).lower()
    
    # Hard drops
    if row['is_competitor']:
        return 'Dropped', 'competitor_brand', 'Dropped: Competitor brand'
    if row['is_typo']:
        return 'Dropped', 'typo_truncated_broken', 'Dropped: Typo, truncated, or broken'
    if row['is_irrelevant']:
        return 'Dropped', 'irrelevant_intent', 'Dropped: Irrelevant category/intent'
    if row['is_noise']:
        return 'Dropped', 'noise_only', 'Dropped: Noise-only generic term'
    if row['NaturalnessFlag'] != 'OK':
        return 'Dropped', 'unnatural', f"Dropped: Unnatural phrase ({row['NaturalnessReason']})"
        
    # Language Mismatches
    if row['LanguageGroup'] == 'FOREIGN':
        return 'Language Mismatch Audit', 'foreign_language_mismatch', 'Foreign language mismatch'
    if row['LanguageGroup'] in ['MIXED', 'UNKNOWN']:
        return 'Manual Review', 'manual_review', 'Mixed or unknown language'
    if row['LanguageGroup'] == 'SECONDARY':
        return 'Consider Keywords', 'secondary_language_handling', 'Secondary language handling'
        
    # Platform Risk
    has_platform_risk = any(re.search(r'\b' + re.escape(term) + r'\b', kw) for term in config['risky_platform_terms'])
    if has_platform_risk:
        return 'Consider Keywords', 'platform_style_risk', 'Platform-style risk'
        
    # Core, Feature, Style
    has_core = any(term in kw for term in config['intent_core_terms'])
    has_feature = any(re.search(r'\b' + re.escape(f.lower()) + r'\b', kw) for f in config['feature_terms'])
    has_style = any(re.search(r'\b' + re.escape(s.lower()) + r'\b', kw) for s in config['style_terms'])
    has_core_word = any(re.search(r'\b' + re.escape(w.lower()) + r'\b', kw) for w in config.get('intent_core_words', []))
    
    if has_core:
        return 'Core Intent Final', 'core_intent_final', 'Strong core camera filter/effect search intent'
        
    # Check style-only held back
    if has_style and not has_core and not has_feature and not has_core_word:
        return 'Generic Style Reserve', 'style_only', 'Generic aesthetic/style-only terms held back from shortlist'
        
    if has_feature:
        return 'Effect / Filter Type', 'feature_keywords', 'Specific features/toggles candidate'
        
    if has_style:
        return 'User Intent / Content Use Case', 'style_keywords', 'Aesthetic/theme candidate'
        
    if row['RelevancyScore'] < 0.45:
        return 'Dropped', 'dropped', 'Dropped: Weak app intent after scoring'
        
    return 'Broad Expansion', 'broad_expansion', 'Broad camera filter expansion'

classifications = df.apply(lambda r: classify_keyword(r, config), axis=1)
df['Bucket'] = [c[0] for c in classifications]
df['DecisionRule'] = [c[1] for c in classifications]
df['Reason'] = [c[2] for c in classifications]

# Apply user overrides
def apply_user_overrides(row, config):
    kw = str(row['Keyword']).lower().strip()
    uo = config.get('user_overrides', {})
    
    if row['is_competitor'] or row['is_typo'] or row['LanguageGroup'] == 'FOREIGN':
        return row['Bucket'], row['DecisionRule'], row['Reason']
        
    force_drops = [t.lower().strip() for t in uo.get('force_drop_terms', [])]
    if kw in force_drops:
        return 'Dropped', 'user_override_force_drop', 'Dropped: Force drop by user override'
        
    force_top30 = [t.lower().strip() for t in uo.get('force_top30_terms', [])]
    if kw in force_top30:
        return 'Core Intent Final', 'user_override_force_top30', 'Core Intent Final: Forced by user override'
        
    force_consider = [t.lower().strip() for t in uo.get('force_consider_terms', [])]
    if kw in force_consider:
        return 'Consider Keywords', 'user_override_force_consider', 'Consider Keywords: Forced by user override'
        
    return row['Bucket'], row['DecisionRule'], row['Reason']

def override_row(row):
    bucket, rule, reason = apply_user_overrides(row, config)
    return pd.Series([bucket, rule, reason])

df[['Bucket', 'DecisionRule', 'Reason']] = df.apply(override_row, axis=1)

# Shortlist building & duplicate checking
print("[Step 8] Near-Duplicate Cleanup & Shortlist building...")
def build_shortlist(df_all, config):
    df_sorted = df_all.sort_values(by=['BalancedScore', 'Rank_numeric', 'KEI', 'Difficulty'], ascending=[False, True, False, True]).copy()
    selected_core, selected_broad, selected_consider = [], [], []
    selected_normalized, selected_tokens = set(), set()
    dedup_log = []
    
    def check_duplicate(kw, original_bucket):
        norm = normalize_text(kw)
        tokens = " ".join(sorted(norm.split()))
        if not norm:
            return True, "Empty normalized keyword", ""
        if norm in selected_normalized:
            all_selected = selected_core + selected_broad + selected_consider
            kept_kw = ""
            for item in all_selected:
                if normalize_text(item['Keyword']) == norm:
                    kept_kw = item['Keyword']
                    break
            return True, f"Exact normalized duplicate of '{kept_kw}'", kept_kw
        if tokens in selected_tokens:
            all_selected = selected_core + selected_broad + selected_consider
            kept_kw = ""
            for item in all_selected:
                t = " ".join(sorted(normalize_text(item['Keyword']).split()))
                if t == tokens:
                    kept_kw = item['Keyword']
                    break
            return True, f"Same normalized token set as '{kept_kw}'", kept_kw
        return False, "", ""
        
    def add_to_shortlist(item, section):
        norm = normalize_text(item['Keyword'])
        tokens = " ".join(sorted(norm.split()))
        selected_normalized.add(norm)
        selected_tokens.add(tokens)
        entry = item.to_dict()
        entry['Section'] = section
        entry['QuotaStatus'] = 'EXACT'
        entry['FillSource'] = ''
        entry['FillReason'] = ''
        return entry

    # Core (Quota: 25)
    core_candidates = df_sorted[df_sorted['Bucket'] == 'Core Intent Final']
    for _, row in core_candidates.iterrows():
        if len(selected_core) >= 25:
            break
        is_dup, reason, kept_kw = check_duplicate(row['Keyword'], 'Core Intent Final')
        if is_dup:
            dedup_log.append({
                'Table': '01_Main_Keyword_Shortlist', 'RemovedKeyword': row['Keyword'],
                'OriginalSection': 'Core Intent Final', 'KeptKeyword': kept_kw,
                'DedupReason': reason, 'BalancedScore': row['BalancedScore'],
                'Note': 'Keyword remains in All Candidates pool'
            })
        else:
            selected_core.append(add_to_shortlist(row, 'Core Intent Final'))
            
    # Core Fallback
    if len(selected_core) < 25:
        fallback_candidates = df_sorted[df_sorted['Bucket'].isin(['Effect / Filter Type', 'Broad Expansion'])]
        for _, row in fallback_candidates.iterrows():
            if len(selected_core) >= 25:
                break
            norm = normalize_text(row['Keyword'])
            if norm in selected_normalized:
                continue
            if row['RelevancyScore'] < 0.45:
                continue
            is_dup, reason, kept_kw = check_duplicate(row['Keyword'], 'Core Intent Final (Fallback)')
            if is_dup:
                dedup_log.append({
                    'Table': '01_Main_Keyword_Shortlist', 'RemovedKeyword': row['Keyword'],
                    'OriginalSection': 'Core Intent Final (Fallback)', 'KeptKeyword': kept_kw,
                    'DedupReason': reason, 'BalancedScore': row['BalancedScore'],
                    'Note': 'Keyword remains in All Candidates pool'
                })
            else:
                entry = add_to_shortlist(row, 'Core Intent Final')
                entry['QuotaStatus'] = 'FILLED'
                entry['FillSource'] = row['Bucket']
                entry['FillReason'] = 'Core Intent Quota Fallback'
                selected_core.append(entry)

    # Broad (Quota: 5)
    broad_candidates = df_sorted[df_sorted['Bucket'] == 'Broad Expansion']
    for _, row in broad_candidates.iterrows():
        if len(selected_broad) >= 5:
            break
        is_dup, reason, kept_kw = check_duplicate(row['Keyword'], 'Broad Expansion')
        if is_dup:
            dedup_log.append({
                'Table': '01_Main_Keyword_Shortlist', 'OriginalSection': 'Broad Expansion',
                'RemovedKeyword': row['Keyword'], 'KeptKeyword': kept_kw,
                'DedupReason': reason, 'BalancedScore': row['BalancedScore'],
                'Note': 'Keyword remains in All Candidates pool'
            })
        else:
            selected_broad.append(add_to_shortlist(row, 'Broad Expansion'))
            
    # Broad Fallback
    if len(selected_broad) < 5:
        fallback_candidates = df_sorted[df_sorted['Bucket'].isin(['Effect / Filter Type', 'User Intent / Content Use Case'])]
        for _, row in fallback_candidates.iterrows():
            if len(selected_broad) >= 5:
                break
            norm = normalize_text(row['Keyword'])
            if norm in selected_normalized:
                continue
            if row['RelevancyScore'] < 0.45:
                continue
            is_dup, reason, kept_kw = check_duplicate(row['Keyword'], 'Broad Expansion (Fallback)')
            if is_dup:
                dedup_log.append({
                    'Table': '01_Main_Keyword_Shortlist', 'RemovedKeyword': row['Keyword'],
                    'OriginalSection': 'Broad Expansion (Fallback)', 'KeptKeyword': kept_kw,
                    'DedupReason': reason, 'BalancedScore': row['BalancedScore'],
                    'Note': 'Keyword remains in All Candidates pool'
                })
            else:
                entry = add_to_shortlist(row, 'Broad Expansion')
                entry['QuotaStatus'] = 'FILLED'
                entry['FillSource'] = row['Bucket']
                entry['FillReason'] = 'Broad Expansion Quota Fallback'
                selected_broad.append(entry)

    # Consider (Quota: 10)
    consider_candidates = df_sorted[df_sorted['Bucket'] == 'Consider Keywords']
    for _, row in consider_candidates.iterrows():
        if len(selected_consider) >= 10:
            break
        is_dup, reason, kept_kw = check_duplicate(row['Keyword'], 'Consider Keywords')
        if is_dup:
            dedup_log.append({
                'Table': '01_Main_Keyword_Shortlist', 'RemovedKeyword': row['Keyword'],
                'OriginalSection': 'Consider Keywords', 'KeptKeyword': kept_kw,
                'DedupReason': reason, 'BalancedScore': row['BalancedScore'],
                'Note': 'Keyword remains in All Candidates pool'
            })
        else:
            selected_consider.append(add_to_shortlist(row, 'Consider Keywords'))
            
    # Consider Fallback
    if len(selected_consider) < 10:
        selected_kws = {item['Keyword'].lower() for item in selected_core + selected_broad}
        missed_opps = df_sorted[df_sorted['Bucket'].isin(['Core Intent Final', 'Broad Expansion']) & 
                                (~df_sorted['Keyword'].str.lower().isin(selected_kws))]
        for _, row in missed_opps.iterrows():
            if len(selected_consider) >= 10:
                break
            norm = normalize_text(row['Keyword'])
            if norm in selected_normalized:
                continue
            if row['RelevancyScore'] < 0.45:
                continue
            is_dup, reason, kept_kw = check_duplicate(row['Keyword'], 'Consider Keywords (Fallback)')
            if is_dup:
                dedup_log.append({
                    'Table': '01_Main_Keyword_Shortlist', 'RemovedKeyword': row['Keyword'],
                    'OriginalSection': 'Consider Keywords (Fallback)', 'KeptKeyword': kept_kw,
                    'DedupReason': reason, 'BalancedScore': row['BalancedScore'],
                    'Note': 'Keyword remains in All Candidates pool'
                })
            else:
                entry = add_to_shortlist(row, 'Consider Keywords')
                entry['QuotaStatus'] = 'FILLED'
                entry['FillSource'] = row['Bucket']
                entry['FillReason'] = 'Consider Keywords Quota Fallback (Missed Opportunity)'
                selected_consider.append(entry)

    return selected_core, selected_broad, selected_consider, dedup_log

selected_core, selected_broad, selected_consider, dedup_log_list = build_shortlist(df, config)

def build_curated_sheet(df_all, bucket_name, sheet_name):
    df_sorted = df_all[df_all['Bucket'] == bucket_name].sort_values(by=['BalancedScore', 'Rank_numeric', 'KEI', 'Difficulty'], ascending=[False, True, False, True]).copy()
    selected = []
    selected_normalized, selected_tokens = set(), set()
    dedup_entries = []
    
    for _, row in df_sorted.iterrows():
        if len(selected) >= 30:
            break
        norm = normalize_text(row['Keyword'])
        tokens = " ".join(sorted(norm.split()))
        
        is_dup = False
        reason = ""
        kept_kw = ""
        
        if norm in selected_normalized:
            is_dup = True
            for item in selected:
                if normalize_text(item['Keyword']) == norm:
                    kept_kw = item['Keyword']
                    break
            reason = f"Exact normalized duplicate of '{kept_kw}'"
        elif tokens in selected_tokens:
            is_dup = True
            for item in selected:
                t = " ".join(sorted(normalize_text(item['Keyword']).split()))
                if t == tokens:
                    kept_kw = item['Keyword']
                    break
            reason = f"Same normalized token set as '{kept_kw}'"
            
        if is_dup:
            dedup_entries.append({
                'Table': sheet_name, 'RemovedKeyword': row['Keyword'],
                'OriginalSection': bucket_name, 'KeptKeyword': kept_kw,
                'DedupReason': reason, 'BalancedScore': row['BalancedScore'],
                'Note': 'Keyword remains in All Candidates pool'
            })
        else:
            selected_normalized.add(norm)
            selected_tokens.add(tokens)
            entry = row.to_dict()
            entry['Section'] = bucket_name
            entry['QuotaStatus'] = 'EXACT'
            entry['FillSource'] = ''
            entry['FillReason'] = ''
            selected.append(entry)
            
    return selected, dedup_entries

selected_feature, dedup_feat = build_curated_sheet(df, 'Effect / Filter Type', '02_Effect_Filter_Type')
selected_style, dedup_style = build_curated_sheet(df, 'User Intent / Content Use Case', '03_User_Intent_Content_UseCase')

dedup_log_list.extend(dedup_feat)
dedup_log_list.extend(dedup_style)
df_dedup_log = pd.DataFrame(dedup_log_list)

# Metadata assignment
print("[Step 9] Metadata slot assignment...")

# Start Interactive Selector Dashboard
selections_file = os.path.join(os.path.dirname(__file__), "selected_keywords.json")

confirmed_selection = None

if os.path.exists(selections_file):
    print(f"\n[Step 9] Found existing keyword selections in {selections_file}. Loading...")
    with open(selections_file, "r", encoding="utf-8") as f:
        confirmed_selection = json.load(f)
else:
    if args.interactive:
        confirmed_selection = start_interactive_server(df, config, app_profile)
        if confirmed_selection:
            with open(selections_file, "w", encoding="utf-8") as f:
                json.dump(confirmed_selection, f, indent=4, ensure_ascii=False)
                
            print("\n" + "="*50)
            print("[SELECTION_CONFIRMED] Keyword selections successfully saved!")
            print("Selections saved to:", selections_file)
            print("="*50 + "\n")
    else:
        print("\n[Step 9] Run in non-interactive/headless mode. Using defaults from shortlist logic.")

if confirmed_selection:
    user_core = confirmed_selection.get("core_keywords", [])
    user_secondary = confirmed_selection.get("secondary_keywords", [])
    user_feature = confirmed_selection.get("feature_keywords", [])
    user_style = confirmed_selection.get("style_keywords", [])
    
    df_lookup = df.set_index('Keyword')
    
    selected_core = []
    for kw in user_core:
        if kw in df_lookup.index:
            row = df_lookup.loc[kw]
            if isinstance(row, pd.DataFrame):
                row = row.iloc[0]
            entry = row.to_dict()
            entry['Keyword'] = kw
            entry['Section'] = 'Core Intent Final'
            entry['QuotaStatus'] = 'EXACT'
            selected_core.append(entry)
            
    selected_broad = []
    selected_consider = []
    for idx, kw in enumerate(user_secondary):
        if kw in df_lookup.index:
            row = df_lookup.loc[kw]
            if isinstance(row, pd.DataFrame):
                row = row.iloc[0]
            entry = row.to_dict()
            entry['Keyword'] = kw
            if idx < 5:
                entry['Section'] = 'Broad Expansion'
                selected_broad.append(entry)
            else:
                entry['Section'] = 'Consider Keywords'
                selected_consider.append(entry)
                
    selected_feature = []
    for kw in user_feature:
        if kw in df_lookup.index:
            row = df_lookup.loc[kw]
            if isinstance(row, pd.DataFrame):
                row = row.iloc[0]
            entry = row.to_dict()
            entry['Keyword'] = kw
            entry['Section'] = 'Effect / Filter Type'
            selected_feature.append(entry)
            
    selected_style = []
    for kw in user_style:
        if kw in df_lookup.index:
            row = df_lookup.loc[kw]
            if isinstance(row, pd.DataFrame):
                row = row.iloc[0]
            entry = row.to_dict()
            entry['Keyword'] = kw
            entry['Section'] = 'User Intent / Content Use Case'
            selected_style.append(entry)
            
    config["app_title_draft"] = confirmed_selection.get("title", "")
    config["short_desc_draft"] = confirmed_selection.get("short_description", "")
    config["full_desc_draft"] = confirmed_selection.get("full_description", "")

all_shortlist = selected_core + selected_broad + selected_consider
for idx, entry in enumerate(all_shortlist):
    sec = entry['Section']
    if sec == 'Core Intent Final':
        if idx < 2:
            entry['WhereToUse'] = '🏷️ Title'
        elif idx < 9:
            entry['WhereToUse'] = '📱 Short Description'
        else:
            entry['WhereToUse'] = '📄 Full Description'
    elif sec == 'Broad Expansion':
        entry['WhereToUse'] = '📄 Full Description'
    else:
        entry['WhereToUse'] = '🔍 Consider / Research Only'

df_shortlist = pd.DataFrame(all_shortlist)

# Stylization & Export
print("[Step 10] Exporting to stylized Excel Workbook...")
wb = Workbook()
wb.remove(wb.active)

def style_sheet(ws, title, is_report=False):
    ws.views.sheetView[0].showGridLines = True
    if not is_report and ws.max_row > 1:
        ws.freeze_panes = 'A2'
        
    navy_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    white_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin = Side(border_style="thin", color="D3D3D3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    if not is_report:
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = navy_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border
                
                col_name = ws.cell(row=1, column=col_idx).value
                if col_name in ['Volume', 'Max. Volume', 'Difficulty', 'Rank', 'MaximumReach']:
                    try: cell.value = int(float(cell.value))
                    except: pass
                elif col_name in ['BalancedScore', 'RelevancyScore', 'KEI', 'VolumeN', 'DifficultyN', 'KEIN', 'CurrentRankN', 'OpportunityRankGap', 'ExpansionValue', 'Traffic Stability']:
                    try:
                        cell.value = round(float(cell.value), 4)
                        if col_name == 'Traffic Stability':
                            cell.number_format = '0.00%'
                        else:
                            cell.number_format = '0.0000'
                    except: pass
                        
    for col in ws.columns:
        max_len = 0
        for cell in col:
            val_str = str(cell.value or '')
            if '\n' in val_str:
                val_str = max(val_str.split('\n'), key=len)
            max_len = max(max_len, len(val_str))
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
        # Hide Traffic Stability and Stability Class columns
        col_name = ws.cell(row=1, column=col[0].column).value
        if col_name in ['Traffic Stability', 'Stability Class']:
            ws.column_dimensions[col_letter].hidden = True

# --- 00_README_CONFIG ---
ws_readme = wb.create_sheet(title="00_README_CONFIG")
ws_readme.views.sheetView[0].showGridLines = True
ws_readme.cell(row=1, column=1, value="ASO Keyword Planner v3.5 - Configuration Summary").font = Font(size=14, bold=True)
configs = [
    ("Pipeline Version", "ASO Keyword Planner v3.5"),
    ("App Name", config["app_name"]),
    ("App ID", config["app_id"]),
    ("Category", config["category"]),
    ("Market", config["market"]),
    ("Platform", config["platform_mode"]),
    ("Run Date", pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"))
]
if "app_title_draft" in config:
    configs.append(("Custom Title Draft", config["app_title_draft"]))
if "short_desc_draft" in config:
    configs.append(("Custom Short Desc Draft", config["short_desc_draft"]))
if "full_desc_draft" in config:
    configs.append(("Custom Full Desc Draft", config["full_desc_draft"]))
for idx, (lbl, val) in enumerate(configs, 3):
    ws_readme.cell(row=idx, column=1, value=lbl).font = Font(bold=True)
    ws_readme.cell(row=idx, column=2, value=val)
    ws_readme.cell(row=idx, column=1).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws_readme.cell(row=idx, column=2).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
ws_readme.column_dimensions['A'].width = 25
ws_readme.column_dimensions['B'].width = 80

# --- 01_Main_Keyword_Shortlist ---
ws_shortlist = wb.create_sheet(title="01_Main_Keyword_Shortlist")
cols_shortlist = ['Keyword', 'EN', 'Volume', 'Max. Volume', 'Difficulty', 'KEI', 'Rank', 'BalancedScore', 'Traffic Stability', 'Stability Class', 'Section', 'RelevancyScore', 
                  'CompetitorProven', 'ProvenDetails', 'DetectedLanguage', 'LanguageGroup', 'NaturalnessFlag', 'WhereToUse', 'QuotaStatus', 'FillSource', 'FillReason', 'Reason']
for col_idx, col in enumerate(cols_shortlist, 1):
    ws_shortlist.cell(row=1, column=col_idx, value=col)
for row_idx, entry in enumerate(all_shortlist, 2):
    for col_idx, col in enumerate(cols_shortlist, 1):
        ws_shortlist.cell(row=row_idx, column=col_idx, value=entry.get(col, ''))
style_sheet(ws_shortlist, "01_Main_Keyword_Shortlist")

# --- 02_Effect_Filter_Type ---
ws_feature = wb.create_sheet(title="02_Effect_Filter_Type")
cols_curated = ['Keyword', 'EN', 'Volume', 'Max. Volume', 'Difficulty', 'KEI', 'Rank', 'BalancedScore', 'Traffic Stability', 'Stability Class', 'Section', 'RelevancyScore', 'Reason']
for col_idx, col in enumerate(cols_curated, 1):
    ws_feature.cell(row=1, column=col_idx, value=col)
for row_idx, entry in enumerate(selected_feature, 2):
    for col_idx, col in enumerate(cols_curated, 1):
        ws_feature.cell(row=row_idx, column=col_idx, value=entry.get(col, ''))
style_sheet(ws_feature, "02_Effect_Filter_Type")

# --- 03_User_Intent_Content_UseCase ---
ws_style = wb.create_sheet(title="03_User_Intent_Content_UseCase")
for col_idx, col in enumerate(cols_curated, 1):
    ws_style.cell(row=1, column=col_idx, value=col)
for row_idx, entry in enumerate(selected_style, 2):
    for col_idx, col in enumerate(cols_curated, 1):
        ws_style.cell(row=row_idx, column=col_idx, value=entry.get(col, ''))
style_sheet(ws_style, "03_User_Intent_Content_UseCase")

# --- 04_Dropped_Audit ---
ws_dropped = wb.create_sheet(title="04_Dropped_Audit")
df_dropped = df[df['Bucket'] == 'Dropped'].sort_values(by=['BalancedScore', 'Rank_numeric', 'KEI', 'Difficulty'], ascending=[False, True, False, True])
cols_audit = ['Keyword', 'EN', 'Volume', 'Max. Volume', 'Difficulty', 'KEI', 'Rank', 'BalancedScore', 'Traffic Stability', 'Stability Class', 'RelevancyScore', 'DecisionRule', 'Reason']
for col_idx, col in enumerate(cols_audit, 1):
    ws_dropped.cell(row=1, column=col_idx, value=col)
for row_idx, (_, row) in enumerate(df_dropped.iterrows(), 2):
    for col_idx, col in enumerate(cols_audit, 1):
        ws_dropped.cell(row=row_idx, column=col_idx, value=row.get(col, ''))
style_sheet(ws_dropped, "04_Dropped_Audit")

# --- 05_Report_Summary ---
ws_report = wb.create_sheet(title="05_Report_Summary")
ws_report.views.sheetView[0].showGridLines = True
ws_report.cell(row=1, column=1, value="ASO Keyword Planner v3.5 - Report Summary").font = Font(size=14, bold=True)
ws_report.cell(row=3, column=1, value="Metric Summary").font = Font(size=12, bold=True)
metrics = [
    ("Total Raw Keywords", len(df)),
    ("Unique Keywords", df['Keyword'].nunique()),
    ("Dropped Keywords", len(df_dropped)),
    ("Core Intent Selected", len(selected_core)),
    ("Broad Expansion Selected", len(selected_broad)),
    ("Consider Selected", len(selected_consider)),
    ("Effect / Filter Type Curated (02)", len(selected_feature)),
    ("User Intent / Content Use Case Curated (03)", len(selected_style)),
    ("Duplicates Filtered (Dedup Log)", len(df_dedup_log))
]
for idx, (lbl, val) in enumerate(metrics, 4):
    ws_report.cell(row=idx, column=1, value=lbl).font = Font(bold=True)
    ws_report.cell(row=idx, column=2, value=val)

ws_report.cell(row=15, column=1, value="Language Summary").font = Font(size=12, bold=True)
lang_counts = df['LanguageGroup'].value_counts()
for idx, (lang_g, count) in enumerate(lang_counts.items(), 17):
    ws_report.cell(row=idx, column=1, value=lang_g).font = Font(bold=True)
    ws_report.cell(row=idx, column=2, value=count)

ws_report.cell(row=24, column=1, value="Naturalness Summary").font = Font(size=12, bold=True)
nat_counts = df['NaturalnessFlag'].value_counts()
for idx, (flag, count) in enumerate(nat_counts.items(), 26):
    ws_report.cell(row=idx, column=1, value=flag).font = Font(bold=True)
    ws_report.cell(row=idx, column=2, value=count)

ws_report.cell(row=3, column=4, value="Sheet Index").font = Font(size=12, bold=True)
sheets_info = [
    ("00_README_CONFIG", "App configuration parameters and run metadata"),
    ("01_Main_Keyword_Shortlist", "Top 25 Core + 5 Broad + 10 Consider shortlist for metadata allocation"),
    ("02_Effect_Filter_Type", "Curated effect and filter type specific candidates (capped <= 30)"),
    ("03_User_Intent_Content_UseCase", "Curated user intent and content use case specific candidates (capped <= 30)"),
    ("04_Dropped_Audit", "Dropped keywords with detailed reasons"),
    ("05_Report_Summary", "Summary stats, language breakdowns, and sheet indices"),
    ("06_All_Candidates", "Full candidate pool with detailed score and policy values"),
    ("07_Language_Mismatch", "Audit sheet for keywords mismatching US_EN market language"),
    ("08_Generic_Style_Reserve", "Broad style-only keywords held back from metadata shortlist"),
    ("09_Manual_Review", "Audit sheet for keywords flagged with mixed or unknown languages"),
    ("10_Top_By_Score", "Candidates sorted by BalancedScore before diversity overlap filtering"),
    ("11_Secondary_Language", "Research candidates matching Spanish (Secondary Language)"),
    ("12_Text_Dedup_Log", "Log of text-level duplicates and variants pruned during optimization")
]
for idx, (title, purpose) in enumerate(sheets_info, 5):
    ws_report.cell(row=idx, column=4, value=title).font = Font(bold=True)
    ws_report.cell(row=idx, column=5, value=purpose)

thin_border = Border(left=Side(style='thin', color='C0C0C0'), right=Side(style='thin', color='C0C0C0'), 
                     top=Side(style='thin', color='C0C0C0'), bottom=Side(style='thin', color='C0C0C0'))
for r in range(4, 13):
    ws_report.cell(row=r, column=1).border = thin_border
    ws_report.cell(row=r, column=2).border = thin_border
for r in range(17, 17 + len(lang_counts)):
    ws_report.cell(row=r, column=1).border = thin_border
    ws_report.cell(row=r, column=2).border = thin_border
for r in range(26, 26 + len(nat_counts)):
    ws_report.cell(row=r, column=1).border = thin_border
    ws_report.cell(row=r, column=2).border = thin_border
for r in range(5, 5 + len(sheets_info)):
    ws_report.cell(row=r, column=4).border = thin_border
    ws_report.cell(row=r, column=5).border = thin_border

ws_report.column_dimensions['A'].width = 30
ws_report.column_dimensions['B'].width = 15
ws_report.column_dimensions['D'].width = 25
ws_report.column_dimensions['E'].width = 65

# --- 06_All_Candidates ---
ws_all = wb.create_sheet(title="06_All_Candidates")
cols_all = ['Keyword', 'EN', 'Volume', 'Max. Volume', 'Difficulty', 'KEI', 'Rank', 'BalancedScore', 'Traffic Stability', 'Stability Class', 'RelevancyScore', 'CompetitorProven', 'ProvenDetails', 'Bucket', 
            'DetectedLanguage', 'LanguageGroup', 'NaturalnessFlag', 'Reason']
for col_idx, col in enumerate(cols_all, 1):
    ws_all.cell(row=1, column=col_idx, value=col)
for row_idx, (_, row) in enumerate(df.sort_values(by=['BalancedScore', 'Rank_numeric', 'KEI', 'Difficulty'], ascending=[False, True, False, True]).iterrows(), 2):
    for col_idx, col in enumerate(cols_all, 1):
        ws_all.cell(row=row_idx, column=col_idx, value=row.get(col, ''))
style_sheet(ws_all, "06_All_Candidates")

# --- 07_Language_Mismatch ---
ws_lang_m = wb.create_sheet(title="07_Language_Mismatch")
df_lang_m = df[df['Bucket'] == 'Language Mismatch Audit'].sort_values(by=['BalancedScore', 'Rank_numeric', 'KEI', 'Difficulty'], ascending=[False, True, False, True])
for col_idx, col in enumerate(cols_curated, 1):
    ws_lang_m.cell(row=1, column=col_idx, value=col)
for row_idx, (_, row) in enumerate(df_lang_m.iterrows(), 2):
    for col_idx, col in enumerate(cols_curated, 1):
        ws_lang_m.cell(row=row_idx, column=col_idx, value=row.get(col, ''))
style_sheet(ws_lang_m, "07_Language_Mismatch")

# --- 08_Generic_Style_Reserve ---
ws_reserve = wb.create_sheet(title="08_Generic_Style_Reserve")
df_reserve = df[df['Bucket'] == 'Generic Style Reserve'].sort_values(by=['BalancedScore', 'Rank_numeric', 'KEI', 'Difficulty'], ascending=[False, True, False, True])
for col_idx, col in enumerate(cols_curated, 1):
    ws_reserve.cell(row=1, column=col_idx, value=col)
for row_idx, (_, row) in enumerate(df_reserve.iterrows(), 2):
    for col_idx, col in enumerate(cols_curated, 1):
        ws_reserve.cell(row=row_idx, column=col_idx, value=row.get(col, ''))
style_sheet(ws_reserve, "08_Generic_Style_Reserve")

# --- 09_Manual_Review ---
ws_mrev = wb.create_sheet(title="09_Manual_Review")
df_mrev = df[df['Bucket'] == 'Manual Review'].sort_values(by=['BalancedScore', 'Rank_numeric', 'KEI', 'Difficulty'], ascending=[False, True, False, True])
for col_idx, col in enumerate(cols_curated, 1):
    ws_mrev.cell(row=1, column=col_idx, value=col)
for row_idx, (_, row) in enumerate(df_mrev.iterrows(), 2):
    for col_idx, col in enumerate(cols_curated, 1):
        ws_mrev.cell(row=row_idx, column=col_idx, value=row.get(col, ''))
style_sheet(ws_mrev, "09_Manual_Review")

# --- 10_Top_By_Score ---
ws_tps = wb.create_sheet(title="10_Top_By_Score")
df_tps = df.sort_values(by=['BalancedScore', 'Rank_numeric', 'KEI', 'Difficulty'], ascending=[False, True, False, True]).head(50)
for col_idx, col in enumerate(cols_curated, 1):
    ws_tps.cell(row=1, column=col_idx, value=col)
for row_idx, (_, row) in enumerate(df_tps.iterrows(), 2):
    for col_idx, col in enumerate(cols_curated, 1):
        ws_tps.cell(row=row_idx, column=col_idx, value=row.get(col, ''))
style_sheet(ws_tps, "10_Top_By_Score")

# --- 11_Secondary_Language ---
ws_seclang = wb.create_sheet(title="11_Secondary_Language")
df_seclang = df[df['LanguageGroup'] == 'SECONDARY'].sort_values(by=['BalancedScore', 'Rank_numeric', 'KEI', 'Difficulty'], ascending=[False, True, False, True])
for col_idx, col in enumerate(cols_curated, 1):
    ws_seclang.cell(row=1, column=col_idx, value=col)
for row_idx, (_, row) in enumerate(df_seclang.iterrows(), 2):
    for col_idx, col in enumerate(cols_curated, 1):
        ws_seclang.cell(row=row_idx, column=col_idx, value=row.get(col, ''))
style_sheet(ws_seclang, "11_Secondary_Language")

# --- 12_Text_Dedup_Log ---
ws_dedup = wb.create_sheet(title="12_Text_Dedup_Log")
cols_dedup = ['Table', 'RemovedKeyword', 'OriginalSection', 'KeptKeyword', 'DedupReason', 'BalancedScore', 'Note']
for col_idx, col in enumerate(cols_dedup, 1):
    ws_dedup.cell(row=1, column=col_idx, value=col)
if not df_dedup_log.empty:
    for row_idx, (_, row) in enumerate(df_dedup_log.iterrows(), 2):
        for col_idx, col in enumerate(cols_dedup, 1):
            ws_dedup.cell(row=row_idx, column=col_idx, value=row.get(col, ''))
style_sheet(ws_dedup, "12_Text_Dedup_Log")

# Save
print(f"Saving stylized workbook to {OUTPUT_PATH}...")
try:
    wb.save(OUTPUT_PATH)
    print("Pipeline for AR Filter complete!")
except PermissionError:
    alt_path = OUTPUT_PATH.replace(".xlsx", "_temp.xlsx")
    print(f"WARNING: Permission denied to write to {OUTPUT_PATH} (file is likely open in another program).")
    print(f"Saving to fallback path: {alt_path}")
    wb.save(alt_path)
    print("Pipeline complete (saved to fallback path)!")
